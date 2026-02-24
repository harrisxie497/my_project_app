"""
完整的DELIVERY任务测试和验证流程
"""
import sys
import os
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

from app.core.database import SessionLocal
from app.services.delivery_processor import DeliveryProcessor
import openpyxl

print("=" * 80)
print("DELIVERY任务测试 - 验证佐川顧客コード和記事欄2修改")
print("=" * 80)

# 测试任务目录
task_dir = "storage/tasks/test_delivery_jkm_001"
result_file = os.path.join(task_dir, "result.xlsx")

# 删除备份文件
backup_file = result_file.replace('.xlsx', '_backup3.xlsx')
if os.path.exists(backup_file):
    os.remove(backup_file)

# 执行DELIVERY处理
print("\n执行DELIVERY处理...")

db = SessionLocal()

try:
    # 获取task_id
    from app.models.task import Task
    task = db.query(Task).filter(Task.file_type == 'DELIVERY').first()
    task_id = task.id if task else 'test_delivery_001'

    header_params = {
        'mawb_no': 'TEST-001',
        'flight_no': 'CA123',
        'arrival_date': '2026-02-10',
        'task_id': task_id
    }

    print(f"Header params: {header_params}")

    processor = DeliveryProcessor(
        task_dir=task_dir,
        db_session=db,
        file_type='DELIVERY',
        header_params=header_params
    )

    result = processor.process()
    print(f"✅ 处理完成")

    # 验证结果
    print("\n验证结果文件...")

    wb_result = openpyxl.load_workbook(result_file)
    ws_result = wb_result.active

    # 获取表头
    headers = [cell.value for cell in ws_result[1]]

    # 验证佐川顧客コード（固定）列
    customer_col_idx = None
    for i, header in enumerate(headers):
        if header == '佐川顧客コード（固定）':
            customer_col_idx = i + 1
            break

    if customer_col_idx:
        customer_value = ws_result.cell(row=2, column=customer_col_idx).value
        print(f"\n佐川顧客コード（固定）列（第{customer_col_idx}列）:")
        print(f"  第2行值: {customer_value}")
        if customer_value == '148202040055':
            print(f"  ✅ 固定值正确！")
        else:
            print(f"  ❌ 固定值错误，应该是 148202040055")

    # 验证記事欄2列
    kiji_col_idx = None
    for i, header in enumerate(headers):
        if header == '記事欄2':
            kiji_col_idx = i + 1
            break

    if kiji_col_idx:
        kiji_value = ws_result.cell(row=2, column=kiji_col_idx).value
        print(f"\n記事欄2列（第{kiji_col_idx}列）:")
        print(f"  第2行值: {kiji_value}")
        print(f"  task_id: {task_id}")
        if kiji_value == task_id:
            print(f"  ✅ task_id正确！")
        else:
            print(f"  ❌ task_id错误，应该是 {task_id}")

    wb_result.close()

except Exception as e:
    print(f"\n❌ 处理失败: {str(e)}")
    import traceback
    traceback.print_exc()
finally:
    db.close()

print("\n" + "=" * 80)
print("测试完成")
print("=" * 80)
print(f"\n结果文件: {result_file}")
print(f"请打开Excel文件手动验证所有列的数据！")
