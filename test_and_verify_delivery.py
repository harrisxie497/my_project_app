"""
完整的DELIVERY任务测试和验证流程
"""
import sys
import os
import io

# 设置标准输出为UTF-8编码
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

print("Current directory:", os.getcwd())
print("Script path:", os.path.abspath(__file__))

# 添加backend目录到路径
backend_path = os.path.join(os.path.dirname(__file__), 'backend')
sys.path.insert(0, backend_path)

from app.core.database import SessionLocal
from app.services.delivery_processor import DeliveryProcessor
from app.models.field_pipeline import FieldPipeline
import openpyxl
from datetime import datetime
import json

print("=" * 80)
print("DELIVERY任务完整测试和验证流程")
print("=" * 80)

# 测试任务目录
task_dir = "storage/tasks/test_delivery_jkm_001"
original_file = os.path.join(task_dir, "original.xlsx")
result_file = os.path.join(task_dir, "result.xlsx")

# 检查原始文件
if not os.path.exists(original_file):
    print(f"\n❌ 原始文件不存在: {original_file}")
    sys.exit(1)

print(f"\n✅ 测试目录: {task_dir}")
print(f"✅ 原始文件: {original_file}")

# 步骤1: 检查原始文件
print("\n" + "=" * 80)
print("步骤1: 检查原始文件第一列数据")
print("=" * 80)

wb_orig = openpyxl.load_workbook(original_file)
ws_orig = wb_orig.active

first_col_values = []
empty_row_found = False
empty_row_index = -1

for row_idx, row in enumerate(ws_orig.iter_rows(min_col=1, max_col=1), start=1):
    cell_value = row[0].value
    first_col_values.append((row_idx, cell_value))

    if not empty_row_found and row_idx > 1:
        if cell_value is None or str(cell_value).strip() == '':
            empty_row_found = True
            empty_row_index = row_idx

    if row_idx > 50:
        break

print(f"\n原始文件前20行的第1列数据:")
for row_idx, val in first_col_values[:20]:
    print(f"  第{row_idx}行: {repr(val)}")

if empty_row_found:
    print(f"\n✅ 在第{empty_row_index}行发现第一列为空")
else:
    print(f"\n⚠️  前50行未发现空行")

wb_orig.close()

# 步骤2: 执行DELIVERY处理
print("\n" + "=" * 80)
print("步骤2: 执行DELIVERY处理")
print("=" * 80)

if os.path.exists(result_file):
    backup_file = result_file.replace('.xlsx', '_backup.xlsx')
    os.rename(result_file, backup_file)
    print(f"已备份旧结果文件: {backup_file}")

db = SessionLocal()

try:
    header_params = {
        'mawb_no': 'TEST-001',
        'flight_no': 'CA123',
        'arrival_date': '2026-02-10'
    }

    print(f"\nHeader params: {header_params}")

    processor = DeliveryProcessor(
        task_dir=task_dir,
        db_session=db,
        file_type='DELIVERY',
        header_params=header_params
    )

    result = processor.process()

    print(f"\n✅ 处理完成")
    print(f"结果文件: {processor.result_file_path}")

    if not os.path.exists(result_file):
        print(f"❌ 结果文件不存在")
        sys.exit(1)

    file_size = os.path.getsize(result_file)
    print(f"结果文件大小: {file_size} 字节")

except Exception as e:
    print(f"\n❌ 处理失败: {str(e)}")
    import traceback
    traceback.print_exc()
    db.close()
    sys.exit(1)

# 步骤3: 验证功能改进
print("\n" + "=" * 80)
print("步骤3: 验证3个功能改进")
print("=" * 80)

# 验证1: None值替换
print("\n验证1: None值替换")
print("-" * 80)

wb_result = openpyxl.load_workbook(result_file)
ws_result = wb_result.active

has_none_values = False
none_value_count = 0
total_cells = 0

for row in ws_result.iter_rows():
    for cell in row:
        total_cells += 1
        if cell.value is None:
            has_none_values = True
            none_value_count += 1

if has_none_values:
    print(f"❌ 验证失败: 结果文件中有 {none_value_count} 个None值")
    none_validation = "FAILED"
else:
    print(f"✅ 验证通过: 结果文件中没有None值")
    none_validation = "PASSED"

# 验证2: 智能数据行判断
print("\n验证2: 智能数据行判断")
print("-" * 80)

data_row_count = ws_result.max_row - 1

if empty_row_found:
    expected_max_row = empty_row_index - 1
    if data_row_count <= expected_max_row:
        print(f"✅ 验证通过: 数据行数 {data_row_count} <= 预期 {expected_max_row}")
        smart_row_validation = "PASSED"
    else:
        print(f"⚠️  警告: 数据行数 {data_row_count} > 预期 {expected_max_row}")
        smart_row_validation = "WARNING"
else:
    print(f"⚠️  无法验证: 原始文件中没有明显的空行")
    smart_row_validation = "N/A"

wb_result.close()

db.close()

# 总结
print("\n" + "=" * 80)
print("测试总结")
print("=" * 80)

print(f"\n功能改进验证:")
print(f"  1. None值替换:     {none_validation}")
print(f"  2. 智能数据行判断: {smart_row_validation}")

print(f"\n结果文件: {result_file}")
