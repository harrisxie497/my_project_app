"""
最终测试：验证D列和記事欄2列的修复
"""
import sys
import os
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

from app.core.database import SessionLocal
from app.services.delivery_processor import DeliveryProcessor
from app.models.field_pipeline import FieldPipeline
import openpyxl

print("=" * 80)
print("DELIVERY任务最终测试")
print("=" * 80)

task_dir = "storage/tasks/test_delivery_jkm_001"
result_file = os.path.join(task_dir, "result.xlsx")

db = SessionLocal()

try:
    # 验证配置
    print("\n验证当前配置:")

    pipeline_kiji = db.query(FieldPipeline).filter(
        FieldPipeline.target_header == '記事欄2',
        FieldPipeline.file_type == 'DELIVERY'
    ).first()

    if pipeline_kiji:
        print(f"\n記事欄2列:")
        print(f"  map_op: {pipeline_kiji.map_op}")
        print(f"  field_type: {pipeline_kiji.field_type}")
        print(f"  rule_params_json: {pipeline_kiji.rule_params_json}")

    pipeline_d = db.query(FieldPipeline).filter(
        FieldPipeline.target_col == 'D',
        FieldPipeline.file_type == 'DELIVERY'
    ).first()

    if pipeline_d:
        print(f"\nD列（時間帯指定）:")
        print(f"  map_op: {pipeline_d.map_op}")
        print(f"  depends_on: {pipeline_d.depends_on}")

    # 执行处理
    header_params = {
        'mawb_no': 'TEST-001',
        'flight_no': 'CA123',
        'arrival_date': '2026-02-10'
    }

    print(f"\nHeader params: {header_params}")

    processor = DeliveryProcessor(
        task_dir=task_dir,
        db_session=db,
        file_type='DELIVERY',
        header_params=header_params
    )

    result = processor.process()
    print(f"\n✅ 处理完成")

    # 验证结果
    print("\n验证结果文件...")

    wb_result = openpyxl.load_workbook(result_file)
    ws_result = wb_result.active

    # 获取表头
    headers = [cell.value for cell in ws_result[1]]

    # 验证D列
    d_col_idx = None
    for i, header in enumerate(headers):
        if header == '時間帯指定':
            d_col_idx = i + 1
            break

    if d_col_idx:
        print(f"\nD列（時間帯指定）前20行数据:")
        d_data = []
        for row_idx in range(2, min(22, ws_result.max_row + 1)):
            d_value = ws_result.cell(row=row_idx, column=d_col_idx).value
            d_data.append(d_value)
            print(f"  第{row_idx}行: {repr(d_value)}")

        # 检查是否全部为空
        non_empty_count = sum(1 for v in d_data if v not in [None, ''])
        print(f"\n  非空值数量: {non_empty_count}/20")
        if non_empty_count > 0:
            print(f"  ✅ D列有数据！")
        else:
            print(f"  ❌ D列全部为空！")

    # 验证記事欄2列
    kiji_col_idx = None
    for i, header in enumerate(headers):
        if header == '記事欄2':
            kiji_col_idx = i + 1
            break

    if kiji_col_idx:
        print(f"\n記事欄2列（第{kiji_col_idx}列）:")
        kiji_value = ws_result.cell(row=2, column=kiji_col_idx).value
        print(f"  第2行值: {repr(kiji_value)}")
        print(f"  mawb_no: {header_params.get('mawb_no', '')}")

        # 期待应该是格式化后的mawb_no
        mawb_no = header_params.get('mawb_no', '')
        if len(mawb_no) >= 8:
            expected_value = mawb_no[:8] + ' ' + mawb_no[8:]
        else:
            expected_value = mawb_no

        print(f"  期望值: {repr(expected_value)}")
        if kiji_value == expected_value:
            print(f"  ✅ 值正确！")
        else:
            print(f"  ❌ 值不正确")

    wb_result.close()

except Exception as e:
    print(f"\n❌ 处理失败: {str(e)}")
    import traceback
    traceback.print_exc()
finally:
    db.close()

print("\n" + "=" * 80)
print("测试完成")
print("=" * 80)
print(f"\n结果文件: {result_file}")
print(f"请打开Excel文件手动验证！")
