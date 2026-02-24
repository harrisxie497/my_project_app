"""
检查原始文件中的D列数据
"""
import openpyxl

original_file = "storage/tasks/test_delivery_jkm_001/original.xlsx"

print(f"检查原始文件: {original_file}\n")

wb = openpyxl.load_workbook(original_file)
ws = wb.active

# 获取表头
headers = [cell.value for cell in ws[1]]
print(f"表头: {headers}\n")

# 找到D列（時間帯指定）和C列（配達指定日）
d_col_idx = None
c_col_idx = None

for i, header in enumerate(headers):
    if header == '時間帯指定':
        d_col_idx = i
    if header == '配達指定日':
        c_col_idx = i

print(f"D列索引: {d_col_idx}")
print(f"C列索引: {c_col_idx}\n")

# 读取D列和C列的前20行数据
print("D列和C列的前20行数据:")
print("-" * 80)

for row_idx in range(2, min(22, ws.max_row + 1)):
    d_value = ws.cell(row=row_idx, column=d_col_idx + 1).value
    c_value = ws.cell(row=row_idx, column=c_col_idx + 1).value
    print(f"第{row_idx}行: D列={repr(d_value)}, C列={repr(c_value)}")

wb.close()
