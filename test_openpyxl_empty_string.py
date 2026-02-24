"""
测试openpyxl写入空字符串的行为
"""
import openpyxl

# 创建一个新的工作簿
wb = openpyxl.Workbook()
ws = wb.active

# 写入不同类型的值
test_data = [
    ("正常字符串", "ABC"),
    ("空字符串", ""),
    ("None值", None),
    ("空字符串2", ""),
    ("数字", 123),
]

for i, (label, value) in enumerate(test_data, start=1):
    ws.cell(row=i, column=1, value=label)
    ws.cell(row=i, column=2, value=value)

# 保存文件
output_file = "test_empty_values.xlsx"
wb.save(output_file)
print(f"已创建测试文件: {output_file}")

# 重新读取并检查
wb2 = openpyxl.load_workbook(output_file)
ws2 = wb2.active

print("\n读取结果:")
for i in range(1, len(test_data) + 1):
    label = ws2.cell(row=i, column=1).value
    value = ws2.cell(row=i, column=2).value
    print(f"  {label}: {repr(value)} (类型: {type(value).__name__})")

wb2.close()
