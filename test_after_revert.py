"""
回退后的DELIVERY任务测试
"""
import sys
import os
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

from app.core.database import SessionLocal
from app.services.delivery_processor import DeliveryProcessor
from app.models.field_pipeline import FieldPipeline
import openpyxl

print("=" * 80)
print("DELIVERY任务测试 - 回退記事欄2配置后")
print("=" * 80)

# 测试任务目录
task_dir = "storage/tasks/test_delivery_jkm_001"
result_file = os.path.join(task_dir, "result.xlsx")

# 执行DELIVERY处理
print("\n执行DELIVERY处理...")

db = SessionLocal()

try:
    # 验证配置
    print("\n验证当前配置:")
    pipeline_customer = db.query(FieldPipeline).filter(
        FieldPipeline.target_header == '佐川顧客コード（固定）',
        FieldPipeline.file_type == 'DELIVERY'
    ).first()

    if pipeline_customer:
        rule_params = pipeline_customer.rule_params_json
        if isinstance(rule_params, dict) and 'policy_const' in rule_params:
            customer_code = rule_params['policy_const'].get('value', '')
            print(f"\n佐川顧客コード（固定）列:")
            print(f"  固定值: {customer_code}")
            if customer_code == '148202040055':
                print(f"  ✅ 配置正确！")
            else:
                print(f"  ❌ 配置错误，应该是 148202040055")

    pipeline_kiji = db.query(FieldPipeline).filter(
        FieldPipeline.target_header == '記事欄2',
        FieldPipeline.file_type == 'DELIVERY'
    ).first()

    if pipeline_kiji:
        print(f"\n記事欄2列:")
        print(f"  map_op: {pipeline_kiji.map_op}")
        print(f"  field_type: {pipeline_kiji.field_type}")
        print(f"  source_cols: {pipeline_kiji.source_cols}")
        if pipeline_kiji.map_op == 'COPY' and pipeline_kiji.field_type == 'TEXT':
            print(f"  ✅ 配置已回退到COPY方式！")
        else:
            print(f"  ❌ 配置不正确")

    # 执行处理
    header_params = {
        'mawb_no': 'TEST-001',
        'flight_no': 'CA123',
        'arrival_date': '2026-02-10'
    }

    print(f"\nHeader params: {header_params}")

    processor = DeliveryProcessor(
        task_dir=task_dir,
        db_session=db,
        file_type='DELIVERY',
        header_params=header_params
    )

    result = processor.process()
    print(f"\n✅ 处理完成")

    # 验证结果
    print("\n验证结果文件...")

    wb_result = openpyxl.load_workbook(result_file)
    ws_result = wb_result.active

    # 获取表头
    headers = [cell.value for cell in ws_result[1]]

    # 验证佐川顧客コード（固定）列
    customer_col_idx = None
    for i, header in enumerate(headers):
        if header == '佐川顧客コード（固定）':
            customer_col_idx = i + 1
            break

    if customer_col_idx:
        customer_value = ws_result.cell(row=2, column=customer_col_idx).value
        print(f"\n佐川顧客コード（固定）列（第{customer_col_idx}列）:")
        print(f"  第2行值: {customer_value}")
        if customer_value == '148202040055':
            print(f"  ✅ 固定值正确！")
        else:
            print(f"  ❌ 固定值错误，应该是 148202040055")

    # 验证記事欄2列
    kiji_col_idx = None
    for i, header in enumerate(headers):
        if header == '記事欄2':
            kiji_col_idx = i + 1
            break

    if kiji_col_idx:
        kiji_value = ws_result.cell(row=2, column=kiji_col_idx).value
        print(f"\n記事欄2列（第{kiji_col_idx}列）:")
        print(f"  第2行值: {kiji_value}")
        print(f"  说明: COPY方式，应该从原始文件复制")

    wb_result.close()

except Exception as e:
    print(f"\n❌ 处理失败: {str(e)}")
    import traceback
    traceback.print_exc()
finally:
    db.close()

print("\n" + "=" * 80)
print("测试完成")
print("=" * 80)
print(f"\n结果文件: {result_file}")
print(f"请打开Excel文件手动验证！")
