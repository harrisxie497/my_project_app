"""
测试脚本：使用process_header_row的返回值写入Excel
"""
import sys
import os
import logging

# 添加项目路径
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from app.services.excel_reader import read_excel_file
from app.services.excel_writer import write_excel_file_by_columns
from app.services.header_processor import process_header_row

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def main():
    """
    主函数：测试使用process_header_row返回值写入Excel
    """
    logger.info("=" * 100)
    logger.info("测试：使用process_header_row返回值写入Excel")
    logger.info("=" * 100)

    # 输入文件路径
    input_file = r"c:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_0a22941a\original.xlsx"
    # 输出文件路径
    output_file = r"c:\Users\harris.xie\Documents\trae_projects\japan\backend\test_output_with_header_row.xlsx"

    try:
        # 步骤1: 读取Excel文件
        logger.info("")
        logger.info("步骤1: 读取Excel文件")
        logger.info("=" * 100)

        result = read_excel_file(
            input_file,
            file_type='CUSTOMS',
            file_role='SOURCE'
        )

        column_data = result["column_data"]
        data_row_count = result["data_row_count"]

        logger.info(f"读取完成 - 列数: {len(column_data)}, 数据行数: {data_row_count}")

        # 提取表头
        headers = [col['head'] for col in column_data]
        logger.info(f"表头数量: {len(headers)}")

        # 构建按列数据字典
        column_data_dict = {}
        for col in column_data:
            col_name = col['head']
            col_values = col['data']
            column_data_dict[col_name] = col_values

        # 步骤2: 使用process_header_row生成特殊第一行
        logger.info("")
        logger.info("步骤2: 生成特殊第一行")
        logger.info("=" * 100)

        header_params = {
            "mawb_no": "16003279161",
            "flight_no": "CX509",
            "arrival_date": "20251210"
        }

        logger.info(f"header_params: {header_params}")

        # 生成特殊第一行（使用表头的列数）
        special_first_row = process_header_row(header_params, total_columns=len(headers))

        logger.info(f"特殊第一行生成完成:")
        logger.info(f"  长度: {len(special_first_row)}")
        logger.info(f"  前8列: {special_first_row[:8]}")
        logger.info(f"  非空列: {[i for i, v in enumerate(special_first_row) if v]}")

        # 步骤3: 写入Excel文件
        logger.info("")
        logger.info("步骤3: 写入Excel文件")
        logger.info("=" * 100)
        logger.info(f"输出文件: {output_file}")

        write_excel_file_by_columns(
            file_path=output_file,
            headers=headers,
            column_data=column_data_dict,
            special_first_row=special_first_row  # 使用process_header_row的返回值
        )

        # 步骤4: 验证写入结果
        logger.info("")
        logger.info("步骤4: 验证写入结果")
        logger.info("=" * 100)

        # 重新读取生成的文件
        from openpyxl import load_workbook

        workbook = load_workbook(output_file, data_only=True)
        worksheet = workbook.active

        # 读取第1行（特殊第一行）
        first_row_values = []
        for col_idx in range(1, min(11, len(headers) + 1)):  # 读取前10列
            cell_value = worksheet.cell(row=1, column=col_idx).value
            first_row_values.append(str(cell_value) if cell_value is not None else "")

        logger.info(f"验证 - 第1行（特殊第一行）前10列:")
        for idx, val in enumerate(first_row_values):
            col_letter = chr(65 + idx) if idx < 26 else f"{chr(65 + idx // 26 - 1)}{chr(65 + idx % 26)}"
            logger.info(f"  {col_letter}1: '{val}'")

        # 检查B1、E1、H1的值
        logger.info("")
        logger.info("验证关键字段:")
        b1_val = worksheet.cell(row=1, column=2).value
        e1_val = worksheet.cell(row=1, column=5).value
        h1_val = worksheet.cell(row=1, column=8).value

        logger.info(f"  B1: {b1_val}")
        logger.info(f"  E1: {e1_val}")
        logger.info(f"  H1: {h1_val}")

        # 验证期望值
        expected_b1 = "MAWB NO：16003279161"
        expected_e1 = "FLIGHT NO：CX509"
        expected_h1 = "ARRIVAL DATE：20251210"

        all_match = True
        if b1_val == expected_b1:
            logger.info("  ✓ B1值正确")
        else:
            logger.warning(f"  ✗ B1值错误: 期望='{expected_b1}', 实际='{b1_val}'")
            all_match = False

        if e1_val == expected_e1:
            logger.info("  ✓ E1值正确")
        else:
            logger.warning(f"  ✗ E1值错误: 期望='{expected_e1}', 实际='{e1_val}'")
            all_match = False

        if h1_val == expected_h1:
            logger.info("  ✓ H1值正确")
        else:
            logger.warning(f"  ✗ H1值错误: 期望='{expected_h1}', 实际='{h1_val}'")
            all_match = False

        # 检查第2行（表头行）
        logger.info("")
        logger.info("验证表头行（第2行）:")
        header_row_values = []
        for col_idx in range(1, min(6, len(headers) + 1)):  # 读取前5列
            cell_value = worksheet.cell(row=2, column=col_idx).value
            header_row_values.append(str(cell_value) if cell_value is not None else "")

        logger.info(f"  前5列: {header_row_values}")
        logger.info(f"  期望前5列: {headers[:5]}")

        if header_row_values == headers[:5]:
            logger.info("  ✓ 表头行正确")
        else:
            logger.warning("  ✗ 表头行不匹配")

        workbook.close()

        logger.info("")
        logger.info("=" * 100)
        if all_match:
            logger.info("✓ 所有验证通过!")
        else:
            logger.warning("⚠ 部分验证失败")
        logger.info(f"✓ 输出文件: {output_file}")
        logger.info("=" * 100)

    except Exception as e:
        logger.error(f"测试失败: {str(e)}", exc_info=True)
        raise


if __name__ == "__main__":
    main()
