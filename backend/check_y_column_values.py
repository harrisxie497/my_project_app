"""
检查结果文件中Y列（收件人地址）的值
"""

from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


def check_y_column_values():
    """检查结果文件中Y列（收件人地址）的值"""
    print("=" * 100)
    print("检查结果文件中Y列（收件人地址）的值")
    print("=" * 100)
    
    result_file_path = "C:\\Users\\harris.xie\\Documents\\trae_projects\\japan\\backend\\storage\\tasks\\t_e4caaadc\\result.xlsx"
    
    try:
        # 加载工作簿
        workbook = load_workbook(result_file_path)
        sheet = workbook.active
        
        print(f"\n结果文件路径: {result_file_path}")
        print(f"工作表名称: {sheet.title}")
        
        # 读取表头行
        header_row = []
        for cell in sheet[2]:
            header_row.append(cell.value)
        
        print(f"\n表头行数据: {header_row[:10]}")
        
        # 找到Y列的索引
        y_col_index = None
        for idx, header in enumerate(header_row):
            if header == '收件人地址':
                y_col_index = idx
                print(f"\n找到Y列，索引: {y_col_index}")
                break
        
        if y_col_index is not None:
            # 读取Y列的数据
            y_column_data = []
            for row_idx in range(3, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=y_col_index + 1).value
                y_column_data.append(cell_value)
            
            print(f"\nY列数据行数: {len(y_column_data)}")
            print(f"Y列前10个数据: {y_column_data[:10]}")
            print(f"Y列最后10个数据: {y_column_data[-10:]}")
        else:
            print(f"\n⚠️ 没有找到Y列")
        
        # 关闭工作簿
        workbook.close()
        
        print("\n" + "=" * 100)
        print("检查完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"\n❌ 检查失败: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    check_y_column_values()
