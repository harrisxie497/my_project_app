import openpyxl
import os

def check_source_columns():
    """检查原始文件中Q列和R列的数据"""
    print("=" * 100)
    print("检查原始文件中Q列和R列的数据")
    print("=" * 100)
    
    try:
        # 原始文件路径
        original_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\original.xlsx'
        
        # 读取原始文件
        wb = openpyxl.load_workbook(original_file_path)
        ws = wb.active
        
        # 显示表头行（第2行）
        print(f"\n表头行（第2行）:")
        for col_idx in range(1, min(20, ws.max_column + 1)):
            cell_value = ws.cell(row=2, column=col_idx).value
            print(f"  列{col_idx} ({chr(64 + col_idx)}): {cell_value}")
        
        # 显示Q列和R列的数据
        print(f"\nQ列（第17列）数据:")
        for row_idx in range(3, min(8, ws.max_row + 1)):
            cell_value = ws.cell(row=row_idx, column=17).value
            print(f"  行{row_idx}: {cell_value}")
        
        print(f"\nR列（第18列）数据:")
        for row_idx in range(3, min(8, ws.max_row + 1)):
            cell_value = ws.cell(row=row_idx, column=18).value
            print(f"  行{row_idx}: {cell_value}")
        
        # 显示L列和M列的数据
        print(f"\nL列（第12列）数据:")
        for row_idx in range(3, min(8, ws.max_row + 1)):
            cell_value = ws.cell(row=row_idx, column=12).value
            print(f"  行{row_idx}: {cell_value}")
        
        print(f"\nM列（第13列）数据:")
        for row_idx in range(3, min(8, ws.max_row + 1)):
            cell_value = ws.cell(row=row_idx, column=13).value
            print(f"  行{row_idx}: {cell_value}")
        
        wb.close()
        
        print("\n" + "=" * 100)
        print("检查完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_source_columns()
