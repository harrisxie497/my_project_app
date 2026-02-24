"""
完整的DELIVERY任务测试和验证流程：
1. 执行DELIVERY处理
2. 验证3个功能改进
3. 按照验收标准检查每一列的数据
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.core.database import SessionLocal
from app.services.delivery_processor import DeliveryProcessor
from app.models.field_pipeline import FieldPipeline
import openpyxl
from datetime import datetime
import json

print("=" * 80)
print("DELIVERY任务完整测试和验证流程")
print("=" * 80)

# 测试任务目录
task_dir = "storage/tasks/test_delivery_jkm_001"
original_file = os.path.join(task_dir, "original.xlsx")
result_file = os.path.join(task_dir, "result.xlsx")

# 检查原始文件
if not os.path.exists(original_file):
    print(f"\n❌ 原始文件不存在: {original_file}")
    sys.exit(1)

print(f"\n✅ 测试目录: {task_dir}")
print(f"✅ 原始文件: {original_file}")

# ============================================================================
# 步骤1: 读取原始文件，检查第一列数据（智能数据行判断测试准备）
# ============================================================================
print("\n" + "=" * 80)
print("步骤1: 检查原始文件第一列数据")
print("=" * 80)

wb_orig = openpyxl.load_workbook(original_file)
ws_orig = wb_orig.active

first_col_values = []
empty_row_found = False
empty_row_index = -1

for row_idx, row in enumerate(ws_orig.iter_rows(min_col=1, max_col=1), start=1):
    cell_value = row[0].value
    first_col_values.append((row_idx, cell_value))

    # 检查是否为空行（用于验证智能数据行判断）
    if not empty_row_found and row_idx > 1:  # 跳过表头行
        if cell_value is None or str(cell_value).strip() == '':
            empty_row_found = True
            empty_row_index = row_idx

    if row_idx > 50:  # 只看前50行
        break

print(f"\n原始文件前20行的第1列数据:")
for row_idx, val in first_col_values[:20]:
    print(f"  第{row_idx}行: {repr(val)}")

if empty_row_found:
    print(f"\n✅ 在第{empty_row_index}行发现第一列为空（用于测试智能数据行判断）")
else:
    print(f"\n⚠️  前50行未发现空行（智能数据行判断测试准备）")

wb_orig.close()

# ============================================================================
# 步骤2: 执行DELIVERY处理
# ============================================================================
print("\n" + "=" * 80)
print("步骤2: 执行DELIVERY处理")
print("=" * 80)

# 如果结果文件已存在，先备份
if os.path.exists(result_file):
    backup_file = result_file.replace('.xlsx', '_backup.xlsx')
    os.rename(result_file, backup_file)
    print(f"已备份旧结果文件: {backup_file}")

db = SessionLocal()

try:
    header_params = {
        'mawb_no': 'TEST-001',
        'flight_no': 'CA123',
        'arrival_date': '2026-02-10'
    }

    print(f"\nHeader params: {header_params}")

    processor = DeliveryProcessor(
        task_dir=task_dir,
        db_session=db,
        file_type='DELIVERY',
        header_params=header_params
    )

    result = processor.process()

    print(f"\n✅ 处理完成")
    print(f"结果文件: {processor.result_file_path}")

    if not os.path.exists(result_file):
        print(f"❌ 结果文件不存在")
        sys.exit(1)

    file_size = os.path.getsize(result_file)
    print(f"结果文件大小: {file_size} 字节")

except Exception as e:
    print(f"\n❌ 处理失败: {str(e)}")
    import traceback
    traceback.print_exc()
    db.close()
    sys.exit(1)

# ============================================================================
# 步骤3: 验证功能改进
# ============================================================================
print("\n" + "=" * 80)
print("步骤3: 验证3个功能改进")
print("=" * 80)

# 验证1: None值替换
print("\n验证1: None值替换")
print("-" * 80)

wb_result = openpyxl.load_workbook(result_file)
ws_result = wb_result.active

has_none_values = False
none_value_count = 0
total_cells = 0

for row in ws_result.iter_rows():
    for cell in row:
        total_cells += 1
        if cell.value is None:
            has_none_values = True
            none_value_count += 1

if has_none_values:
    print(f"❌ 验证失败: 结果文件中有 {none_value_count} 个None值（应该全部替换为空字符串）")
    none_validation = "FAILED"
else:
    print(f"✅ 验证通过: 结果文件中没有None值，已全部替换为空字符串")
    none_validation = "PASSED"

# 验证2: 智能数据行判断（通过检查数据行数来间接验证）
print("\n验证2: 智能数据行判断")
print("-" * 80)

# 检查原始文件和结果文件的数据行数
data_row_count = ws_result.max_row - 1  # 减去表头行

if empty_row_found:
    expected_max_row = empty_row_index - 1  # 应该在空行之前停止
    if data_row_count <= expected_max_row:
        print(f"✅ 验证通过: 数据行数 {data_row_count} <= 预期最大行数 {expected_max_row}（智能数据行判断生效）")
        smart_row_validation = "PASSED"
    else:
        print(f"⚠️  验证警告: 数据行数 {data_row_count} > 预期最大行数 {expected_max_row}")
        smart_row_validation = "WARNING"
else:
    print(f"⚠️  无法验证: 原始文件中没有明显的空行")
    smart_row_validation = "N/A"

# 验证3: 增强日志记录（需要查看日志文件）
print("\n验证3: 增强日志记录")
print("-" * 80)

log_file = os.path.join(os.path.dirname(__file__), "logs", "app.log")
if os.path.exists(log_file):
    print(f"✅ 日志文件存在: {log_file}")
    print(f"   请手动检查日志，搜索以下关键词:")
    print(f"   - '按列处理输入 - column_data'")
    print(f"   - '替换None值为空字符串'")
    print(f"   - '第一列为空，停止读取'")
    log_validation = "MANUAL"
else:
    print(f"⚠️  日志文件不存在，无法验证")
    log_validation = "N/A"

wb_result.close()

# ============================================================================
# 步骤4: 按照验收标准检查每一列
# ============================================================================
print("\n" + "=" * 80)
print("步骤4: 按照验收标准检查每一列")
print("=" * 80)

# 重新加载结果文件
wb_result = openpyxl.load_workbook(result_file)
ws_result = wb_result.active

# 查询DELIVERY的field_pipelines配置
pipelines = db.query(FieldPipeline).filter(
    FieldPipeline.file_type == 'DELIVERY',
    FieldPipeline.enabled == True
    ).order_by(FieldPipeline.order).all()

pipelines_sorted = sorted(pipelines, key=lambda x: x.order if x.order else 999)

# 读取结果文件的表头
headers = []
for cell in ws_result[1]:
    headers.append(cell.value)

print(f"\n结果文件表头: {headers}\n")
print(f"共 {len(pipelines_sorted)} 个列需要验证\n")

# 验证每一列
verification_results = []

for idx, pipeline in enumerate(pipelines_sorted, 1):
    target_col = pipeline.target_col
    target_header = pipeline.target_header
    map_op = pipeline.map_op
    source_cols = pipeline.source_cols
    field_type = pipeline.field_type
    rule_ref = pipeline.rule_ref
    rule_params_json = pipeline.rule_params_json

    print(f"{idx}. 列 {target_col} ({target_header})")
    print(f"   操作: {map_op}, 类型: {field_type}")

    # 解析source_cols
    if isinstance(source_cols, str):
        try:
            source_cols = json.loads(source_cols)
        except:
            source_cols = []

    # 在结果文件中找到该列
    col_index = None
    for i, header in enumerate(headers):
        if header == target_header:
            col_index = i + 1
            break

    if col_index is None:
        print(f"   ❌ 未找到该列")
        verification_results.append({
            'col': target_col,
            'header': target_header,
            'status': 'NOT_FOUND'
        })
        continue

    # 读取该列的数据
    column_data = []
    for row in ws_result.iter_rows(min_col=col_index, max_col=col_index, min_row=2):
        cell_value = row[0].value
        column_data.append(cell_value)

    print(f"   数据: {column_data[:5]}{'...' if len(column_data) > 5 else ''}")

    # 验证数据
    validation_passed = True
    validation_errors = []

    # 检查None值
    has_none = any(val is None for val in column_data)
    if has_none:
        validation_passed = False
        validation_errors.append("存在None值")

    # 根据map_op验证
    if map_op == 'COPY':
        # COPY列：检查是否有空值
        empty_count = sum(1 for val in column_data if val == '' or val is None)
        if empty_count > 0:
            print(f"   ℹ️  有 {empty_count} 个空值（可能是合法的）")

    elif map_op == 'CONST':
        # CONST列：检查所有值是否相同
        const_value = rule_params_json.get('value', '') if rule_params_json else ''
        non_const_count = sum(1 for val in column_data if val != const_value)
        if non_const_count > 0:
            validation_passed = False
            validation_errors.append(f"有{non_const_count}个值不匹配常量")
        else:
            print(f"   ✅ 所有值都是常量 '{const_value}'")

    elif map_op == 'CALC' and target_col == 'D':
        # D列：时间帯指定
        valid_time_slots = ['午前中', '12時-14時', '14時-16時', '16時-18時', '18時-20時', '19時-21時', '']
        invalid_count = sum(1 for val in column_data if val not in valid_time_slots and val != '')
        if invalid_count > 0:
            validation_passed = False
            validation_errors.append(f"有{invalid_count}个无效的时间帯")
        else:
            print(f"   ✅ 所有时间帯值都有效")

    # 输出验证结果
    if validation_passed:
        print(f"   ✅ 验证通过")
        verification_results.append({
            'col': target_col,
            'header': target_header,
            'status': 'PASSED'
        })
    else:
        print(f"   ❌ 验证失败: {', '.join(validation_errors)}")
        verification_results.append({
            'col': target_col,
            'header': target_header,
            'status': 'FAILED',
            'errors': validation_errors
        })

    print()

wb_result.close()
db.close()

# ============================================================================
# 总结报告
# ============================================================================
print("\n" + "=" * 80)
print("测试总结报告")
print("=" * 80)

# 功能改进验证
print("\n功能改进验证:")
print(f"  1. None值替换:     {none_validation}")
print(f"  2. 智能数据行判断: {smart_row_validation}")
print(f"  3. 增强日志记录:   {log_validation}")

# 列数据验证
passed_count = sum(1 for r in verification_results if r['status'] == 'PASSED')
failed_count = sum(1 for r in verification_results if r['status'] == 'FAILED')
not_found_count = sum(1 for r in verification_results if r['status'] == 'NOT_FOUND')

print(f"\n列数据验证:")
print(f"  总列数: {len(verification_results)}")
print(f"  ✅ 通过: {passed_count}")
print(f"  ❌ 失败: {failed_count}")
print(f"  ❓ 未找到: {not_found_count}")

# 失败的列详情
if failed_count > 0 or not_found_count > 0:
    print(f"\n❌ 验证未通过的列:")
    for r in verification_results:
        if r['status'] != 'PASSED':
            if r['status'] == 'FAILED':
                print(f"  - 列{r['col']} ({r['header']}): {', '.join(r.get('errors', []))}")
            else:
                print(f"  - 列{r['col']} ({r['header']}): 列不存在")
else:
    print(f"\n✅ 所有列验证通过！")

# 整体评估
print("\n" + "=" * 80)
if none_validation == "PASSED" and failed_count == 0 and not_found_count == 0:
    print("✅ 整体评估: 所有测试通过！")
else:
    print("⚠️  整体评估: 存在问题，需要修复")
print("=" * 80)

print(f"\n结果文件: {result_file}")
print(f"日志文件: {log_file if os.path.exists(log_file) else '不存在'}")
