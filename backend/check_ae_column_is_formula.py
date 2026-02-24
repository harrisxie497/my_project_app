"""
检查AE列的数据是否真的是公式
"""

from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


def check_ae_column_is_formula():
    """检查AE列的数据是否真的是公式"""
    print("=" * 100)
    print("检查AE列的数据是否真的是公式")
    print("=" * 100)
    
    original_file_path = "C:\\Users\\harris.xie\\Documents\\trae_projects\\japan\\backend\\storage\\tasks\\t_0fc5b76e\\original.xlsx"
    
    try:
        # 加载工作簿
        workbook = load_workbook(original_file_path)
        sheet = workbook.active
        
        print(f"\n原始文件路径: {original_file_path}")
        print(f"工作表名称: {sheet.title}")
        
        # 找到AE列的索引
        ae_col_index = None
        for idx, header in enumerate(sheet[1]):
            if header == '依赖人地址':
                ae_col_index = idx
                print(f"\n找到AE列，索引: {ae_col_index}")
                break
        
        if ae_col_index is not None:
            # 读取AE列的前10个单元格
            for row_idx in range(3, 13):
                cell = sheet.cell(row=row_idx, column=ae_col_index + 1)
                print(f"行{row_idx} - cell.value: {cell.value}, cell.internal_value: {cell.internal_value}, cell.data_type: {cell.data_type}")
        
        # 关闭工作簿
        workbook.close()
        
        print("\n" + "=" * 100)
        print("检查完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"\n❌ 检查失败: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    check_ae_column_is_formula()
