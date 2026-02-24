"""
验证生成的输出文件
"""
from app.services.excel_reader import read_excel_file
import openpyxl
import os

print("=" * 80)
print("验证DELIVERY输出文件")
print("=" * 80)

result_file = os.path.join(
    os.path.dirname(__file__),
    "test_results",
    "result.xlsx"
)

try:
    # 使用openpyxl读取
    wb = openpyxl.load_workbook(result_file)
    sheet = wb.active
    
    print(f"\n工作表名称: {sheet.title}")
    print(f"数据行数: {sheet.max_row}")
    print(f"数据列数: {sheet.max_column}")
    
    print("\n【表头行（第1行）】")
    header_row = []
    for col in range(1, min(18, sheet.max_column + 1)):
        cell_value = sheet.cell(row=1, column=col).value
        header_row.append(cell_value)
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"  {col_letter}: {cell_value}")
    
    print("\n【数据行检查（前2行）】")
    for row_idx in range(2, min(4, sheet.max_row + 1)):
        print(f"\n第{row_idx}行:")
        for col in range(1, min(18, sheet.max_column + 1)):
            cell_value = sheet.cell(row=row_idx, column=col).value
            col_letter = openpyxl.utils.get_column_letter(col)
            header = header_row[col-1] if col-1 < len(header_row) else ""
            print(f"  {col_letter} ({header}): {cell_value}")
    
    print("\n【关键字段验证】")
    print(f"N列（佐川顧客コード）第一行数据: {sheet.cell(row=2, column=14).value}")
    
    print("\n" + "=" * 80)
    print("验证完成！")
    print("=" * 80)
    
except Exception as e:
    print(f"错误: {str(e)}")
    import traceback
    traceback.print_exc()
