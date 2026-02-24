"""
查找最新的CUSTOMS任务并检查收件地址列
"""

import sys
import os

# 添加项目根目录到Python路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.core.database import SessionLocal
from app.models.task import Task
import openpyxl

def find_latest_customs_task():
    """查找最新的CUSTOMS任务并检查收件地址列"""
    print("=" * 100)
    print("查找最新的CUSTOMS任务并检查收件地址列")
    print("=" * 100)
    
    db_session = SessionLocal()
    
    try:
        # 查询最新的CUSTOMS类型任务
        task = db_session.query(Task).filter(
            Task.file_type == 'customs'
        ).order_by(Task.created_at.desc()).first()
        
        if task:
            print(f"\n最新CUSTOMS任务: {task.id}")
            
            # 从files字段获取任务目录
            files = task.files if task.files else {}
            task_dir = files.get('task_dir', '')
            print(f"  任务目录: {task_dir}")
            
            if not task_dir:
                print("\n任务目录为空")
                return
            
            original_file = os.path.join(task_dir, 'original.xlsx')
            result_file = os.path.join(task_dir, 'result.xlsx')
            
            # 读取原始文件
            original_workbook = openpyxl.load_workbook(original_file)
            original_sheet = original_workbook.active
            
            # 获取表头
            original_headers = []
            for cell in original_sheet[1]:
                original_headers.append(cell.value)
            
            print(f"\n原始文件表头: {original_headers}")
            
            # 查找收件人地址列
            if '收件人地址' in original_headers:
                col_index = original_headers.index('收件人地址') + 1
                print(f"\n收件人地址列（列{col_index}）:")
                
                # 获取数据
                original_data = []
                for row_idx in range(2, original_sheet.max_row + 1):
                    cell_value = original_sheet.cell(row=row_idx, column=col_index).value
                    original_data.append(cell_value)
                
                print(f"  原始数据数量: {len(original_data)}")
                print(f"  前10个数据: {original_data[:10]}")
            
            # 读取结果文件
            result_workbook = openpyxl.load_workbook(result_file)
            result_sheet = result_workbook.active
            
            # 获取表头
            result_headers = []
            for cell in result_sheet[1]:
                result_headers.append(cell.value)
            
            print(f"\n结果文件表头: {result_headers}")
            
            # 查找收件人地址列
            if '收件人地址' in result_headers:
                col_index = result_headers.index('收件人地址') + 1
                print(f"\n收件人地址列（列{col_index}）:")
                
                # 获取数据
                result_data = []
                for row_idx in range(2, result_sheet.max_row + 1):
                    cell_value = result_sheet.cell(row=row_idx, column=col_index).value
                    result_data.append(cell_value)
                
                print(f"  结果数据数量: {len(result_data)}")
                print(f"  前10个数据: {result_data[:10]}")
                
                # 检查是否有空值
                empty_indices = [idx for idx, v in enumerate(result_data) if v is None or (isinstance(v, str) and v.strip() == '')]
                if empty_indices:
                    print(f"\n空值位置（行号）: {[idx + 2 for idx in empty_indices]}")
                    print(f"  对应的原始数据: {[original_data[idx] for idx in empty_indices if idx < len(original_data)]}")
            else:
                print("\n结果文件中未找到收件人地址列")
        else:
            print("\n未找到CUSTOMS类型任务")
    
    finally:
        db_session.close()
    
    print("\n" + "=" * 100)
    print("检查完成")
    print("=" * 100)

if __name__ == "__main__":
    find_latest_customs_task()
