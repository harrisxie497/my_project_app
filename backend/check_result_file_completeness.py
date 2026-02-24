"""
检查结果文件的完整性
"""

from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


def check_result_file_completeness():
    """检查结果文件的完整性"""
    print("=" * 100)
    print("检查结果文件的完整性")
    print("=" * 100)
    
    result_file_path = "C:\\Users\\harris.xie\\Documents\\trae_projects\\japan\\backend\\storage\\tasks\\t_f37e2c5b\\result.xlsx"
    
    try:
        # 加载工作簿
        workbook = load_workbook(result_file_path)
        sheet = workbook.active
        
        print(f"\n结果文件路径: {result_file_path}")
        print(f"工作表名称: {sheet.title}")
        print(f"最大行数: {sheet.max_row}")
        print(f"最大列数: {sheet.max_column}")
        
        # 1. 检查第一行
        print("\n" + "=" * 100)
        print("1. 检查第一行")
        print("=" * 100)
        first_row = []
        for cell in sheet[1]:
            first_row.append(cell.value)
        
        print(f"第一行数据: {first_row[:10]}")
        
        # 检查第一行是否包含MAWB NO、FLIGHT NO、ARRIVAL DATE
        has_mawb = any('MAWB NO' in str(cell.value) for cell in sheet[1])
        has_flight = any('FLIGHT NO' in str(cell.value) for cell in sheet[1])
        has_arrival = any('ARRIVAL DATE' in str(cell.value) for cell in sheet[1])
        
        print(f"第一行包含MAWB NO: {has_mawb}")
        print(f"第一行包含FLIGHT NO: {has_flight}")
        print(f"第一行包含ARRIVAL DATE: {has_arrival}")
        
        # 2. 检查列的顺序是否正确
        print("\n" + "=" * 100)
        print("2. 检查列的顺序是否正确")
        print("=" * 100)
        
        # 读取表头行
        header_row = []
        for cell in sheet[2]:
            header_row.append(cell.value)
        
        print(f"表头行数据: {header_row[:10]}")
        
        # 检查列顺序是否正确
        expected_headers = ['会员编号', '序号', 'HAWB番号', '現地問合せ番号', '貨物個数', '貨物重量', '重量単位コード', '品名', '材质', '輸入者名', '輸入者住所', '輸入者 郵便番号', '輸入者電話番号', '輸出者名', '輸出者住所', 'インボイス価格条件コード', 'インボイス通貨コード', 'インボイス価格', '運賃区分コード', '運賃通貨コード', '運賃', '原産地コード', '備考', '收件人名（日文）', '收件人地址', '收件人电话', '收件人邮编', '依赖人名', '依赖人地址', '依赖人电话', '收件地址识别码', '电商货识别码', '电商平台码', '电商平台名称', '系统预留列，不可使用']
        
        is_headers_correct = header_row == expected_headers
        print(f"列顺序是否正确: {is_headers_correct}")
        
        # 3. 检查AI列的数据完整性
        print("\n" + "=" * 100)
        print("3. 检查AI列的数据完整性")
        print("=" * 100)
        
        # 检查Y列（收件人地址）的数据完整性
        y_col_index = None
        for idx, header in enumerate(header_row):
            if header == '收件人地址':
                y_col_index = idx
                break
        
        if y_col_index is not None:
            y_column_data = []
            for row_idx in range(3, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=y_col_index + 1).value
                y_column_data.append(cell_value)
            
            # 检查Y列的数据完整性
            y_non_empty_count = sum(1 for value in y_column_data if value is not None and value != '')
            y_total_count = len(y_column_data)
            y_completeness = y_non_empty_count / y_total_count if y_total_count > 0 else 0
            
            print(f"Y列（收件人地址）数据完整性: {y_completeness:.2%} ({y_non_empty_count}/{y_total_count})")
            print(f"Y列前10个数据: {y_column_data[:10]}")
        
        # 检查K列（輸入者住所）的数据完整性
        k_col_index = None
        for idx, header in enumerate(header_row):
            if header == '輸入者住所':
                k_col_index = idx
                break
        
        if k_col_index is not None:
            k_column_data = []
            for row_idx in range(3, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=k_col_index + 1).value
                k_column_data.append(cell_value)
            
            # 检查K列的数据完整性
            k_non_empty_count = sum(1 for value in k_column_data if value is not None and value != '')
            k_total_count = len(k_column_data)
            k_completeness = k_non_empty_count / k_total_count if k_total_count > 0 else 0
            
            print(f"K列（輸入者住所）数据完整性: {k_completeness:.2%} ({k_non_empty_count}/{k_total_count})")
            print(f"K列前10个数据: {k_column_data[:10]}")
        
        # 检查J列（輸入者名）的数据完整性
        j_col_index = None
        for idx, header in enumerate(header_row):
            if header == '輸入者名':
                j_col_index = idx
                break
        
        if j_col_index is not None:
            j_column_data = []
            for row_idx in range(3, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=j_col_index + 1).value
                j_column_data.append(cell_value)
            
            # 检查J列的数据完整性
            j_non_empty_count = sum(1 for value in j_column_data if value is not None and value != '')
            j_total_count = len(j_column_data)
            j_completeness = j_non_empty_count / j_total_count if j_total_count > 0 else 0
            
            print(f"J列（輸入者名）数据完整性: {j_completeness:.2%} ({j_non_empty_count}/{j_total_count})")
            print(f"J列前10个数据: {j_column_data[:10]}")
        
        # 4. 检查计算列的数据完整性
        print("\n" + "=" * 100)
        print("4. 检查计算列的数据完整性")
        print("=" * 100)
        
        # 检查F列（貨物重量）的数据完整性
        f_col_index = None
        for idx, header in enumerate(header_row):
            if header == '貨物重量':
                f_col_index = idx
                break
        
        if f_col_index is not None:
            f_column_data = []
            for row_idx in range(3, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=f_col_index + 1).value
                f_column_data.append(cell_value)
            
            # 检查F列的数据完整性
            f_non_empty_count = sum(1 for value in f_column_data if value is not None and value != '')
            f_total_count = len(f_column_data)
            f_completeness = f_non_empty_count / f_total_count if f_total_count > 0 else 0
            
            print(f"F列（貨物重量）数据完整性: {f_completeness:.2%} ({f_non_empty_count}/{f_total_count})")
            print(f"F列前10个数据: {f_column_data[:10]}")
        
        # 5. 检查每一列的数据完整性
        print("\n" + "=" * 100)
        print("5. 检查每一列的数据完整性")
        print("=" * 100)
        
        # 检查每一列的数据完整性
        for idx, header in enumerate(header_row):
            column_data = []
            for row_idx in range(3, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=idx + 1).value
                column_data.append(cell_value)
            
            # 检查列的数据完整性
            non_empty_count = sum(1 for value in column_data if value is not None and value != '')
            total_count = len(column_data)
            completeness = non_empty_count / total_count if total_count > 0 else 0
            
            print(f"{header}: {completeness:.2%} ({non_empty_count}/{total_count})")
        
        # 关闭工作簿
        workbook.close()
        
        print("\n" + "=" * 100)
        print("检查完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"\n❌ 检查失败: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    check_result_file_completeness()
