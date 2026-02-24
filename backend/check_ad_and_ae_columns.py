"""
检查AD列和AE列的数据
"""

from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


def check_ad_and_ae_columns():
    """检查AD列和AE列的数据"""
    print("=" * 100)
    print("检查AD列和AE列的数据")
    print("=" * 100)
    
    original_file_path = "C:\\Users\\harris.xie\\Documents\\trae_projects\\japan\\backend\\storage\\tasks\\t_0fc5b76e\\original.xlsx"
    
    try:
        # 加载工作簿
        workbook = load_workbook(original_file_path)
        sheet = workbook.active
        
        print(f"\n原始文件路径: {original_file_path}")
        print(f"工作表名称: {sheet.title}")
        
        # 读取表头行
        header_row = []
        for cell in sheet[2]:
            header_row.append(cell.value)
        
        print(f"\n表头行数据: {header_row}")
        
        # 找到AD列和AE列的索引
        ad_col_index = None
        ae_col_index = None
        for idx, header in enumerate(header_row):
            if header == '依赖人名':
                ad_col_index = idx
                print(f"\n找到AD列，索引: {ad_col_index}")
            if header == '依赖人地址':
                ae_col_index = idx
                print(f"找到AE列，索引: {ae_col_index}")
        
        if ad_col_index is not None:
            # 读取AD列的数据
            ad_column_data = []
            for row_idx in range(3, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=ad_col_index + 1).value
                ad_column_data.append(cell_value)
            
            print(f"\nAD列数据行数: {len(ad_column_data)}")
            print(f"AD列前5个数据: {ad_column_data[:5]}")
            print(f"AD列最后5个数据: {ad_column_data[-5:]}")
            
            # 统计None值的数量
            none_count = sum(1 for value in ad_column_data if value is None)
            print(f"AD列None值的数量: {none_count}")
            print(f"AD列非None值的数量: {len(ad_column_data) - none_count}")
        else:
            print(f"\n⚠️ 没有找到AD列")
        
        if ae_col_index is not None:
            # 读取AE列的数据
            ae_column_data = []
            for row_idx in range(3, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=ae_col_index + 1).value
                ae_column_data.append(cell_value)
            
            print(f"\nAE列数据行数: {len(ae_column_data)}")
            print(f"AE列前5个数据: {ae_column_data[:5]}")
            print(f"AE列最后5个数据: {ae_column_data[-5:]}")
            
            # 统计None值的数量
            none_count = sum(1 for value in ae_column_data if value is None)
            print(f"AE列None值的数量: {none_count}")
            print(f"AE列非None值的数量: {len(ae_column_data) - none_count}")
        else:
            print(f"\n⚠️ 没有找到AE列")
        
        # 关闭工作簿
        workbook.close()
        
        print("\n" + "=" * 100)
        print("检查完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"\n❌ 检查失败: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    check_ad_and_ae_columns()
