import openpyxl
import os

def check_k_and_n_columns():
    """检查原始文件中K列和N列的数据"""
    print("=" * 100)
    print("检查原始文件中K列和N列的数据")
    print("=" * 100)
    
    try:
        # 原始文件路径
        original_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\original.xlsx'
        
        # 读取原始文件
        wb = openpyxl.load_workbook(original_file_path, data_only=True)
        ws = wb.active
        
        print(f"\n工作表名称: {ws.title}")
        print(f"数据行数: {ws.max_row}")
        print(f"数据列数: {ws.max_column}")
        
        # 检查K列和N列的数据
        columns_to_check = {
            'K': '輸入者名',
            'N': '輸入者住所'
        }
        
        print(f"\n{'=' * 100}")
        print("K列和N列的数据（使用data_only=True）")
        print(f"{'=' * 100}")
        
        for col_letter, col_header in columns_to_check.items():
            col_idx = ord(col_letter) - ord('A') + 1
            
            print(f"\n{col_letter}列 ({col_header}):")
            print(f"  列索引: {col_idx}")
            
            # 显示前10行数据
            print(f"  前10行数据:")
            for row_idx in range(3, min(13, ws.max_row + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                print(f"    行{row_idx}: {cell_value}")
        
        wb.close()
        
        print("\n" + "=" * 100)
        print("检查完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_k_and_n_columns()
