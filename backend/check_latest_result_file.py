"""
检查最新任务的结果文件
"""

import sys
import os

# 添加项目根目录到Python路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook

def check_latest_result_file():
    """检查最新任务的结果文件"""
    print("=" * 100)
    print("检查最新任务的结果文件")
    print("=" * 100)
    
    # 读取结果文件
    file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_b6617ca6\result.xlsx'
    
    try:
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        
        # 获取表头行（第2行，索引为1）
        header_row = 2
        headers = {}
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=header_row, column=col).value
            if cell_value:
                headers[col] = str(cell_value)
        
        print(f"\n结果文件的表头顺序（第{header_row}行，共{len(headers)}列）:")
        for col, header in headers.items():
            print(f"  列{col}: {header}")
        
        # 检查数据行数
        data_start_row = 3
        print(f"\n数据行数: {worksheet.max_row - data_start_row + 1}")
        
        # 检查特定列的数据
        print(f"\n检查特定列的数据:")
        
        # 查找輸入者電話番号列
        for col, header in headers.items():
            if header == '輸入者電話番号':
                print(f"\n輸入者電話番号（列{col}）:")
                for row in range(data_start_row, worksheet.max_row + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    print(f"  第{row}行: {cell_value}")
                break
        
        # 查找收件人电话列
        for col, header in headers.items():
            if header == '收件人电话':
                print(f"\n收件人电话（列{col}）:")
                for row in range(data_start_row, worksheet.max_row + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    print(f"  第{row}行: {cell_value}")
                break
    
    except Exception as e:
        print(f"读取失败: {e}")
        import traceback
        traceback.print_exc()
    
    print("\n" + "=" * 100)
    print("检查完成")
    print("=" * 100)

if __name__ == "__main__":
    check_latest_result_file()
