import openpyxl
import os

def check_original_file_structure():
    """检查原始文件的详细结构"""
    print("=" * 100)
    print("检查原始文件的详细结构")
    print("=" * 100)
    
    try:
        # 原始文件路径
        original_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\original.xlsx'
        
        # 读取原始文件
        wb = openpyxl.load_workbook(original_file_path)
        ws = wb.active
        
        print(f"\n工作表名称: {ws.title}")
        print(f"数据行数: {ws.max_row}")
        print(f"数据列数: {ws.max_column}")
        
        # 显示前10行的完整数据
        print(f"\n前10行完整数据:")
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(36, ws.max_column + 1)):  # 只显示前35列
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(str(cell_value) if cell_value else '')
            print(f"\n行{row_idx}:")
            for col_idx, value in enumerate(row_data, 1):
                print(f"  列{col_idx}: {value}")
        
        wb.close()
        
        print("\n" + "=" * 100)
        print("检查完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_original_file_structure()
