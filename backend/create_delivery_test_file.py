"""创建DELIVERY测试Excel文件"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# 创建测试目录
test_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_results")
os.makedirs(test_dir, exist_ok=True)

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "Delivery"

# 设置表头（第1行）
headers = [
    "お客様管理番号",
    "佐川問合せ番号HAWB",
    "配達指定日",
    "時間帯指定",
    "貨物個数",
    "お届け先人名",
    "お届け先住所",
    "お届け先電話",
    "お届け先郵便",
    "依頼主",
    "依頼主住所",
    "依頼主郵便番号",
    "依頼主電話",
    "佐川顧客コード（固定）",
    "記事欄2（品名）",
    "記事欄2",
    "記事欄3",
    "收件人省州（删）",
    "收件人城市（删）",
    "收件人地址1（删）",
    "收件人地址2（删）",
    "発貨省（删）",
    "発貨市（删）",
    "発貨住所（删）"
]

# 写入表头
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# 写入测试数据（第2-5行）
test_data = [
    [
        "CUST001",
        "HAWB001",
        "2026-02-08",
        "12:00-14:00",
        2,
        "田中 太郎",
        "東京都渋谷区渋谷1-1-1",
        "03-1234-5678",
        "150-0001",
        "株式会社佐川急便",
        "東京都品川区品川1-1-1",
        "140-0001",
        "03-9876-5432",
        "12345",
        "衣類",
        "新品",
        "東京都",
        "渋谷区",
        "渋谷1-1-1",
        "",
        "東京都",
        "品川区",
        "品川1-1-1"
    ],
    [
        "CUST002",
        "HAWB002",
        "2026-02-08",
        "14:00-16:00",
        3,
        "山田 花子",
        "東京都港区赤坂1-1-1",
        "03-1111-2222",
        "107-0052",
        "株式会社ヤマト運輸",
        "東京都千代田区丸の内1-1-1",
        "100-0001",
        "03-3333-4444",
        "12345",
        "靴",
        "中古",
        "東京都",
        "港区",
        "赤坂1-1-1",
        "",
        "東京都",
        "千代田区",
        "丸の内1-1-1"
    ],
    [
        "CUST003",
        "HAWB003",
        "2026-02-09",
        "16:00-18:00",
        1,
        "佐藤 健一",
        "神奈川県横浜市中区本町1-1-1",
        "045-123-4567",
        "231-0021",
        "佐川急便株式会社",
        "東京都品川区品川2-2-2",
        "140-0002",
        "03-5555-6666",
        "12345",
        "バッグ",
        "新品",
        "神奈川県",
        "横浜市",
        "中区本町1-1-1",
        "",
        "東京都",
        "品川区",
        "品川2-2-2"
    ],
    [
        "CUST004",
        "HAWB004",
        "2026-02-09",
        "18:00-20:00",
        5,
        "鈴木 美咲",
        "大阪府大阪市中央区本町1-1-1",
        "06-1234-5678",
        "541-0041",
        "株式会社エイエイ",
        "東京都港区赤坂2-2-2",
        "107-0053",
        "03-7777-8888",
        "12345",
        "電子製品",
        "新品",
        "大阪府",
        "大阪市",
        "中央区本町1-1-1",
        "",
        "東京都",
        "港区",
        "赤坂2-2-2"
    ]
]

for row_idx, row_data in enumerate(test_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 保存文件
test_file = os.path.join(test_dir, "delivery_original.xlsx")
wb.save(test_file)
print(f"DELIVERY测试文件已创建: {test_file}")
