"""
检查源列AE（收件人地 址）的有效值数量
"""

from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


def check_ae_column_valid_values():
    """检查源列AE（收件人地 址）的有效值数量"""
    print("=" * 100)
    print("检查源列AE（收件人地 址）的有效值数量")
    print("=" * 100)
    
    original_file_path = "C:\\Users\\harris.xie\\Documents\\trae_projects\\japan\\backend\\storage\\tasks\\t_f37e2c5b\\original.xlsx"
    
    try:
        # 加载工作簿
        workbook = load_workbook(original_file_path)
        sheet = workbook.active
        
        print(f"\n原始文件路径: {original_file_path}")
        print(f"工作表名称: {sheet.title}")
        print(f"最大行数: {sheet.max_row}")
        print(f"最大列数: {sheet.max_column}")
        
        # 读取表头行
        header_row = []
        for cell in sheet[2]:
            header_row.append(cell.value)
        
        # 找到AE列（收件人地 址）的索引
        ae_col_index = None
        for idx, header in enumerate(header_row):
            if header and '收件人' in str(header) and '地址' in str(header):
                ae_col_index = idx
                print(f"\n找到AE列，索引: {ae_col_index}, 表头: {header}")
                break
        
        if ae_col_index is not None:
            # 读取AE列的数据
            ae_column_data = []
            for row_idx in range(3, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=ae_col_index + 1).value
                ae_column_data.append(cell_value)
            
            # 计算有效值的数量
            valid_values = [v for v in ae_column_data if v and str(v).strip() != '']
            print(f"\nAE列数据行数: {len(ae_column_data)}")
            print(f"AE列有效值数量: {len(valid_values)}")
            print(f"AE列空值数量: {len(ae_column_data) - len(valid_values)}")
            print(f"\nAE列前20个数据: {ae_column_data[:20]}")
        
        # 关闭工作簿
        workbook.close()
        
        print("\n" + "=" * 100)
        print("检查完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"\n❌ 检查失败: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    check_ae_column_valid_values()
