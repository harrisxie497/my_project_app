"""
检查原始文件中的DEFAULT类型列
"""

import openpyxl

def check_default_columns_in_original():
    """检查原始文件中的DEFAULT类型列"""
    print("=" * 100)
    print("检查原始文件中的DEFAULT类型列")
    print("=" * 100)
    
    original_file = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_fc2fe5d3\original.xlsx'
    
    # 读取原始文件
    workbook = openpyxl.load_workbook(original_file)
    sheet = workbook.active
    
    # 获取表头
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    
    print(f"\n表头: {headers}")
    
    # 查找DEFAULT类型的列（依頼主、依頼主住所、依頼主電話）
    default_columns = ['依頼主', '依頼主住所', '依頼主電話']
    
    for col_name in default_columns:
        if col_name in headers:
            col_index = headers.index(col_name) + 1
            print(f"\n{col_name}列（列{col_index}）:")
            
            # 获取前10行数据
            data = []
            for row_idx in range(2, min(12, sheet.max_row + 1)):
                cell_value = sheet.cell(row=row_idx, column=col_index).value
                data.append(cell_value)
            
            print(f"  前10行数据: {data}")
            
            # 检查是否有空值
            empty_count = sum(1 for v in data if v is None or (isinstance(v, str) and v.strip() == ''))
            print(f"  空值数量: {empty_count}")
        else:
            print(f"\n{col_name}列不存在")
    
    print("\n" + "=" * 100)
    print("检查完成")
    print("=" * 100)

if __name__ == "__main__":
    check_default_columns_in_original()
