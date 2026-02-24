"""
调试为什么X、Y、J、K列没有被处理
"""

import openpyxl
import pymysql

def debug_ai_columns_not_processed():
    """调试为什么X、Y、J、K列没有被处理"""
    print("=" * 100)
    print("调试为什么X、Y、J、K列没有被处理")
    print("=" * 100)
    
    # 原始文件路径
    original_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\original.xlsx'
    
    # 读取原始文件
    print(f"\n读取原始文件...")
    wb = openpyxl.load_workbook(original_file_path, data_only=True)
    ws = wb.active
    
    print(f"\n工作表名称: {ws.title}")
    print(f"数据行数: {ws.max_row}")
    print(f"数据列数: {ws.max_column}")
    
    # 检查AI列的source_cols数据
    ai_columns_config = [
        {'target_col': 'X', 'source_cols': ['AD'], 'depends_on': []},
        {'target_col': 'Y', 'source_cols': ['M'], 'depends_on': []},
        {'target_col': 'J', 'source_cols': ['K'], 'depends_on': ['X']},
        {'target_col': 'K', 'source_cols': ['N'], 'depends_on': ['Y']}
    ]
    
    print(f"\n检查AI列的source_cols数据:")
    for config in ai_columns_config:
        target_col = config['target_col']
        source_cols = config['source_cols']
        depends_on = config['depends_on']
        
        print(f"\n{target_col}列:")
        print(f"  source_cols: {source_cols}")
        print(f"  depends_on: {depends_on}")
        
        for source_col in source_cols:
            # 计算列索引（支持多列字母，如AD）
            col_idx = 0
            for i, c in enumerate(reversed(source_col)):
                col_idx += (ord(c) - ord('A') + 1) * (26 ** i)
            
            print(f"\n  检查源列 {source_col}:")
            print(f"    列索引: {col_idx}")
            
            # 显示前5行数据
            print(f"    前5行数据:")
            for row_idx in range(3, min(8, ws.max_row + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                print(f"      行{row_idx}: {cell_value}")
    
    wb.close()
    
    # 检查field_pipelines中的enabled字段
    print(f"\n" + "=" * 100)
    print("检查field_pipelines中的enabled字段")
    print("=" * 100)
    
    try:
        connection = pymysql.connect(
            host='172.18.207.224',
            port=3306,
            user='app',
            password='app123456',
            database='demo',
            charset='utf8mb4'
        )
        cursor = connection.cursor()
        
        # 查询AI列的配置
        sql = """
        SELECT target_col, target_header, enabled
        FROM field_pipelines
        WHERE file_type = 'CUSTOMS' AND target_col IN ('X', 'Y', 'J', 'K')
        ORDER BY target_col
        """
        
        cursor.execute(sql)
        results = cursor.fetchall()
        
        print(f"\n找到 {len(results)} 个AI列:\n")
        
        for result in results:
            target_col, target_header, enabled = result
            
            print(f"{target_col} ({target_header}):")
            print(f"  enabled: {enabled}")
            print()
        
        connection.close()
        
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()
    
    print("\n" + "=" * 100)
    print("调试完成！")
    print("=" * 100)

if __name__ == "__main__":
    debug_ai_columns_not_processed()
