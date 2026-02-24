"""
检查用户上传的实际文件
"""
import openpyxl
import os

# 从日志中找到的用户上传文件路径
user_file = r"C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_ec161292\original.xlsx"

print("=" * 80)
print("检查用户上传的文件")
print("=" * 80)

if not os.path.exists(user_file):
    print(f"[FAIL] 文件不存在: {user_file}")
else:
    wb = openpyxl.load_workbook(user_file)
    sheet = wb.active
    
    print(f"\n工作表名称: {sheet.title}")
    print(f"总行数: {sheet.max_row}")
    print(f"总列数: {sheet.max_column}")
    
    print("\n【第1行（表头）所有列】")
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"  {col_letter}: {cell_value}")
    
    print("\n【第2行（第一行数据）所有列】")
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=2, column=col).value
        col_letter = openpyxl.utils.get_column_letter(col)
        if cell_value is not None:
            print(f"  {col_letter}: {cell_value}")
    
    print("\n" + "=" * 80)
