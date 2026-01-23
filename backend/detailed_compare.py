#!/usr/bin/env python3
"""
详细比较生成的结果文件与成品文件之间的差异，输出具体哪些字段不同
"""

import os
import sys
from openpyxl import load_workbook

def detailed_compare(file1_path: str, file2_path: str):
    """
    详细比较两个Excel文件的差异
    
    Args:
        file1_path: 第一个文件路径（生成的结果文件）
        file2_path: 第二个文件路径（成品文件）
    """
    print(f"\n=== 详细比较文件 ===")
    print(f"文件1: {file1_path}")
    print(f"文件2: {file2_path}")
    
    # 加载工作簿
    wb1 = load_workbook(file1_path, data_only=True)
    wb2 = load_workbook(file2_path, data_only=True)
    
    # 获取活动工作表
    ws1 = wb1.active
    ws2 = wb2.active
    
    # 读取所有数据
    data1 = list(ws1.iter_rows(values_only=True))
    data2 = list(ws2.iter_rows(values_only=True))
    
    # 获取表头
    header1 = list(data1[0])
    header2 = list(data2[0])
    
    # 确保表头一致
    if header1 != header2:
        print("表头不一致，无法进行详细比较")
        return
    
    # 统计各字段的差异数
    field_diff_count = {}
    for col in header1:
        field_diff_count[col] = 0
    
    total_diff = 0
    
    # 比较前10行数据的具体差异
    print(f"\n=== 前10行数据详细差异 ===")
    max_rows = min(11, len(data1), len(data2))  # 比较前10行数据（跳过表头）
    
    for i in range(1, max_rows):
        row1 = list(data1[i])
        row2 = list(data2[i])
        
        print(f"\n第{i+1}行:")
        row_has_diff = False
        
        for j in range(len(header1)):
            val1 = row1[j]
            val2 = row2[j]
            field_name = header1[j]
            
            if val1 != val2:
                row_has_diff = True
                field_diff_count[field_name] += 1
                total_diff += 1
                print(f"  {field_name}: 文件1='{val1}', 文件2='{val2}'")
        
        if not row_has_diff:
            print("  ✅ 该行无差异")
    
    # 输出字段差异统计
    print(f"\n=== 字段差异统计 ===")
    for field, count in field_diff_count.items():
        if count > 0:
            print(f"{field}: {count} 处差异")
    
    print(f"\n总差异数: {total_diff} 处")
    print(f"比较的行数: {max_rows - 1} 行")

if __name__ == "__main__":
    # 文件路径
    generated_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_results", "result.xlsx")
    finished_file = "c:\\Users\\harris.xie\\Documents\\trae_projects\\japan\\docs\\派送文件成品.xlsx"
    
    # 详细比较
    detailed_compare(generated_file, finished_file)
