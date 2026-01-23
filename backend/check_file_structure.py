#!/usr/bin/env python3
"""
检查源文件和成品文件的结构
"""

import os
from openpyxl import load_workbook

def check_file_structure(file_path: str):
    """
    检查Excel文件的结构
    
    Args:
        file_path: 文件路径
    """
    print(f"\n=== 检查文件结构: {os.path.basename(file_path)} ===")
    print(f"文件路径: {file_path}")
    
    # 加载工作簿
    wb = load_workbook(file_path)
    
    # 获取所有工作表名称
    print(f"\n工作表列表: {wb.sheetnames}")
    
    # 检查每个工作表
    for sheet_name in wb.sheetnames:
        print(f"\n--- 工作表: {sheet_name} ---")
        ws = wb[sheet_name]
        
        # 获取数据
        data = list(ws.iter_rows(values_only=True))
        
        # 打印基本信息
        print(f"总行数: {len(data)}")
        print(f"总列数: {len(data[0]) if data else 0}")
        
        # 打印表头
        if data:
            header = list(data[0])
            print(f"\n表头: {header}")
            print(f"表头列数: {len(header)}")
            
            # 打印前5行数据（包含表头）
            print(f"\n前5行数据:")
            for i, row in enumerate(data[:5]):
                print(f"第{i+1}行: {list(row)}")
            
            # 打印列索引映射
            print(f"\n列索引映射:")
            for idx, col_name in enumerate(header):
                print(f"{idx+1}: {col_name}")

if __name__ == "__main__":
    # 源文件路径
    source_file = "c:\\Users\\harris.xie\\Documents\\trae_projects\\japan\\docs\\佐川指定时间带表格源文件（对应 派送文件成品）.xlsx"
    finished_file = "c:\\Users\\harris.xie\\Documents\\trae_projects\\japan\\docs\\派送文件成品.xlsx"
    
    # 检查源文件结构
    check_file_structure(source_file)
    
    # 检查成品文件结构
    check_file_structure(finished_file)
