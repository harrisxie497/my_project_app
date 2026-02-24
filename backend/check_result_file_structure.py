"""
检查结果文件的结构
"""

import openpyxl
from openpyxl import load_workbook

def check_result_file_structure():
    """检查结果文件的结构"""
    print("=" * 100)
    print("检查结果文件的结构")
    print("=" * 100)
    
    # 结果文件路径
    result_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\result.xlsx'
    
    # 加载结果文件
    print(f"\n加载结果文件: {result_file_path}")
    workbook = load_workbook(result_file_path, data_only=True)
    
    # 获取工作表
    worksheet = workbook.active
    print(f"工作表名称: {worksheet.title}")
    print(f"总行数: {worksheet.max_row}")
    print(f"总列数: {worksheet.max_column}")
    
    # 获取第一行（表头）
    header_row = list(worksheet[1])
    header_values = [cell.value for cell in header_row]
    
    print("\n结果文件表头:")
    for i, header in enumerate(header_values):
        if header:
            print(f"  列{i+1}: {header}")
    
    # 获取前5行数据
    print("\n前5行数据:")
    for row_idx in range(1, min(6, worksheet.max_row + 1)):
        row_data = []
        for col_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            row_data.append(cell_value)
        print(f"  行{row_idx}: {row_data}")
    
    workbook.close()
    
    print("\n" + "=" * 100)
    print("检查完成！")
    print("=" * 100)

if __name__ == "__main__":
    check_result_file_structure()
