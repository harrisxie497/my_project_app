"""
测试D列（時間帯指定）的计算逻辑
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.core.database import SessionLocal
from app.services.delivery_processor import DeliveryProcessor
import openpyxl

print("=" * 80)
print("测试D列（時間帯指定）的计算逻辑")
print("=" * 80)

# 测试任务目录
task_dir = "storage/tasks/test_d_column_001"
original_file = os.path.join(task_dir, "original.xlsx")

# 创建数据库会话
db = SessionLocal()

try:
    print("\n【步骤1】初始化DeliveryProcessor")
    print("-" * 80)

    header_params = {
        'mawb_no': '160-03270890',
        'flight_no': 'CA123',
        'arrival_date': '2026-02-10'
    }

    processor = DeliveryProcessor(
        task_dir=task_dir,
        db_session=db,
        file_type='DELIVERY',
        header_params=header_params
    )

    print("DeliveryProcessor初始化成功")
    print(f"header_params: {header_params}")

    print("\n【步骤2】处理文件")
    print("-" * 80)

    result = processor.process()

    print("文件处理完成")
    print(f"输出文件: {result['output_file']}")
    print(f"统计信息: {result['stats']}")

    print("\n【步骤3】验证结果文件中的D列")
    print("-" * 80)

    # 读取结果文件
    result_file = result['output_file']
    if os.path.exists(result_file):
        wb = openpyxl.load_workbook(result_file)
        ws = wb.active

        # 检查D列的数据
        expected_results = [
            {
                "row": 2,
                "C": "2026-02-10",
                "D": "00",
                "desc": "C不为空，D=0 -> 应为00"
            },
            {
                "row": 3,
                "C": "2026-02-11",
                "D": "05",
                "desc": "C不为空，D=5 -> 应为05"
            },
            {
                "row": 4,
                "C": "2026-02-12",
                "D": "12",
                "desc": "C不为空，D=12 -> 应为12"
            },
            {
                "row": 5,
                "C": "",
                "D": "",
                "desc": "C为空，D=0 -> 应为空"
            },
            {
                "row": 6,
                "C": "",
                "D": "08",
                "desc": "C为空，D=8 -> 应为08"
            },
            {
                "row": 7,
                "C": "",
                "D": "15",
                "desc": "C为空，D=15 -> 应为15"
            }
        ]

        all_passed = True

        for expected in expected_results:
            row_num = expected["row"]
            c_col = ws.cell(row=row_num, column=3).value  # C列
            d_col = ws.cell(row=row_num, column=4).value  # D列

            print(f"\n第{row_num}行 ({expected['desc']}):")
            print(f"  C列(配達指定日): 实际={repr(c_col)}, 期望={repr(expected['C'])}")
            print(f"  D列(時間帯指定): 实际={repr(d_col)}, 期望={repr(expected['D'])}")

            # 检查C列
            c_passed = (c_col == expected["C"] or (c_col is None and expected["C"] == "") or (str(c_col).strip() == "" and expected["C"] == ""))

            # 检查D列：考虑openpyxl将空字符串保存为None
            d_expected = expected["D"]
            d_actual = d_col
            if d_expected == "":
                d_expected = None  # 空字符串会被保存为None
            if d_actual == "" and d_expected is None:
                d_actual = None  # 空字符串会被保存为None

            d_passed = (d_actual == d_expected or (d_actual is None and d_expected is None) or (str(d_actual).strip() == "" and d_expected == ""))

            if c_passed and d_passed:
                print(f"  [通过]")
            else:
                print(f"  [失败]")
                all_passed = False

        print("\n" + "=" * 80)
        if all_passed:
            print("[成功] D列计算逻辑测试通过！")
        else:
            print("[失败] D列计算逻辑测试失败！")
        print("=" * 80)

        wb.close()
    else:
        print(f"错误: 结果文件不存在 - {result_file}")

except Exception as e:
    print(f"\n错误: {str(e)}")
    import traceback
    traceback.print_exc()
finally:
    db.close()
