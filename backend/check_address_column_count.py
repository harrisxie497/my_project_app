"""
检查收件人地址列的输入输出数量
"""

import openpyxl

def check_address_column_count():
    """检查收件人地址列的输入输出数量"""
    print("=" * 100)
    print("检查收件人地址列的输入输出数量")
    print("=" * 100)
    
    original_file = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_fc2fe5d3\original.xlsx'
    result_file = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_fc2fe5d3\result.xlsx'
    
    # 读取原始文件
    original_workbook = openpyxl.load_workbook(original_file)
    original_sheet = original_workbook.active
    
    # 获取表头
    original_headers = []
    for cell in original_sheet[1]:
        original_headers.append(cell.value)
    
    print(f"\n原始文件表头: {original_headers}")
    
    # 查找收件人地址列
    if 'お届け先住所' in original_headers:
        col_index = original_headers.index('お届け先住所') + 1
        print(f"\nお届け先住所列（列{col_index}）:")
        
        # 获取数据
        original_data = []
        for row_idx in range(2, original_sheet.max_row + 1):
            cell_value = original_sheet.cell(row=row_idx, column=col_index).value
            original_data.append(cell_value)
        
        print(f"  原始数据数量: {len(original_data)}")
        print(f"  非空数据数量: {sum(1 for v in original_data if v is not None and (not isinstance(v, str) or v.strip() != ''))}")
        print(f"  前10个数据: {original_data[:10]}")
    
    # 读取结果文件
    result_workbook = openpyxl.load_workbook(result_file)
    result_sheet = result_workbook.active
    
    # 获取表头
    result_headers = []
    for cell in result_sheet[1]:
        result_headers.append(cell.value)
    
    print(f"\n结果文件表头: {result_headers}")
    
    # 查找收件人地址列
    if 'お届け先住所' in result_headers:
        col_index = result_headers.index('お届け先住所') + 1
        print(f"\nお届け先住所列（列{col_index}）:")
        
        # 获取数据
        result_data = []
        for row_idx in range(2, result_sheet.max_row + 1):
            cell_value = result_sheet.cell(row=row_idx, column=col_index).value
            result_data.append(cell_value)
        
        print(f"  结果数据数量: {len(result_data)}")
        print(f"  非空数据数量: {sum(1 for v in result_data if v is not None and (not isinstance(v, str) or v.strip() != ''))}")
        print(f"  前10个数据: {result_data[:10]}")
    
    print("\n" + "=" * 100)
    print("检查完成")
    print("=" * 100)

if __name__ == "__main__":
    check_address_column_count()
