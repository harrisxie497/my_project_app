"""
测试脚本：验证customs_processor使用process_header_row返回值
"""
import sys
import os
import logging

# 添加项目路径
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from app.services.customs_processor import CustomsProcessor
from app.core.database import SessionLocal

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def main():
    """
    主函数：测试customs_processor
    """
    logger.info("=" * 100)
    logger.info("测试：customs_processor使用process_header_row返回值")
    logger.info("=" * 100)

    # 任务目录
    task_dir = r"c:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_0a22941a"

    try:
        # 创建数据库会话
        db = SessionLocal()

        try:
            # 创建处理器实例
            header_params = {
                "mawb_no": "16003279161",
                "flight_no": "CX509",
                "arrival_date": "20251210"
            }

            logger.info(f"创建处理器 - task_dir: {task_dir}")
            logger.info(f"header_params: {header_params}")

            processor = CustomsProcessor(
                task_dir=task_dir,
                db_session=db,
                file_type='CUSTOMS',
                header_params=header_params
            )

            # 执行处理
            logger.info("")
            logger.info("开始执行处理...")
            stats = processor.process()

            logger.info("")
            logger.info("=" * 100)
            logger.info("处理完成!")
            logger.info("=" * 100)
            logger.info(f"统计信息: {stats}")
            logger.info(f"结果文件路径: {processor.result_file_path}")

            # 验证输出文件
            logger.info("")
            logger.info("验证输出文件...")

            from openpyxl import load_workbook

            workbook = load_workbook(processor.result_file_path, data_only=True)
            worksheet = workbook.active

            # 读取第1行（特殊第一行）
            first_row_values = []
            for col_idx in range(1, min(11, 42)):  # 读取前10列
                cell_value = worksheet.cell(row=1, column=col_idx).value
                first_row_values.append(str(cell_value) if cell_value is not None else "")

            logger.info("")
            logger.info("验证 - 第1行（特殊第一行）前10列:")
            for idx, val in enumerate(first_row_values):
                col_letter = chr(65 + idx) if idx < 26 else f"{chr(65 + idx // 26 - 1)}{chr(65 + idx % 26)}"
                logger.info(f"  {col_letter}1: '{val}'")

            # 检查B1、E1、H1的值
            logger.info("")
            logger.info("验证关键字段:")
            b1_val = worksheet.cell(row=1, column=2).value
            e1_val = worksheet.cell(row=1, column=5).value
            h1_val = worksheet.cell(row=1, column=8).value

            logger.info(f"  B1: {b1_val}")
            logger.info(f"  E1: {e1_val}")
            logger.info(f"  H1: {h1_val}")

            # 验证期望值
            expected_b1 = "MAWB NO：16003279161"
            expected_e1 = "FLIGHT NO：CX509"
            expected_h1 = "ARRIVAL DATE：20251210"

            all_match = True
            if b1_val == expected_b1:
                logger.info("  ✓ B1值正确")
            else:
                logger.warning(f"  ✗ B1值错误: 期望='{expected_b1}', 实际='{b1_val}'")
                all_match = False

            if e1_val == expected_e1:
                logger.info("  ✓ E1值正确")
            else:
                logger.warning(f"  ✗ E1值错误: 期望='{expected_e1}', 实际='{e1_val}'")
                all_match = False

            if h1_val == expected_h1:
                logger.info("  ✓ H1值正确")
            else:
                logger.warning(f"  ✗ H1值错误: 期望='{expected_h1}', 实际='{h1_val}'")
                all_match = False

            workbook.close()

            logger.info("")
            logger.info("=" * 100)
            if all_match:
                logger.info("✓ 所有验证通过!")
            else:
                logger.warning("⚠ 部分验证失败")
            logger.info("=" * 100)

        finally:
            db.close()

    except Exception as e:
        logger.error(f"测试失败: {str(e)}", exc_info=True)
        raise


if __name__ == "__main__":
    main()
