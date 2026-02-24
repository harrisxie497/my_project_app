"""
检查C列（HAWB番号）的源数据
"""

from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


def check_c_column():
    """检查C列（HAWB番号）的源数据"""
    print("=" * 100)
    print("检查C列（HAWB番号）的源数据")
    print("=" * 100)
    
    original_file_path = "C:\\Users\\harris.xie\\Documents\\trae_projects\\japan\\backend\\storage\\tasks\\t_0fc5b76e\\original.xlsx"
    
    try:
        # 加载工作簿
        workbook = load_workbook(original_file_path)
        sheet = workbook.active
        
        print(f"\n原始文件路径: {original_file_path}")
        print(f"工作表名称: {sheet.title}")
        
        # 读取表头行
        header_row = []
        for cell in sheet[2]:
            header_row.append(cell.value)
        
        print(f"\n表头行数据: {header_row[:10]}")
        
        # 找到C列的索引
        c_col_index = None
        for idx, header in enumerate(header_row):
            if header == 'HAWB番号':
                c_col_index = idx
                print(f"\n找到C列，索引: {c_col_index}")
                break
        
        if c_col_index is not None:
            # 读取C列的数据
            c_column_data = []
            for row_idx in range(3, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=c_col_index + 1).value
                c_column_data.append(cell_value)
            
            print(f"\nC列数据行数: {len(c_column_data)}")
            print(f"C列前10个数据: {c_column_data[:10]}")
            print(f"C列最后10个数据: {c_column_data[-10:]}")
        else:
            print(f"\n⚠️ 没有找到C列")
        
        # 关闭工作簿
        workbook.close()
        
        print("\n" + "=" * 100)
        print("检查完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"\n❌ 检查失败: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    check_c_column()
