import openpyxl
import os

def check_ai_columns_result():
    """检查结果文件中AI列的数据"""
    print("=" * 100)
    print("检查结果文件中AI列的数据")
    print("=" * 100)
    
    try:
        # 结果文件路径
        result_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\result.xlsx'
        
        if not os.path.exists(result_file_path):
            print(f"\n❌ 结果文件不存在: {result_file_path}")
            return
        
        print(f"\n✅ 结果文件存在: {result_file_path}")
        
        # 读取结果文件
        wb = openpyxl.load_workbook(result_file_path)
        ws = wb.active
        
        print(f"\n工作表名称: {ws.title}")
        print(f"数据行数: {ws.max_row}")
        print(f"数据列数: {ws.max_column}")
        
        # 获取表头（第2行）
        result_headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=2, column=col_idx).value
            result_headers.append(str(cell_value) if cell_value else '')
        
        # 创建表头到列索引的映射
        header_to_index = {}
        for idx, header in enumerate(result_headers):
            if header:
                header_to_index[header] = idx + 1
        
        # 定义要检查的AI列
        ai_columns = {
            'X': '收件人名（日文）',
            'Y': '收件人地址',
            'J': '輸入者名',
            'K': '輸入者住所'
        }
        
        # 检查AI列的数据
        print(f"\n{'=' * 100}")
        print("AI列的数据")
        print(f"{'=' * 100}")
        
        for col_name, header in ai_columns.items():
            if header in header_to_index:
                col_idx = header_to_index[header]
                print(f"\n{header} ({col_name}列):")
                print(f"  列索引: {col_idx}")
                
                # 显示前10行数据
                print(f"  前10行数据:")
                for row_idx in range(3, min(13, ws.max_row + 1)):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    print(f"    行{row_idx}: {cell_value}")
            else:
                print(f"\n{header}: 未找到列")
        
        wb.close()
        
        print("\n" + "=" * 100)
        print("检查完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_ai_columns_result()
