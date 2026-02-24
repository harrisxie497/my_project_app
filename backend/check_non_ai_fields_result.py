"""
检查非AI字段的结果
"""

import openpyxl
from openpyxl import load_workbook
import pymysql

def check_non_ai_fields_result():
    """检查非AI字段的结果"""
    print("=" * 100)
    print("检查非AI字段的结果")
    print("=" * 100)
    
    # 结果文件路径
    result_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\result.xlsx'
    
    # 加载结果文件
    print(f"\n加载结果文件: {result_file_path}")
    workbook = load_workbook(result_file_path, data_only=True)
    
    # 获取工作表
    worksheet = workbook.active
    print(f"工作表名称: {worksheet.title}")
    
    # 查询非AI字段的配置
    print("\n查询非AI字段的配置...")
    connection = pymysql.connect(
        host='172.18.207.224',
        port=3306,
        user='app',
        password='app123456',
        database='demo',
        charset='utf8mb4'
    )
    cursor = connection.cursor()
    
    # 查询非AI字段的配置
    sql = """
    SELECT target_col, target_header, map_op, source_cols, field_type, rule_ref, depends_on, enabled
    FROM field_pipelines
    WHERE file_type = 'CUSTOMS' AND field_type != 'AI'
    ORDER BY target_col
    """
    
    cursor.execute(sql)
    results = cursor.fetchall()
    
    print(f"\n找到 {len(results)} 个非AI字段的配置:\n")
    
    # 获取结果文件中的列数据
    # 获取第一行（表头）
    header_row = list(worksheet[1])
    header_values = [cell.value for cell in header_row]
    
    print("结果文件表头:")
    for i, header in enumerate(header_values):
        if header:
            print(f"  列{i+1}: {header}")
    
    print("\n" + "=" * 100)
    print("检查非AI字段的数据:")
    print("=" * 100)
    
    # 检查每个非AI字段的数据
    for result in results:
        target_col, target_header, map_op, source_cols, field_type, rule_ref, depends_on, enabled = result
        
        print(f"\n{target_col} ({target_header}):")
        print(f"  map_op: {map_op}")
        print(f"  source_cols: {source_cols}")
        print(f"  field_type: {field_type}")
        print(f"  rule_ref: {rule_ref}")
        print(f"  depends_on: {depends_on}")
        print(f"  enabled: {enabled}")
        
        # 在结果文件中查找该列的数据
        col_index = None
        for i, header in enumerate(header_values):
            if header == target_header:
                col_index = i
                break
        
        if col_index is not None:
            print(f"  ✅ 在结果文件中找到该列（列{col_index+1}）")
            # 获取前5行数据
            data = []
            for row_idx in range(2, min(7, worksheet.max_row + 1)):
                cell_value = worksheet.cell(row=row_idx, column=col_index + 1).value
                data.append(cell_value)
            print(f"  数据（前5行）: {data}")
        else:
            print(f"  ❌ 在结果文件中未找到该列")
    
    connection.close()
    workbook.close()
    
    print("\n" + "=" * 100)
    print("检查完成！")
    print("=" * 100)

if __name__ == "__main__":
    check_non_ai_fields_result()
