"""
创建测试DELIVERY文件，包含空的依頼主、依頼主住所、依頼主電話列
"""
import openpyxl
from openpyxl.styles import Font
import os

# 测试文件路径
test_file_path = "storage/tasks/test_delivery_jkm_001/original/test_delivery.xlsx"

# 确保目录存在
os.makedirs(os.path.dirname(test_file_path), exist_ok=True)

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active

# 设置工作表名称
ws.title = "Speedy"

# 写入表头（第1行）
headers = [
    "お客様管理番号", "佐川問合せ番号HAWB", "配達指定日", "時間帯指定", "貨物個数",
    "お届け先人名", "お届け先住所", "お届け先電話", "お届け先郵便",
    "依頼主", "依頼主住所", "依頼主郵便番号", "依頼主電話",
    "佐川顧客コード（固定）", "記事欄2（品名）", "記事欄2", "記事欄3"
]

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True)

# 写入测试数据（第2行开始）- J、K、M列为空
test_data = [
    ["TEST001", "SG123456789", "2026-02-10", "午前", 1, "テスト受取人", "東京都渋谷区恵比寿1-1-1", "03-1234-5678", "150-0012",
     "", "", "100-0001", "",  # J列(依頼主)、K列(依頼主住所)、M列(依頼主電話)为空
     "DIDA", "テスト品名", "160-0327 0890", "備考テスト"],  # 第2行
    
    ["TEST002", "SG987654321", "2026-02-10", "午後", 2, "テスト受取人2", "東京都新宿区新宿1-1-1", "03-8765-4321", "160-0022",
     "別の依頼主", "東京都千代田区千代田1-1-1", "100-0002", "03-1111-1111",  # J、K、M列有值
     "DIDA", "テスト品名2", "160-0327 0890", "備考テスト2"],  # 第3行
    
    ["TEST003", "SG456789123", "2026-02-11", "午前", 1, "テスト受取人3", "東京都港区虎ノ門1-1-1", "03-2222-2222", "105-0001",
     "", "", "100-0003", "",  # J列(依頼主)、K列(依頼主住所)、M列(依頼主電話)为空
     "DIDA", "テスト品名3", "160-0327 0890", "備考テスト3"],  # 第4行
]

for row_idx, row_data in enumerate(test_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 保存文件
wb.save(test_file_path)

print(f"测试文件已创建: {test_file_path}")
print(f"包含 {len(test_data)} 行数据")
print(f"第2行和第4行的J、K、M列为空，第3行的J、K、M列有值")
print("\n预期结果:")
print("  第2行: J列='DIDA', K列='千葉県流山市平方8061GLPALFALINK81F13番シャッター', M列='0471377848'")
print("  第3行: J列='別の依頼主', K列='東京都千代田区千代田1-1-1', M列='03-1111-1111' (保持源值)")
print("  第4行: J列='DIDA', K列='千葉県流山市平方8061GLPALFALINK81F13番シャッター', M列='0471377848'")
