import openpyxl
import os

def check_calc_copy_format_by_header():
    """根据表头检查CALC、COPY、FORMAT类型的列"""
    print("=" * 100)
    print("根据表头检查CALC、COPY、FORMAT类型的列")
    print("=" * 100)
    
    try:
        # 结果文件路径
        result_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\result.xlsx'
        
        if not os.path.exists(result_file_path):
            print(f"\n❌ 结果文件不存在: {result_file_path}")
            return
        
        print(f"\n✅ 结果文件存在: {result_file_path}")
        
        # 读取结果文件
        wb = openpyxl.load_workbook(result_file_path)
        ws = wb.active
        
        print(f"\n工作表名称: {ws.title}")
        print(f"数据行数: {ws.max_row}")
        print(f"数据列数: {ws.max_column}")
        
        # 获取表头（第二行）
        header_row = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=2, column=col_idx).value
            header_row.append(str(cell_value) if cell_value else '')
        
        # 创建表头到列索引的映射
        header_to_index = {}
        for idx, header in enumerate(header_row):
            if header:
                header_to_index[header] = idx + 1  # 列索引从1开始
        
        print(f"\n表头到列索引的映射:")
        for header, idx in sorted(header_to_index.items()):
            print(f"  {header}: 列{idx}")
        
        # 定义要检查的列
        test_columns = {
            'CALC': {
                '序号': {'预期': '从1开始，每次递增1', 'col': 'B'},
                'インボイス価格': {'预期': '计算发票价格，四舍五入', 'col': 'R'}
            },
            'COPY': {
                '电商平台码': {'预期': '复制源列的值', 'col': 'AG'},
                '电商平台名称': {'预期': '复制源列的值', 'col': 'AH'},
                'HAWB番号': {'预期': '复制源列的值', 'col': 'C'},
                '備考': {'预期': '复制源列的值', 'col': 'W'}
            },
            'FORMAT': {
                '收件人邮编': {'预期': '格式化数据（如移除横杠、验证格式等）', 'col': 'AA'},
                '輸入者 郵便番号': {'预期': '格式化数据（如移除横杠、验证格式等）', 'col': 'L'},
                '輸入者電話番号': {'预期': '格式化数据（如移除横杠、验证格式等）', 'col': 'M'},
                '運賃': {'预期': '格式化数据（如移除横杠、验证格式等）', 'col': 'U'},
                '收件人电话': {'预期': '格式化数据（如移除横杠、验证格式等）', 'col': 'Z'}
            }
        }
        
        # 检查CALC类型的列
        print(f"\n{'=' * 100}")
        print("CALC类型的列")
        print(f"{'=' * 100}")
        
        for header, info in test_columns['CALC'].items():
            if header in header_to_index:
                col_idx = header_to_index[header]
                print(f"\n{header} ({info['col']}列):")
                print(f"  列索引: {col_idx}")
                print(f"  预期结果: {info['预期']}")
                
                # 显示前5行数据（从第3行开始，因为第2行是表头）
                print(f"  前5行数据:")
                for row_idx in range(3, min(8, ws.max_row + 1)):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    print(f"    行{row_idx}: {cell_value}")
            else:
                print(f"\n{header}: 未找到列")
        
        # 检查COPY类型的列
        print(f"\n{'=' * 100}")
        print("COPY类型的列")
        print(f"{'=' * 100}")
        
        for header, info in test_columns['COPY'].items():
            if header in header_to_index:
                col_idx = header_to_index[header]
                print(f"\n{header} ({info['col']}列):")
                print(f"  列索引: {col_idx}")
                print(f"  预期结果: {info['预期']}")
                
                # 显示前5行数据
                print(f"  前5行数据:")
                for row_idx in range(3, min(8, ws.max_row + 1)):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    print(f"    行{row_idx}: {cell_value}")
            else:
                print(f"\n{header}: 未找到列")
        
        # 检查FORMAT类型的列
        print(f"\n{'=' * 100}")
        print("FORMAT类型的列")
        print(f"{'=' * 100}")
        
        for header, info in test_columns['FORMAT'].items():
            if header in header_to_index:
                col_idx = header_to_index[header]
                print(f"\n{header} ({info['col']}列):")
                print(f"  列索引: {col_idx}")
                print(f"  预期结果: {info['预期']}")
                
                # 显示前5行数据
                print(f"  前5行数据:")
                for row_idx in range(3, min(8, ws.max_row + 1)):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    print(f"    行{row_idx}: {cell_value}")
            else:
                print(f"\n{header}: 未找到列")
        
        wb.close()
        
        print("\n" + "=" * 100)
        print("检查完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_calc_copy_format_by_header()
