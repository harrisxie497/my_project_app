import openpyxl
import os

def find_ai_columns_in_original():
    """在原始文件中查找AI列的正确位置"""
    print("=" * 100)
    print("在原始文件中查找AI列的正确位置")
    print("=" * 100)
    
    try:
        # 原始文件路径
        original_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\original.xlsx'
        
        # 读取原始文件
        wb = openpyxl.load_workbook(original_file_path)
        ws = wb.active
        
        print(f"\n工作表名称: {ws.title}")
        print(f"数据行数: {ws.max_row}")
        print(f"数据列数: {ws.max_column}")
        
        # 获取表头（第2行）
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=2, column=col_idx).value
            headers.append(str(cell_value) if cell_value else '')
        
        # 查找包含"收件人名"、"收件人地址"、"輸入者名"、"輸入者住所"的列
        target_headers = ['收件人名（日文）', '收件人地址', '輸入者名', '輸入者住所']
        
        print(f"\n查找目标列:")
        for target_header in target_headers:
            for idx, header in enumerate(headers):
                if target_header in header:
                    col_letter = chr(65 + idx) if idx < 26 else f"A{chr(65 + idx - 26)}"
                    print(f"\n{target_header}:")
                    print(f"  列索引: {idx + 1}")
                    print(f"  列字母: {col_letter}")
                    print(f"  表头: {header}")
                    
                    # 显示前5行数据
                    print(f"  前5行数据:")
                    for row_idx in range(3, min(8, ws.max_row + 1)):
                        cell_value = ws.cell(row=row_idx, column=idx + 1).value
                        print(f"    行{row_idx}: {cell_value}")
                    break
        
        wb.close()
        
        print("\n" + "=" * 100)
        print("查找完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    find_ai_columns_in_original()
