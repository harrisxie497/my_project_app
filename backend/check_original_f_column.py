"""
检查原始文件中F列（貨物重量）的值
"""

from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


def check_original_f_column():
    """检查原始文件中F列（貨物重量）的值"""
    print("=" * 100)
    print("检查原始文件中F列（貨物重量）的值")
    print("=" * 100)
    
    original_file_path = "C:\\Users\\harris.xie\\Documents\\trae_projects\\japan\\backend\\storage\\tasks\\t_f37e2c5b\\original.xlsx"
    
    try:
        # 加载工作簿
        workbook = load_workbook(original_file_path)
        sheet = workbook.active
        
        print(f"\n原始文件路径: {original_file_path}")
        print(f"工作表名称: {sheet.title}")
        print(f"最大行数: {sheet.max_row}")
        print(f"最大列数: {sheet.max_column}")
        
        # 读取表头行
        header_row = []
        for cell in sheet[2]:
            header_row.append(cell.value)
        
        print(f"\n表头行数据: {header_row[:10]}")
        
        # 找到F列（貨物重量）的索引
        f_col_index = None
        for idx, header in enumerate(header_row):
            if header == '貨物重量':
                f_col_index = idx
                print(f"\n找到F列，索引: {f_col_index}")
                break
        
        if f_col_index is not None:
            # 读取F列的数据
            f_column_data = []
            for row_idx in range(3, min(sheet.max_row + 1, 10)):
                cell_value = sheet.cell(row=row_idx, column=f_col_index + 1).value
                f_column_data.append(cell_value)
            
            print(f"\nF列数据行数: {len(f_column_data)}")
            print(f"F列前10个数据: {f_column_data[:10]}")
        
        # 关闭工作簿
        workbook.close()
        
        print("\n" + "=" * 100)
        print("检查完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"\n❌ 检查失败: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    check_original_f_column()
