"""
检查DELIVERY类型任务的原始文件和结果文件
"""

import sys
import os

# 添加项目根目录到Python路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook

def check_delivery_task_files():
    """检查DELIVERY类型任务的原始文件和结果文件"""
    print("=" * 100)
    print("检查DELIVERY类型任务的原始文件和结果文件")
    print("=" * 100)
    
    try:
        # 读取原始文件
        original_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_fc2fe5d3\original.xlsx'
        workbook = load_workbook(original_file_path)
        sheet = workbook.active
        
        print(f"\n原始文件信息:")
        print(f"  工作表名称: {sheet.title}")
        print(f"  最大行数: {sheet.max_row}")
        print(f"  最大列数: {sheet.max_column}")
        
        # 检查表头
        print(f"\n表头（第1行）:")
        for col_idx in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col_idx).value
            print(f"  列{col_idx}: {header}")
        
        # 检查数据
        print(f"\n数据（第2-3行）:")
        for row_idx in range(2, min(4, sheet.max_row + 1)):
            print(f"  第{row_idx}行:")
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                print(f"    列{col_idx}: {cell_value}")
        
        workbook.close()
        
        # 读取结果文件
        result_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_fc2fe5d3\result.xlsx'
        workbook = load_workbook(result_file_path)
        sheet = workbook.active
        
        print(f"\n结果文件信息:")
        print(f"  工作表名称: {sheet.title}")
        print(f"  最大行数: {sheet.max_row}")
        print(f"  最大列数: {sheet.max_column}")
        
        # 检查表头
        print(f"\n表头（第1行）:")
        for col_idx in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col_idx).value
            print(f"  列{col_idx}: {header}")
        
        # 检查数据
        print(f"\n数据（第2-3行）:")
        for row_idx in range(2, min(4, sheet.max_row + 1)):
            print(f"  第{row_idx}行:")
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                print(f"    列{col_idx}: {cell_value}")
        
        workbook.close()
    
    except Exception as e:
        print(f"检查失败: {e}")
        import traceback
        traceback.print_exc()
    
    print("\n" + "=" * 100)
    print("检查完成")
    print("=" * 100)

if __name__ == "__main__":
    check_delivery_task_files()
