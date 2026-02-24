"""
测试A、B、C列的格式化处理
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.core.database import SessionLocal
from app.services.delivery_processor import DeliveryProcessor
import openpyxl

print("=" * 80)
print("测试A、B、C列的格式化处理")
print("=" * 80)

# 测试任务目录
task_dir = "storage/tasks/test_abc_columns_001"
original_file = os.path.join(task_dir, "original.xlsx")

# 创建数据库会话
db = SessionLocal()

try:
    print("\n【步骤1】初始化DeliveryProcessor")
    print("-" * 80)

    header_params = {
        'mawb_no': '160-03270890',
        'flight_no': 'CA123',
        'arrival_date': '2026-02-10'
    }

    processor = DeliveryProcessor(
        task_dir=task_dir,
        db_session=db,
        file_type='DELIVERY',
        header_params=header_params
    )

    print("DeliveryProcessor初始化成功")
    print(f"header_params: {header_params}")

    print("\n【步骤2】处理文件")
    print("-" * 80)

    result = processor.process()

    print("文件处理完成")
    print(f"输出文件: {result['output_file']}")
    print(f"统计信息: {result['stats']}")

    print("\n【步骤3】验证结果文件中的A、B、C列")
    print("-" * 80)

    # 读取结果文件
    result_file = result['output_file']
    if os.path.exists(result_file):
        wb = openpyxl.load_workbook(result_file)
        ws = wb.active

        # 检查A、B、C列的数据（从第2行开始）
        expected_results = [
            {
                "row": 2,
                "A": "123456",
                "B": "SG123456789",
                "C": "2026-02-10",
                "desc": "A列数字转字符串，C列datetime格式化"
            },
            {
                "row": 3,
                "A": "789012",
                "B": "SG987654321",
                "C": "2026-02-11",
                "desc": "A列数字转字符串，C列字符串保持"
            },
            {
                "row": 4,
                "A": "345678",
                "B": "SG456789123",
                "C": "2026-02-12",
                "desc": "A列数字转字符串，C列中文日期格式化"
            }
        ]

        all_passed = True

        for expected in expected_results:
            row_num = expected["row"]
            a_col = ws.cell(row=row_num, column=1).value  # A列
            b_col = ws.cell(row=row_num, column=2).value  # B列
            c_col = ws.cell(row=row_num, column=3).value  # C列

            # 检查数据类型
            a_type = type(a_col).__name__
            b_type = type(b_col).__name__
            c_type = type(c_col).__name__

            print(f"\n第{row_num}行 ({expected['desc']}):")
            print(f"  A列(お客様管理番号): 实际={repr(a_col)}, 类型={a_type}, 期望={repr(expected['A'])}, 期望类型=str")
            print(f"  B列(佐川問合せ番号HAWB): 实际={repr(b_col)}, 类型={b_type}, 期望={repr(expected['B'])}, 期望类型=str")
            print(f"  C列(配達指定日): 实际={repr(c_col)}, 类型={c_type}, 期望={repr(expected['C'])}, 期望类型=str")

            # 检查值和类型
            a_passed = (a_col == expected["A"] and isinstance(a_col, str))
            b_passed = (b_col == expected["B"] and isinstance(b_col, str))
            c_passed = (c_col == expected["C"] and isinstance(c_col, str))

            if a_passed and b_passed and c_passed:
                print(f"  [通过]")
            else:
                print(f"  [失败]")
                all_passed = False

        print("\n" + "=" * 80)
        if all_passed:
            print("[成功] A、B、C列格式化测试通过！")
        else:
            print("[失败] A、B、C列格式化测试失败！")
        print("=" * 80)

        wb.close()
    else:
        print(f"错误: 结果文件不存在 - {result_file}")

except Exception as e:
    print(f"\n错误: {str(e)}")
    import traceback
    traceback.print_exc()
finally:
    db.close()
