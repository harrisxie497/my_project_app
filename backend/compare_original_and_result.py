import openpyxl
import os

def compare_original_and_result():
    """对比原始文件和结果文件的数据"""
    print("=" * 100)
    print("对比原始文件和结果文件的数据")
    print("=" * 100)
    
    try:
        # 文件路径
        original_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\original.xlsx'
        result_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\result.xlsx'
        
        # 读取原始文件
        print(f"\n读取原始文件: {original_file_path}")
        wb_orig = openpyxl.load_workbook(original_file_path)
        ws_orig = wb_orig.active
        
        print(f"工作表名称: {ws_orig.title}")
        print(f"数据行数: {ws_orig.max_row}")
        print(f"数据列数: {ws_orig.max_column}")
        
        # 读取结果文件
        print(f"\n读取结果文件: {result_file_path}")
        wb_result = openpyxl.load_workbook(result_file_path)
        ws_result = wb_result.active
        
        print(f"工作表名称: {ws_result.title}")
        print(f"数据行数: {ws_result.max_row}")
        print(f"数据列数: {ws_result.max_column}")
        
        # 获取结果文件的表头（第2行）
        result_headers = []
        for col_idx in range(1, ws_result.max_column + 1):
            cell_value = ws_result.cell(row=2, column=col_idx).value
            result_headers.append(str(cell_value) if cell_value else '')
        
        # 创建表头到列索引的映射
        header_to_index = {}
        for idx, header in enumerate(result_headers):
            if header:
                header_to_index[header] = idx + 1
        
        # 定义要检查的COPY和CALC类型的列
        check_columns = {
            'COPY': ['HAWB番号', '現地問合せ番号', '貨物重量', '輸入者 郵便番号', '輸入者電話番号'],
            'CALC': ['序号', 'インボイス価格']
        }
        
        # 检查COPY类型的列
        print(f"\n{'=' * 100}")
        print("COPY类型的列对比")
        print(f"{'=' * 100}")
        
        for header in check_columns['COPY']:
            if header in header_to_index:
                col_idx = header_to_index[header]
                print(f"\n{header}:")
                print(f"  结果文件列索引: {col_idx}")
                
                # 显示原始文件和结果文件的前5行数据
                print(f"  前5行数据对比:")
                for row_idx in range(3, min(8, ws_result.max_row + 1)):
                    # 原始文件的数据（第2行是表头，所以数据从第3行开始）
                    orig_value = ws_orig.cell(row=row_idx, column=col_idx).value
                    # 结果文件的数据
                    result_value = ws_result.cell(row=row_idx, column=col_idx).value
                    print(f"    行{row_idx}: 原始={orig_value}, 结果={result_value}")
            else:
                print(f"\n{header}: 未找到列")
        
        # 检查CALC类型的列
        print(f"\n{'=' * 100}")
        print("CALC类型的列对比")
        print(f"{'=' * 100}")
        
        for header in check_columns['CALC']:
            if header in header_to_index:
                col_idx = header_to_index[header]
                print(f"\n{header}:")
                print(f"  结果文件列索引: {col_idx}")
                
                # 显示结果文件的前5行数据
                print(f"  前5行数据:")
                for row_idx in range(3, min(8, ws_result.max_row + 1)):
                    result_value = ws_result.cell(row=row_idx, column=col_idx).value
                    print(f"    行{row_idx}: {result_value}")
            else:
                print(f"\n{header}: 未找到列")
        
        wb_orig.close()
        wb_result.close()
        
        print("\n" + "=" * 100)
        print("对比完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    compare_original_and_result()
