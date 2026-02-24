"""
检查Y列的输入数据
"""

import sys
import os
import openpyxl

# 添加项目根目录到Python路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.core.database import SessionLocal
from app.models.task import Task

def check_y_column_input_data():
    """检查Y列的输入数据"""
    print("=" * 100)
    print("检查Y列的输入数据")
    print("=" * 100)
    
    db_session = SessionLocal()
    
    try:
        # 查询最新的CUSTOMS类型任务
        task = db_session.query(Task).filter(
            Task.file_type == 'customs'
        ).order_by(Task.created_at.desc()).first()
        
        if task:
            print(f"\n最新CUSTOMS任务: {task.id}")
            
            # 从files字段获取任务目录
            files = task.files if task.files else {}
            task_dir = files.get('task_dir', '')
            print(f"  任务目录: {task_dir}")
            
            if not task_dir:
                print("\n任务目录为空")
                return
            
            original_file = os.path.join(task_dir, 'original.xlsx')
            
            # 读取原始文件
            workbook = openpyxl.load_workbook(original_file)
            sheet = workbook.active
            
            # 获取表头
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)
            
            print(f"\n表头: {headers}")
            
            # 查找M列（收件人地址）
            if '收件人地址(删)2' in headers:
                m_col_index = headers.index('收件人地址(删)2') + 1
                print(f"\nM列（收件人地址(删)2）索引: {m_col_index}")
                
                # 获取数据
                m_data = []
                for row_idx in range(2, min(sheet.max_row + 1, 200)):
                    cell_value = sheet.cell(row=row_idx, column=m_col_index).value
                    m_data.append(cell_value)
                
                print(f"  M列数据数量: {len(m_data)}")
                print(f"  非空数据数量: {sum(1 for v in m_data if v is not None and (not isinstance(v, str) or v.strip() != ''))}")
                print(f"  前20个数据: {m_data[:20]}")
            else:
                print("\n未找到M列")
        else:
            print("\n未找到CUSTOMS类型任务")
    
    finally:
        db_session.close()
    
    print("\n" + "=" * 100)
    print("检查完成")
    print("=" * 100)

if __name__ == "__main__":
    check_y_column_input_data()
