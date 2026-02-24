"""
检查原始文件的数据行数
"""

import sys
import os

# 添加项目根目录到Python路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook

def check_original_file_data():
    """检查原始文件的数据行数"""
    print("=" * 100)
    print("检查原始文件的数据行数")
    print("=" * 100)
    
    try:
        # 读取原始文件
        original_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\original.xlsx'
        workbook = load_workbook(original_file_path)
        sheet = workbook.active
        
        print(f"\n原始文件信息:")
        print(f"  工作表名称: {sheet.title}")
        print(f"  最大行数: {sheet.max_row}")
        print(f"  最大列数: {sheet.max_column}")
        
        # 检查第一列的数据
        print(f"\n第一列（A列）的数据:")
        for row_idx in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=1).value
            print(f"  第{row_idx}行: {cell_value}")
        
        # 检查輸入者電話番号列的数据
        print(f"\n輸入者電話番号列（R列）的数据:")
        for row_idx in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=18).value
            print(f"  第{row_idx}行: {cell_value}")
        
        # 检查收件人电话列的数据
        print(f"\n收件人电话列（AF列）的数据:")
        for row_idx in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=32).value
            print(f"  第{row_idx}行: {cell_value}")
        
        workbook.close()
    
    except Exception as e:
        print(f"检查失败: {e}")
        import traceback
        traceback.print_exc()
    
    print("\n" + "=" * 100)
    print("检查完成")
    print("=" * 100)

if __name__ == "__main__":
    check_original_file_data()
