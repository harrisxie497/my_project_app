"""
检查CUSTOMS类型任务的收件地址列处理情况
"""

import openpyxl
import sys
import os

# 添加项目根目录到Python路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.core.database import SessionLocal
from app.models.task import Task

def check_customs_address_column():
    """检查CUSTOMS类型任务的收件地址列处理情况"""
    print("=" * 100)
    print("检查CUSTOMS类型任务的收件地址列处理情况")
    print("=" * 100)
    
    # 使用测试任务t_aa9d170a
    task_dir = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a'
    
    original_file = os.path.join(task_dir, 'original.xlsx')
    result_file = os.path.join(task_dir, 'result.xlsx')
    
    print(f"\n任务目录: {task_dir}")
    print(f"原始文件: {original_file}")
    print(f"结果文件: {result_file}")
    
    # 读取原始文件
    original_workbook = openpyxl.load_workbook(original_file)
    original_sheet = original_workbook.active
    
    # 获取表头
    original_headers = []
    for cell in original_sheet[1]:
        original_headers.append(cell.value)
    
    print(f"\n原始文件表头: {original_headers}")
    
    # 查找收件人地址列
    if '收件人地址' in original_headers:
        col_index = original_headers.index('收件人地址') + 1
        print(f"\n收件人地址列（列{col_index}）:")
        
        # 获取数据
        original_data = []
        for row_idx in range(2, original_sheet.max_row + 1):
            cell_value = original_sheet.cell(row=row_idx, column=col_index).value
            original_data.append(cell_value)
        
        print(f"  原始数据数量: {len(original_data)}")
        print(f"  非空数据数量: {sum(1 for v in original_data if v is not None and (not isinstance(v, str) or v.strip() != ''))}")
        print(f"  前10个数据: {original_data[:10]}")
    else:
        print("\n未找到收件人地址列")
    
    # 读取结果文件
    result_workbook = openpyxl.load_workbook(result_file)
    result_sheet = result_workbook.active
    
    # 获取表头
    result_headers = []
    for cell in result_sheet[1]:
        result_headers.append(cell.value)
    
    print(f"\n结果文件表头: {result_headers}")
    
    # 查找收件人地址列
    if '收件人地址' in result_headers:
        col_index = result_headers.index('收件人地址') + 1
        print(f"\n收件人地址列（列{col_index}）:")
        
        # 获取数据
        result_data = []
        for row_idx in range(2, result_sheet.max_row + 1):
            cell_value = result_sheet.cell(row=row_idx, column=col_index).value
            result_data.append(cell_value)
        
        print(f"  结果数据数量: {len(result_data)}")
        print(f"  非空数据数量: {sum(1 for v in result_data if v is not None and (not isinstance(v, str) or v.strip() != ''))}")
        print(f"  前10个数据: {result_data[:10]}")
        
        # 检查是否有空值
        empty_indices = [idx for idx, v in enumerate(result_data) if v is None or (isinstance(v, str) and v.strip() == '')]
        if empty_indices:
            print(f"\n空值位置（行号）: {[idx + 2 for idx in empty_indices]}")
            print(f"  对应的原始数据: {[original_data[idx] for idx in empty_indices if idx < len(original_data)]}")
    else:
        print("\n未找到收件人地址列")
    
    print("\n" + "=" * 100)
    print("检查完成")
    print("=" * 100)

if __name__ == "__main__":
    check_customs_address_column()
