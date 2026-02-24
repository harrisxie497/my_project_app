"""
检查收件人地址列及其依赖列的数据情况
"""

import openpyxl

def check_address_and_dependencies():
    """检查收件人地址列及其依赖列的数据情况"""
    print("=" * 100)
    print("检查收件人地址列及其依赖列的数据情况")
    print("=" * 100)
    
    original_file = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_fc2fe5d3\original.xlsx'
    
    # 读取原始文件
    workbook = openpyxl.load_workbook(original_file)
    sheet = workbook.active
    
    # 获取表头
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    
    print(f"\n表头: {headers}")
    
    # 查找收件人地址列及其依赖列
    address_col_idx = None
    dependency_col_indices = []
    
    for idx, header in enumerate(headers):
        if header == 'お届け先住所':
            address_col_idx = idx
            print(f"\nお届け先住所列索引: {idx + 1}")
        elif header in ['收件人省州（删）', '收件人城市（删）', '收件人地址1（删）', '收件人地址2（删）', '收件人地址3（删）']:
            dependency_col_indices.append(idx)
            print(f"依赖列 - {header}: 索引 {idx + 1}")
    
    # 获取收件人地址列的数据
    print(f"\nお届け先住所列数据:")
    address_data = []
    for row_idx in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_idx, column=address_col_idx + 1).value
        address_data.append(cell_value)
    
    print(f"  总数据量: {len(address_data)}")
    print(f"  非空数据量: {sum(1 for v in address_data if v is not None and (not isinstance(v, str) or v.strip() != ''))}")
    print(f"  前10个数据: {address_data[:10]}")
    
    # 获取依赖列的数据
    print(f"\n依赖列数据:")
    for col_idx in dependency_col_indices:
        header = headers[col_idx]
        col_data = []
        for row_idx in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx + 1).value
            col_data.append(cell_value)
        
        print(f"\n{header}:")
        print(f"  总数据量: {len(col_data)}")
        print(f"  非空数据量: {sum(1 for v in col_data if v is not None and (not isinstance(v, str) or v.strip() != ''))}")
        print(f"  前10个数据: {col_data[:10]}")
    
    print("\n" + "=" * 100)
    print("检查完成")
    print("=" * 100)

if __name__ == "__main__":
    check_address_and_dependencies()
