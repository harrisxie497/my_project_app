"""
检查结果文件中AI字段的数据
"""

import openpyxl
from openpyxl import load_workbook

def check_ai_fields_in_result():
    """检查结果文件中AI字段的数据"""
    print("=" * 100)
    print("检查结果文件中AI字段的数据")
    print("=" * 100)
    
    # 结果文件路径
    result_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\result.xlsx'
    
    # 加载结果文件
    print(f"\n加载结果文件: {result_file_path}")
    workbook = load_workbook(result_file_path, data_only=True)
    
    # 获取工作表
    worksheet = workbook.active
    
    # 获取表头行（第2行）
    header_row = list(worksheet[2])
    header_values = [cell.value for cell in header_row]
    
    # AI字段列表
    ai_fields = [
        '品名',
        '材质',
        '輸入者名',
        '輸入者住所',
        '收件人名（日文）',
        '收件人地址'
    ]
    
    # 创建列名到列索引的映射
    col_name_to_index = {}
    for i, header in enumerate(header_values):
        if header:
            col_name_to_index[header] = i
    
    print("\nAI字段的数据（前5行）:")
    for field_name in ai_fields:
        if field_name in col_name_to_index:
            col_index = col_name_to_index[field_name]
            
            # 获取前5行数据
            data = []
            for row_idx in range(3, min(8, worksheet.max_row + 1)):
                cell_value = worksheet.cell(row=row_idx, column=col_index + 1).value
                data.append(cell_value)
            
            print(f"\n{field_name}:")
            print(f"  数据（前5行）: {data}")
            
            # 检查数据是否为空
            all_none = all(d is None for d in data)
            if all_none:
                print(f"  ❌ 所有数据都为空")
            else:
                print(f"  ✅ 有数据")
        else:
            print(f"\n{field_name}:")
            print(f"  ❌ 在结果文件中未找到该列")
    
    workbook.close()
    
    print("\n" + "=" * 100)
    print("检查完成！")
    print("=" * 100)

if __name__ == "__main__":
    check_ai_fields_in_result()
