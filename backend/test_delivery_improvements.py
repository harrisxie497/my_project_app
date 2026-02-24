"""
测试DELIVERY的3个功能改进：
1. 智能数据行判断 - 第1列为空则停止
2. None值替换 - 结果文件中无空值
3. 增强日志记录
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.core.database import SessionLocal
from app.services.delivery_processor import DeliveryProcessor
import openpyxl

print("=" * 80)
print("测试DELIVERY的3个功能改进")
print("=" * 80)

# 测试任务目录 - 使用已有的测试任务
task_dir = "storage/tasks/test_delivery_jkm_001"

# 检查任务目录
if not os.path.exists(task_dir):
    print(f"\n❌ 任务目录不存在: {task_dir}")
    sys.exit(1)

original_file = os.path.join(task_dir, "original.xlsx")
if not os.path.exists(original_file):
    print(f"\n❌ 原始文件不存在: {original_file}")
    sys.exit(1)

print(f"\n✅ 测试目录: {task_dir}")
print(f"✅ 原始文件: {original_file}")

# 读取原始文件，检查第一列数据以验证智能数据行判断
print("\n" + "=" * 80)
print("步骤1: 检查原始文件第一列数据（用于测试智能数据行判断）")
print("=" * 80)

wb_orig = openpyxl.load_workbook(original_file)
ws_orig = wb_orig.active

# 读取A列数据（第1列）
first_col_values = []
for row_idx, row in enumerate(ws_orig.iter_rows(min_col=1, max_col=1), start=1):
    cell_value = row[0].value
    first_col_values.append((row_idx, cell_value))
    if row_idx > 20:  # 只看前20行
        break

print("\n原始文件前20行的第1列数据:")
for row_idx, val in first_col_values:
    print(f"  第{row_idx}行: {repr(val)}")

wb_orig.close()

# 创建数据库会话
db = SessionLocal()

try:
    print("\n" + "=" * 80)
    print("步骤2: 执行DELIVERY处理")
    print("=" * 80)

    header_params = {
        'mawb_no': 'TEST-001',
        'flight_no': 'CA123',
        'arrival_date': '2026-02-10'
    }

    print(f"Header params: {header_params}")

    processor = DeliveryProcessor(
        task_dir=task_dir,
        db_session=db,
        file_type='DELIVERY',
        header_params=header_params
    )

    result = processor.process()

    result_file = processor.result_file_path
    print(f"\n✅ 处理完成")
    print(f"结果文件: {result_file}")

    if result_file and os.path.exists(result_file):
        file_size = os.path.getsize(result_file)
        print(f"结果文件大小: {file_size} 字节")
    else:
        print(f"❌ 结果文件不存在")
        sys.exit(1)

except Exception as e:
    print(f"\n❌ 处理失败: {str(e)}")
    import traceback
    traceback.print_exc()
    db.close()
    sys.exit(1)

finally:
    db.close()

# 验证结果文件
print("\n" + "=" * 80)
print("步骤3: 验证结果文件（测试None值替换）")
print("=" * 80)

wb_result = openpyxl.load_workbook(result_file)
ws_result = wb_result.active

# 检查所有单元格是否有None值
has_none_values = False
none_value_count = 0
total_cells = 0

print("\n检查结果文件中的空值:")
for row in ws_result.iter_rows():
    for cell in row:
        total_cells += 1
        if cell.value is None:
            has_none_values = True
            none_value_count += 1
            print(f"  ❌ 发现None值: 行{cell.row}, 列{cell.column_letter}")

if has_none_values:
    print(f"\n❌ 验证失败: 结果文件中有 {none_value_count} 个None值（应该全部替换为空字符串）")
else:
    print(f"\n✅ 验证通过: 结果文件中没有None值，已全部替换为空字符串")

wb_result.close()

# 总结
print("\n" + "=" * 80)
print("测试总结")
print("=" * 80)
print("""
测试结果：
1. 智能数据行判断 - 请检查日志中是否有"第一列为空，停止读取"的日志
2. None值替换 - """ + ("✅ 通过" if not has_none_values else "❌ 失败"))
print("""
3. 增强日志记录 - 请查看日志文件，应该包含更详细的处理日志
""")

print("请查看日志文件以确认所有改进都已生效:")
print(f"  - 日志文件: backend/logs/app.log")
print("\n关键日志搜索:")
print("  1. 智能数据行判断: 搜索 '第一列为空，停止读取'")
print("  2. None值替换: 搜索 '替换None值为空字符串'")
print("  3. 增强日志: 搜索 '按列处理输入 - column_data'")
