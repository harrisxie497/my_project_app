import openpyxl
import os

def check_original_file():
    """查看任务t_aa9d170a的原始文件"""
    print("=" * 100)
    print("查看任务t_aa9d170a的原始文件")
    print("=" * 100)
    
    try:
        # 原始文件路径
        original_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\original.xlsx'
        
        if not os.path.exists(original_file_path):
            print(f"\n❌ 原始文件不存在: {original_file_path}")
            return
        
        print(f"\n✅ 原始文件存在: {original_file_path}")
        
        # 读取原始文件
        wb = openpyxl.load_workbook(original_file_path)
        ws = wb.active
        
        print(f"\n工作表名称: {ws.title}")
        print(f"数据行数: {ws.max_row}")
        print(f"数据列数: {ws.max_column}")
        
        # 显示前10行数据
        print(f"\n前10行数据:")
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(str(cell_value) if cell_value else '')
            print(f"  行{row_idx}: {', '.join(row_data)}")
        
        wb.close()
        
        print("\n" + "=" * 100)
        print("查看完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_original_file()
