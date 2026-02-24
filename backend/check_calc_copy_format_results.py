import openpyxl
import os

def check_calc_copy_format_results():
    """检查任务t_aa9d170a的结果文件中CALC、COPY、FORMAT类型的列"""
    print("=" * 100)
    print("检查任务t_aa9d170a的结果文件中CALC、COPY、FORMAT类型的列")
    print("=" * 100)
    
    try:
        # 结果文件路径
        result_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\result.xlsx'
        
        if not os.path.exists(result_file_path):
            print(f"\n❌ 结果文件不存在: {result_file_path}")
            return
        
        print(f"\n✅ 结果文件存在: {result_file_path}")
        
        # 读取结果文件
        wb = openpyxl.load_workbook(result_file_path)
        ws = wb.active
        
        print(f"\n工作表名称: {ws.title}")
        print(f"数据行数: {ws.max_row}")
        print(f"数据列数: {ws.max_column}")
        
        # 定义要检查的列
        test_columns = {
            'CALC': {'B': '序号', 'R': 'インボイス価格'},
            'COPY': {'AG': '电商平台码', 'AH': '电商平台名称', 'C': 'HAWB番号', 'W': '備考'},
            'FORMAT': {'AA': '收件人邮编', 'L': '輸入者 郵便番号', 'M': '輸入者電話番号', 'U': '運賃', 'Z': '收件人电话'}
        }
        
        # 获取列名到列索引的映射
        col_name_to_index = {}
        for col_idx in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col_idx)
            if header_cell.value:
                col_name_to_index[header_cell.value] = col_idx
        
        print(f"\n列名到列索引的映射: {col_name_to_index}")
        
        # 检查CALC类型的列
        print(f"\n{'=' * 100}")
        print("CALC类型的列")
        print(f"{'=' * 100}")
        
        for col_name, description in test_columns['CALC'].items():
            if col_name in col_name_to_index:
                col_idx = col_name_to_index[col_name]
                print(f"\n{col_name} ({description}):")
                print(f"  列索引: {col_idx}")
                
                # 显示前5行数据
                print(f"  前5行数据:")
                for row_idx in range(2, min(7, ws.max_row + 1)):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    print(f"    行{row_idx}: {cell_value}")
            else:
                print(f"\n{col_name} ({description}): 未找到列")
        
        # 检查COPY类型的列
        print(f"\n{'=' * 100}")
        print("COPY类型的列")
        print(f"{'=' * 100}")
        
        for col_name, description in test_columns['COPY'].items():
            if col_name in col_name_to_index:
                col_idx = col_name_to_index[col_name]
                print(f"\n{col_name} ({description}):")
                print(f"  列索引: {col_idx}")
                
                # 显示前5行数据
                print(f"  前5行数据:")
                for row_idx in range(2, min(7, ws.max_row + 1)):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    print(f"    行{row_idx}: {cell_value}")
            else:
                print(f"\n{col_name} ({description}): 未找到列")
        
        # 检查FORMAT类型的列
        print(f"\n{'=' * 100}")
        print("FORMAT类型的列")
        print(f"{'=' * 100}")
        
        for col_name, description in test_columns['FORMAT'].items():
            if col_name in col_name_to_index:
                col_idx = col_name_to_index[col_name]
                print(f"\n{col_name} ({description}):")
                print(f"  列索引: {col_idx}")
                
                # 显示前5行数据
                print(f"  前5行数据:")
                for row_idx in range(2, min(7, ws.max_row + 1)):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    print(f"    行{row_idx}: {cell_value}")
            else:
                print(f"\n{col_name} ({description}): 未找到列")
        
        wb.close()
        
        print("\n" + "=" * 100)
        print("检查完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_calc_copy_format_results()
