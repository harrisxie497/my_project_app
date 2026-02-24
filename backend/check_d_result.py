"""
检查D列测试结果
"""
import openpyxl

result_file = "storage/tasks/test_d_column_001/result.xlsx"

print("=" * 80)
print("检查D列测试结果")
print("=" * 80)

wb = openpyxl.load_workbook(result_file)
ws = wb.active

print(f"\n工作表名称: {ws.title}")
print(f"最大行数: {ws.max_row}")

print("\n数据行（第2-7行）:")
for row_idx in range(2, 8):
    c_col = ws.cell(row=row_idx, column=3).value  # C列
    d_col = ws.cell(row=row_idx, column=4).value  # D列
    print(f"  第{row_idx}行: C={repr(c_col)}, D={repr(d_col)}")

wb.close()
