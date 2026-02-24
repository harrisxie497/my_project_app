"""
检查任务ID为t_e4caaadc的结果文件
"""

from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


def check_result_file_e4caaadc():
    """检查任务ID为t_e4caaadc的结果文件"""
    print("=" * 100)
    print("检查任务ID为t_e4caaadc的结果文件")
    print("=" * 100)
    
    result_file_path = "C:\\Users\\harris.xie\\Documents\\trae_projects\\japan\\backend\\storage\\tasks\\t_e4caaadc\\result.xlsx"
    
    try:
        # 检查结果文件
        print(f"\n检查结果文件: {result_file_path}")
        workbook = load_workbook(result_file_path)
        sheet = workbook.active
        
        print(f"工作表名称: {sheet.title}")
        print(f"最大行数: {sheet.max_row}")
        print(f"最大列数: {sheet.max_column}")
        
        # 读取第一行（特殊行）
        first_row = []
        for cell in sheet[1]:
            first_row.append(cell.value)
        
        print(f"\n第一行（特殊行）数据: {first_row[:10]}")
        
        # 读取表头行
        header_row = []
        for cell in sheet[2]:
            header_row.append(cell.value)
        
        print(f"\n表头行数据: {header_row[:10]}")
        
        # 读取数据行
        print(f"\n数据行数: {sheet.max_row - 2}")
        
        # 关闭工作簿
        workbook.close()
        
        print("\n" + "=" * 100)
        print("检查完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"\n❌ 检查失败: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    check_result_file_e4caaadc()
