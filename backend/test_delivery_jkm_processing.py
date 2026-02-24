"""
测试DELIVERY文件的J、K、M列DEFAULT操作处理
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.core.database import SessionLocal
from app.services.delivery_processor import DeliveryProcessor
import openpyxl

print("=" * 80)
print("测试DELIVERY文件的J、K、M列DEFAULT操作处理")
print("=" * 80)

# 测试任务目录
task_dir = "storage/tasks/test_delivery_jkm_001"
original_file = os.path.join(task_dir, "original/test_delivery.xlsx")

# 创建数据库会话
db = SessionLocal()

try:
    print("\n【步骤1】初始化DeliveryProcessor")
    print("-" * 80)

    header_params = {
        'mawb_no': '160-03270890',
        'flight_no': 'CA123',
        'arrival_date': '2026-02-10'
    }

    processor = DeliveryProcessor(
        task_dir=task_dir,
        db_session=db,
        file_type='DELIVERY',
        header_params=header_params
    )

    print("DeliveryProcessor初始化成功")
    print(f"header_params: {header_params}")

    print("\n【步骤2】处理文件")
    print("-" * 80)

    result = processor.process()

    print("文件处理完成")
    print(f"输出文件: {result['output_file']}")
    print(f"统计信息: {result['stats']}")

    print("\n【步骤3】验证结果文件")
    print("-" * 80)

    # 读取结果文件
    result_file = result['output_file']
    if os.path.exists(result_file):
        wb = openpyxl.load_workbook(result_file)
        ws = wb.active

        # 检查J、K、M列的数据（从第2行开始）
        expected_results = [
            {"row": 2, "J": "DIDA", "K": "千葉県流山市平方8061GLPALFALINK81F13番シャッター", "M": "0471377848", "desc": "空值应使用默认值"},
            {"row": 3, "J": "別の依頼主", "K": "東京都千代田区千代田1-1-1", "M": "03-1111-1111", "desc": "有值应保持源值"},
            {"row": 4, "J": "DIDA", "K": "千葉県流山市平方8061GLPALFALINK81F13番シャッター", "M": "0471377848", "desc": "空值应使用默认值"}
        ]

        all_passed = True

        for expected in expected_results:
            row_num = expected["row"]
            j_col = ws.cell(row=row_num, column=10).value  # J列
            k_col = ws.cell(row=row_num, column=11).value  # K列
            m_col = ws.cell(row=row_num, column=13).value  # M列

            print(f"\n第{row_num}行 ({expected['desc']}):")
            print(f"  J列(依頼主): 实际={repr(j_col)}, 期望={repr(expected['J'])}")
            print(f"  K列(依頼主住所): 实际={repr(k_col)}, 期望={repr(expected['K'])}")
            print(f"  M列(依頼主電話): 实际={repr(m_col)}, 期望={repr(expected['M'])}")

            if j_col == expected["J"] and k_col == expected["K"] and m_col == expected["M"]:
                print(f"  [通过]")
            else:
                print(f"  [失败]")
                all_passed = False

        print("\n" + "=" * 80)
        if all_passed:
            print("[成功] DEFAULT操作测试通过！")
        else:
            print("[失败] DEFAULT操作测试失败！")
        print("=" * 80)

        wb.close()
    else:
        print(f"错误: 结果文件不存在 - {result_file}")

except Exception as e:
    print(f"\n错误: {str(e)}")
    import traceback
    traceback.print_exc()
finally:
    db.close()
