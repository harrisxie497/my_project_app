"""
检查原始文件的表头顺序
"""

import sys
import os
from openpyxl import load_workbook

# 添加项目根目录到Python路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

def check_original_headers():
    """检查原始文件的表头顺序"""
    print("=" * 100)
    print("检查原始文件的表头顺序")
    print("=" * 100)
    
    # 任务目录
    task_dir = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a'
    original_file = os.path.join(task_dir, 'original.xlsx')
    
    if not os.path.exists(original_file):
        print(f"原始文件不存在: {original_file}")
        return
    
    # 读取原始文件
    workbook = load_workbook(original_file)
    worksheet = workbook.active
    
    # 获取表头行（第2行，索引为1）
    header_row = 2
    headers = []
    for col in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=header_row, column=col).value
        if cell_value:
            headers.append(str(cell_value))
    
    print(f"\n原始文件的表头顺序（第{header_row}行，共{len(headers)}列）:")
    for idx, header in enumerate(headers, start=1):
        print(f"  {idx}. {header}")
    
    workbook.close()
    
    # 用户提到的表头顺序
    user_headers = ['会员编号', '序号', 'HAWB番号', '貨物個数', '重量単位コード', '品名', '材质', '貨物重量', '現地問合せ番号', '輸入者 郵便番号', '輸入者電話番号', '輸出者名', '輸出者住所', 'インボイス価格条件コード', 'インボイス通貨コード', 'インボイス価格', '運賃区分コード', '運賃通貨コード', '運賃', '原産地コード', '備考', '收件人名（日文）', '收件人地址', '收件人电话', '收件人邮编', '依赖人名', '依赖人地址', '依赖人电话', '收件地址识别码', '电商货识别码', '电商平台码', '电商平台名称', '系统预留列，不可使用', '輸入者名', '輸入者住所']
    
    print(f"\n\n用户提到的表头顺序（共{len(user_headers)}列）:")
    for idx, header in enumerate(user_headers, start=1):
        print(f"  {idx}. {header}")
    
    # 比较两个顺序
    print(f"\n\n比较原始文件表头顺序与用户提到的表头顺序:")
    if headers == user_headers:
        print("  ✓ 两个顺序完全一致")
    else:
        print("  ✗ 两个顺序不一致")
        
        # 找出不一致的地方
        print(f"\n  不一致的列:")
        max_len = max(len(headers), len(user_headers))
        for idx in range(max_len):
            header = headers[idx] if idx < len(headers) else "（无）"
            user_header = user_headers[idx] if idx < len(user_headers) else "（无）"
            if header != user_header:
                print(f"    第{idx+1}列: 原始文件='{header}', 用户提到的='{user_header}'")
    
    print("\n" + "=" * 100)
    print("检查完成")
    print("=" * 100)

if __name__ == "__main__":
    check_original_headers()
