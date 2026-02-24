"""
检查非AI字段的数据是否正确
"""

import openpyxl
from openpyxl import load_workbook
import pymysql

def check_non_ai_fields_data():
    """检查非AI字段的数据是否正确"""
    print("=" * 100)
    print("检查非AI字段的数据是否正确")
    print("=" * 100)
    
    # 结果文件路径
    result_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\result.xlsx'
    
    # 加载结果文件
    print(f"\n加载结果文件: {result_file_path}")
    workbook = load_workbook(result_file_path, data_only=True)
    
    # 获取工作表
    worksheet = workbook.active
    
    # 获取表头行（第2行）
    header_row = list(worksheet[2])
    header_values = [cell.value for cell in header_row]
    
    # 创建列名到列索引的映射
    col_name_to_index = {}
    for i, header in enumerate(header_values):
        if header:
            col_name_to_index[header] = i
    
    # 查询非AI字段的配置
    print("\n查询非AI字段的配置...")
    connection = pymysql.connect(
        host='172.18.207.224',
        port=3306,
        user='app',
        password='app123456',
        database='demo',
        charset='utf8mb4'
    )
    cursor = connection.cursor()
    
    # 查询非AI字段的配置
    sql = """
    SELECT target_col, target_header, map_op, source_cols, field_type, rule_ref, depends_on, enabled
    FROM field_pipelines
    WHERE file_type = 'CUSTOMS' AND field_type != 'AI'
    ORDER BY target_col
    """
    
    cursor.execute(sql)
    results = cursor.fetchall()
    
    print(f"\n找到 {len(results)} 个非AI字段的配置\n")
    
    # 检查每个非AI字段的数据
    for result in results:
        target_col, target_header, map_op, source_cols, field_type, rule_ref, depends_on, enabled = result
        
        # 在结果文件中查找该列
        if target_header in col_name_to_index:
            col_index = col_name_to_index[target_header]
            
            # 获取前5行数据
            data = []
            for row_idx in range(3, min(8, worksheet.max_row + 1)):
                cell_value = worksheet.cell(row=row_idx, column=col_index + 1).value
                data.append(cell_value)
            
            print(f"{target_col} ({target_header}):")
            print(f"  map_op: {map_op}")
            print(f"  field_type: {field_type}")
            print(f"  数据（前5行）: {data}")
            
            # 检查数据是否为空
            all_none = all(d is None for d in data)
            if all_none:
                print(f"  ❌ 所有数据都为空")
            else:
                print(f"  ✅ 有数据")
            print()
    
    connection.close()
    workbook.close()
    
    print("=" * 100)
    print("检查完成！")
    print("=" * 100)

if __name__ == "__main__":
    check_non_ai_fields_data()
