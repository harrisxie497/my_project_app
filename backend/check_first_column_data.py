"""
检查原始文件第一列A的数据情况
"""

import openpyxl

def check_first_column_data():
    """检查原始文件第一列A的数据情况"""
    print("=" * 100)
    print("检查原始文件第一列A的数据情况")
    print("=" * 100)
    
    original_file = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_fc2fe5d3\original.xlsx'
    
    # 读取原始文件
    workbook = openpyxl.load_workbook(original_file)
    sheet = workbook.active
    
    print(f"\n原始文件总行数: {sheet.max_row}")
    
    # 获取第一列A的数据
    first_column_data = []
    for row_idx in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_idx, column=1).value
        first_column_data.append(cell_value)
    
    print(f"\n第一列A的数据:")
    print(f"  总数据量: {len(first_column_data)}")
    print(f"  非空数据量: {sum(1 for v in first_column_data if v is not None and (not isinstance(v, str) or v.strip() != ''))}")
    print(f"  空数据量: {sum(1 for v in first_column_data if v is None or (isinstance(v, str) and v.strip() == ''))}")
    
    # 找出第一个空值的位置
    for idx, value in enumerate(first_column_data):
        if value is None or (isinstance(value, str) and value.strip() == ''):
            print(f"\n第一个空值位置: 行{idx + 2}（索引{idx}）")
            print(f"  前10个数据: {first_column_data[:10]}")
            print(f"  第一个空值前后的数据: {first_column_data[max(0, idx-2):idx+3]}")
            break
    
    print("\n" + "=" * 100)
    print("检查完成")
    print("=" * 100)

if __name__ == "__main__":
    check_first_column_data()
