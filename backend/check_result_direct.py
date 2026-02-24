"""
直接检查DELIVERY任务结果文件
"""
import openpyxl
import os

# 直接使用文件路径
result_file = r"C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_fd68ef09\result.xlsx"

print("=" * 80)
print("检查DELIVERY结果文件: " + result_file)
print("=" * 80)

if os.path.exists(result_file):
    print("\n文件存在，开始读取...")
    
    workbook = openpyxl.load_workbook(result_file)
    worksheet = workbook.active
    
    print(f"\n工作表: {worksheet.title}")
    print(f"最大行数: {worksheet.max_row}")
    print(f"最大列数: {worksheet.max_column}")
    
    print(f"\n表头:")
    headers = []
    for cell in worksheet[1]:
        headers.append(cell.value)
        print(f"  {len(headers)}: {cell.value}")
    
    # 找到P列（記事欄2）
    p_col_idx = None
    for idx, header in enumerate(headers):
        if header == '記事欄2':
            p_col_idx = idx
            print(f"\n找到P列（記事欄2），是第{p_col_idx + 1}列")
            break
    
    if p_col_idx is not None:
        print(f"\n数据行中的P列值（前10行）:")
        for row_idx in range(2, min(12, worksheet.max_row + 1)):
            cell_value = worksheet.cell(row=row_idx, column=p_col_idx + 1).value
            print(f"  第{row_idx}行: {cell_value}")
        
        # 检查第一行数据
        first_value = worksheet.cell(row=2, column=p_col_idx + 1).value
        print(f"\n第一个数据值: '{first_value}'")
        
        if first_value == "160-0327 0890":
            print("\n[OK] P列格式正确: '160-0327 0890'")
        elif first_value == "160-03270890":
            print("\n[FAIL] P列格式错误: '160-03270890' (未格式化)")
            print("[INFO] 应该是: '160-0327 0890'")
            print("[INFO] 问题: 格式化逻辑可能没有执行")
        else:
            print(f"\n[INFO] P列值为: '{first_value}'")
    else:
        print("\n[FAIL] 未找到'記事欄2'列")
        print("可用的列:")
        for idx, header in enumerate(headers):
            print(f"  {idx + 1}: {header}")
else:
    print(f"\n[FAIL] 文件不存在: {result_file}")
    
print("\n" + "=" * 80)
