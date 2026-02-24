"""
测试openpyxl如何处理空字符串的不同写入方式
"""
import openpyxl
from openpyxl import Workbook

print("=" * 80)
print("测试openpyxl如何处理空字符串的不同写入方式")
print("=" * 80)

wb = Workbook()
ws = wb.active

# 方法1: 直接设置cell.value = ''
cell1 = ws.cell(row=1, column=1)
cell1.value = ''

# 方法2: 创建单元格后设置value
cell2 = ws.cell(row=1, column=2)
cell2.value = ''

# 方法3: 使用set_explicit_value (如果存在)
cell3 = ws.cell(row=1, column=3)
try:
    cell3.set_explicit_value('')
    print("方法3: set_explicit_value 可用")
except AttributeError:
    cell3.value = ''
    print("方法3: set_explicit_value 不可用")

# 方法4: 不设置value，只创建单元格
cell4 = ws.cell(row=1, column=4)
# 不设置value

# 方法5: 设置为None
cell5 = ws.cell(row=1, column=5)
cell5.value = None

wb.save("test_empty_ways.xlsx")

# 读取并验证
wb_read = openpyxl.load_workbook("test_empty_ways.xlsx")
ws_read = wb_read.active

print("\n读取结果:")
print(f"  方法1 (cell.value = ''): {repr(ws_read.cell(row=1, column=1).value)}")
print(f"  方法2 (cell.value = ''): {repr(ws_read.cell(row=1, column=2).value)}")
print(f"  方法3 (set_explicit_value): {repr(ws_read.cell(row=1, column=3).value)}")
print(f"  方法4 (不设置value): {repr(ws_read.cell(row=1, column=4).value)}")
print(f"  方法5 (cell.value = None): {repr(ws_read.cell(row=1, column=5).value)}")

wb_read.close()
os.remove("test_empty_ways.xlsx")

print("\n测试完成！")
