"""
查看processed_column_data中D列的实际数据
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# 临时修改delivery_processor.py以打印processed_column_data

import openpyxl
from datetime import datetime
import json

# 创建一个简单的测试Excel文件
test_file_path = "test_simple.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Speedy"

# 写入完整的表头
headers = [
    "お客様管理番号", "佐川問合せ番号HAWB", "配達指定日", "時間帯指定", "貨物個数",
    "お届け先人名", "お届け先住所", "お届け先電話", "お届け先郵便",
    "依頼主", "依頼主住所", "依頼主郵便番号", "依頼主電話",
    "佐川顧客コード（固定）", "記事欄2（品名）", "記事欄2", "記事欄3"
]

for col_idx, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col_idx, value=header)

# 写入测试数据
ws.cell(row=2, column=1, value=123456)  # A
ws.cell(row=2, column=2, value="SG123456789")  # B
ws.cell(row=2, column=3, value="2026-02-10")  # C
ws.cell(row=2, column=4, value=0)  # D
ws.cell(row=2, column=5, value=1)  # E

ws.cell(row=3, column=1, value=789012)  # A
ws.cell(row=3, column=2, value="SG987654321")  # B
ws.cell(row=3, column=3, value="2026-02-11")  # C
ws.cell(row=3, column=4, value=5)  # D
ws.cell(row=3, column=5, value=2)  # E

ws.cell(row=4, column=1, value=345678)  # A
ws.cell(row=4, column=2, value="SG456789123")  # B
ws.cell(row=4, column=3, value="")  # C
ws.cell(row=4, column=4, value=0)  # D
ws.cell(row=4, column=5, value=1)  # E

wb.save(test_file_path)

# 现在读取并处理
from app.services.excel_reader import read_excel_file
result = read_excel_file(test_file_path, file_type='DELIVERY', file_role='SOURCE')

print("=" * 80)
print("查看processed_column_data的内容")
print("=" * 80)

print(f"\n原始数据:")
for col in result['column_data']:
    if col.get('head') in ['D', 'C']:
        print(f"  {col.get('head')}: {col.get('data')}")

# 模拟D列处理
from app.services.field_handlers import calc_time_slot_with_delivery_date

d_data = None
c_data = None
for col in result['column_data']:
    if col.get('head') == 'D':
        d_data = col.get('data')
    elif col.get('head') == 'C':
        c_data = col.get('data')

print(f"\n模拟D列处理:")
for idx in range(len(d_data)):
    d_val = d_data[idx]
    c_val = c_data[idx] if idx < len(c_data) else None
    result = calc_time_slot_with_delivery_date(d_val, c_val)
    print(f"  第{idx+2}行: D={d_val}, C={c_val} -> {repr(result)}")

os.remove(test_file_path)
