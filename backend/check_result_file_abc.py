"""
检查结果文件内容
"""
import openpyxl

result_file = "storage/tasks/test_abc_columns_001/result.xlsx"

print("=" * 80)
print("检查结果文件内容")
print("=" * 80)

wb = openpyxl.load_workbook(result_file)
ws = wb.active

print(f"\n工作表名称: {ws.title}")
print(f"最大行数: {ws.max_row}")
print(f"最大列数: {ws.max_column}")

print("\n表头行（第2行，如果第1行是特殊行）:")
for col_idx in range(1, min(ws.max_column + 1, 18)):
    cell = ws.cell(row=1, column=col_idx)
    print(f"  列{col_idx}: {cell.value}")

print("\n数据行（第2行）:")
for col_idx in range(1, min(ws.max_column + 1, 18)):
    cell = ws.cell(row=2, column=col_idx)
    print(f"  列{col_idx} ({ws.cell(row=1, column=col_idx).value}): {cell.value}")

print("\n数据行（第3行）:")
for col_idx in range(1, min(ws.max_column + 1, 18)):
    cell = ws.cell(row=3, column=col_idx)
    print(f"  列{col_idx} ({ws.cell(row=1, column=col_idx).value}): {cell.value}")

wb.close()
