"""
测试openpyxl如何处理空字符串和'00'
"""
import openpyxl
from openpyxl import Workbook

print("=" * 80)
print("测试openpyxl如何处理空字符串和'00'")
print("=" * 80)

wb = Workbook()
ws = wb.active

# 写入测试数据
test_values = [
    ['', '00', '05', '12', '', '08', '15'],
    ['00', '05', '12', '', '08', '15']
]

for row_idx, values in enumerate(test_values, start=2):
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

wb.save("test_openpyxl_values.xlsx")

# 读取并验证
wb_read = openpyxl.load_workbook("test_openpyxl_values.xlsx")
ws_read = wb_read.active

print("\n写入的值:")
for row_idx in range(2, 4):
    row_data = []
    for col_idx in range(1, 7):
        value = ws_read.cell(row=row_idx, column=col_idx).value
        row_data.append(repr(value))
    print(f"  第{row_idx}行: {', '.join(row_data)}")

wb_read.close()
os.remove("test_openpyxl_values.xlsx")

print("\n测试完成！")
