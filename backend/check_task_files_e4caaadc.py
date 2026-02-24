"""
检查任务ID为t_e4caaadc的原始文件和结果文件
"""

from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


def check_task_files_e4caaadc():
    """检查任务ID为t_e4caaadc的原始文件和结果文件"""
    print("=" * 100)
    print("检查任务ID为t_e4caaadc的原始文件和结果文件")
    print("=" * 100)
    
    original_file_path = "C:\\Users\\harris.xie\\Documents\\trae_projects\\japan\\backend\\storage\\tasks\\t_e4caaadc\\original.xlsx"
    result_file_path = "C:\\Users\\harris.xie\\Documents\\trae_projects\\japan\\backend\\storage\\tasks\\t_e4caaadc\\result.xlsx"
    
    try:
        # 检查原始文件
        print(f"\n检查原始文件: {original_file_path}")
        workbook = load_workbook(original_file_path)
        sheet = workbook.active
        
        print(f"工作表名称: {sheet.title}")
        print(f"最大行数: {sheet.max_row}")
        print(f"最大列数: {sheet.max_column}")
        
        # 读取第一行（特殊行）
        first_row = []
        for cell in sheet[1]:
            first_row.append(cell.value)
        
        print(f"\n第一行（特殊行）数据: {first_row[:10]}")
        
        # 读取表头行
        header_row = []
        for cell in sheet[2]:
            header_row.append(cell.value)
        
        print(f"\n表头行数据: {header_row[:10]}")
        
        # 检查结果文件
        print(f"\n检查结果文件: {result_file_path}")
        workbook_result = load_workbook(result_file_path)
        sheet_result = workbook_result.active
        
        print(f"工作表名称: {sheet_result.title}")
        print(f"最大行数: {sheet_result.max_row}")
        print(f"最大列数: {sheet_result.max_column}")
        
        # 读取第一行（特殊行）
        first_row_result = []
        for cell in sheet_result[1]:
            first_row_result.append(cell.value)
        
        print(f"\n第一行（特殊行）数据: {first_row_result[:10]}")
        
        # 读取表头行
        header_row_result = []
        for cell in sheet_result[2]:
            header_row_result.append(cell.value)
        
        print(f"\n表头行数据: {header_row_result[:10]}")
        
        # 关闭工作簿
        workbook.close()
        workbook_result.close()
        
        print("\n" + "=" * 100)
        print("检查完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"\n❌ 检查失败: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    check_task_files_e4caaadc()
