"""
检查原始文件中第一列A的数据
"""

from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


def check_first_column_a():
    """检查原始文件中第一列A的数据"""
    print("=" * 100)
    print("检查原始文件中第一列A的数据")
    print("=" * 100)
    
    original_file_path = "C:\\Users\\harris.xie\\Documents\\trae_projects\\japan\\backend\\storage\\tasks\\t_0fc5b76e\\original.xlsx"
    
    try:
        # 加载工作簿
        workbook = load_workbook(original_file_path)
        sheet = workbook.active
        
        print(f"\n原始文件路径: {original_file_path}")
        print(f"工作表名称: {sheet.title}")
        
        # 读取第一行（特殊行）
        first_row = []
        for cell in sheet[1]:
            first_row.append(cell.value)
        
        print(f"\n第一行（特殊行）数据: {first_row[:10]}")
        
        # 读取表头行
        header_row = []
        for cell in sheet[2]:
            header_row.append(cell.value)
        
        print(f"\n表头行数据: {header_row[:10]}")
        
        # 读取前10行数据
        print(f"\n前10行数据:")
        for row_idx in range(3, 13):
            row = []
            for cell in sheet[row_idx]:
                row.append(cell.value)
            print(f"行{row_idx}: {row[:10]}")
        
        # 读取第一列A的前10个数据
        print(f"\n第一列A的前10个数据:")
        for row_idx in range(3, 13):
            cell_value = sheet.cell(row=row_idx, column=1).value
            print(f"行{row_idx}: {cell_value}")
        
        # 关闭工作簿
        workbook.close()
        
        print("\n" + "=" * 100)
        print("检查完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"\n❌ 检查失败: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    check_first_column_a()
