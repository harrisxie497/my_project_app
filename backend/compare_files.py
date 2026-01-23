#!/usr/bin/env python3
"""
比较生成的结果文件与成品文件之间的差异
"""

import os
import sys
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

def compare_excel_files(file1_path: str, file2_path: str) -> dict:
    """
    比较两个Excel文件的差异
    
    Args:
        file1_path: 第一个文件路径（生成的结果文件）
        file2_path: 第二个文件路径（成品文件）
        
    Returns:
        差异报告字典
    """
    print(f"\n=== 比较文件 ===")
    print(f"文件1: {file1_path}")
    print(f"文件2: {file2_path}")
    
    # 加载工作簿
    wb1 = load_workbook(file1_path)
    wb2 = load_workbook(file2_path)
    
    # 获取活动工作表
    ws1 = wb1.active
    ws2 = wb2.active
    
    # 读取所有数据
    data1 = list(ws1.iter_rows(values_only=True))
    data2 = list(ws2.iter_rows(values_only=True))
    
    # 初始化差异报告
    report = {
        "file1_name": os.path.basename(file1_path),
        "file2_name": os.path.basename(file2_path),
        "file1_rows": len(data1),
        "file2_rows": len(data2),
        "file1_cols": len(data1[0]) if data1 else 0,
        "file2_cols": len(data2[0]) if data2 else 0,
        "header_diff": [],
        "data_diff": [],
        "total_diff_count": 0
    }
    
    # 比较表头
    print(f"\n=== 比较表头 ===")
    if data1 and data2:
        header1 = list(data1[0])
        header2 = list(data2[0])
        
        print(f"文件1表头: {header1}")
        print(f"文件2表头: {header2}")
        
        # 找出差异
        max_cols = max(len(header1), len(header2))
        for i in range(max_cols):
            val1 = header1[i] if i < len(header1) else None
            val2 = header2[i] if i < len(header2) else None
            if val1 != val2:
                diff = {
                    "row": 0,
                    "col": i,
                    "file1_val": val1,
                    "file2_val": val2
                }
                report["header_diff"].append(diff)
                report["total_diff_count"] += 1
        
        if report["header_diff"]:
            print(f"发现表头差异: {len(report['header_diff'])} 处")
            for diff in report["header_diff"]:
                print(f"  第{diff['row']+1}行, 第{diff['col']+1}列: 文件1='{diff['file1_val']}', 文件2='{diff['file2_val']}'")
        else:
            print("✅ 表头完全一致")
    
    # 比较数据
    print(f"\n=== 比较数据 ===")
    max_rows = max(len(data1), len(data2))
    
    for i in range(1, max_rows):  # 从第二行开始比较数据
        if i < len(data1) and i < len(data2):
            row1 = list(data1[i])
            row2 = list(data2[i])
            
            max_cols = max(len(row1), len(row2))
            row_has_diff = False
            
            for j in range(max_cols):
                val1 = row1[j] if j < len(row1) else None
                val2 = row2[j] if j < len(row2) else None
                
                if val1 != val2:
                    diff = {
                        "row": i,
                        "col": j,
                        "file1_val": val1,
                        "file2_val": val2
                    }
                    report["data_diff"].append(diff)
                    report["total_diff_count"] += 1
                    row_has_diff = True
            
            if row_has_diff:
                print(f"发现第{i+1}行有差异")
        elif i < len(data1):
            # 文件1有更多行
            diff = {
                "row": i,
                "col": 0,
                "file1_val": f"[整行数据]",
                "file2_val": f"[无数据]"
            }
            report["data_diff"].append(diff)
            report["total_diff_count"] += 1
            print(f"文件1第{i+1}行有数据，文件2无数据")
        elif i < len(data2):
            # 文件2有更多行
            diff = {
                "row": i,
                "col": 0,
                "file1_val": f"[无数据]",
                "file2_val": f"[整行数据]"
            }
            report["data_diff"].append(diff)
            report["total_diff_count"] += 1
            print(f"文件2第{i+1}行有数据，文件1无数据")
    
    # 输出差异统计
    print(f"\n=== 差异统计 ===")
    print(f"文件1行数: {report['file1_rows']}")
    print(f"文件2行数: {report['file2_rows']}")
    print(f"文件1列数: {report['file1_cols']}")
    print(f"文件2列数: {report['file2_cols']}")
    print(f"表头差异: {len(report['header_diff'])} 处")
    print(f"数据差异: {len(report['data_diff'])} 处")
    print(f"总差异数: {report['total_diff_count']} 处")
    
    if report['total_diff_count'] == 0:
        print("\n✅ 两个文件完全一致！")
    else:
        print(f"\n❌ 发现 {report['total_diff_count']} 处差异")
    
    return report

def main():
    """
    主函数
    """
    # 文件路径
    generated_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_results", "result.xlsx")
    finished_file = "c:\\Users\\harris.xie\\Documents\\trae_projects\\japan\\docs\\派送文件成品.xlsx"
    
    # 比较文件
    compare_excel_files(generated_file, finished_file)

if __name__ == "__main__":
    main()
