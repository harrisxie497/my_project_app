from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form, BackgroundTasks
from sqlalchemy.orm import Session
from uuid import uuid4
import os
from datetime import datetime
from typing import Optional, List
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from app.core.auth import get_current_active_user
from app.core.config import settings
from app.services.task_executor import TaskExecutor
from app.core.database import get_db
from app.core.logging_config import get_logger
from app.models.user import User
from app.models.task import Task, TaskStatus, FileType
from app.models.file_definition import FileDefinition
from app.schemas.task import (
    TaskCreate, TaskRun, TaskListResponse, TaskDetailResponse,
    TaskCreateResponse, PaginatedTaskList
)
from app.schemas.response import GeneralResponse

router = APIRouter(redirect_slashes=False)
logger = get_logger(__name__)

def execute_task_background_with_validation(task_id: str, file_type: str, db: Session) -> None:
    """
    后台异步执行任务（包含文件验证）

    :param task_id: 任务ID
    :param file_type: 文件类型
    :param db: 数据库会话
    """
    # 创建新的数据库会话（后台任务需要独立的session）
    from app.core.database import SessionLocal
    task_db = SessionLocal()

    try:
        # 获取任务
        task = task_db.query(Task).filter(Task.id == task_id).first()
        if not task:
            logger.error(f"后台任务执行失败 - 任务不存在: {task_id}")
            return

        logger.info(f"后台任务开始执行 - 任务ID: {task_id}, 文件类型: {task.file_type}")

        # 更新任务状态为处理中
        task.status = TaskStatus.PROCESSING
        task.started_at = datetime.utcnow()
        task.progress_stage = "processing"
        task.progress_message = "Task is running, please wait approximately 3 minutes"
        task_db.commit()

        # 获取任务文件存储目录
        task_dir = os.path.join(settings.TASKS_STORAGE_PATH, task_id)
        file_path = os.path.join(task_dir, "original.xlsx")

        # 验证文件是否符合配置
        try:
            validate_uploaded_file(file_path, file_type, task_db)
            logger.info(f"文件验证通过 - 任务ID: {task_id}")
        except HTTPException as e:
            # 文件验证失败，更新任务状态为失败
            task.status = TaskStatus.FAILED
            task.finished_at = datetime.utcnow()
            task.progress_stage = "failed"
            task.progress_message = f"File validation failed: {e.detail}"
            task_db.commit()
            logger.error(f"文件验证失败 - 任务ID: {task_id}, 错误: {e.detail}")
            return

        # 构建header_params
        header_params_dict = {
            'mawb_no': task.unique_code or '',
            'flight_no': task.flight_no or '',
            'arrival_date': task.declare_date or ''
        }

        # 根据文件类型选择不同的处理器
        if task.file_type == FileType.CUSTOMS:
            from app.services.task_executor import TaskExecutor
            task_executor = TaskExecutor(
                db_session=task_db,
                task_id=task_id,
                file_type='CUSTOMS',
                header_params=header_params_dict
            )
            logger.info(f"使用TaskExecutor处理清关文件 - 任务ID: {task_id}")
        elif task.file_type == FileType.DELIVERY:
            from app.services.task_executor import TaskExecutor
            task_executor = TaskExecutor(
                db_session=task_db,
                task_id=task_id,
                file_type='DELIVERY',
                header_params=header_params_dict
            )
            logger.info(f"使用TaskExecutor处理派送文件 - 任务ID: {task_id}")
        else:
            from app.services.file_processor import FileProcessor
            processor = FileProcessor(task_dir, task.file_type.value, task_db)
            logger.info(f"使用通用文件处理器 - 任务ID: {task_id}")

        # 执行文件处理
        if task.file_type == FileType.CUSTOMS:
            original_file_path = os.path.join(task_dir, "original.xlsx")
            result_file_path = os.path.join(task_dir, f"result_{task.id}.xlsx")
            stats = task_executor.execute(original_file_path, result_file_path)
        elif task.file_type == FileType.DELIVERY:
            original_file_path = os.path.join(task_dir, "original.xlsx")
            result_file_path = os.path.join(task_dir, f"result_{task.id}.xlsx")
            stats = task_executor.execute(original_file_path, result_file_path)
        else:
            stats = processor.process()

        logger.info(f"文件处理完成 - 任务ID: {task_id}, 统计信息: {stats}")

        # 更新任务的文件信息
        task.files.update({
            "result": {
                "file_name": f"result_{task.id}.xlsx",
                "download_url": f"{settings.API_V1_STR}/tasks/{task_id}/files/result"
            },
            "diff": {
                "file_name": f"diff_{task.id}.xlsx",
                "download_url": f"{settings.API_V1_STR}/tasks/{task_id}/files/diff"
            }
        })

        # 更新任务统计信息
        task.stats = stats

        # 更新任务状态为成功
        task.status = TaskStatus.SUCCESS
        task.finished_at = datetime.utcnow()
        task.progress_stage = "done"
        task.progress_message = "Task completed successfully"

        task_db.commit()

        logger.info(f"后台任务执行成功 - 任务ID: {task_id}, 最终状态: {task.status}")

    except Exception as e:
        # 获取任务并更新状态为失败
        task = task_db.query(Task).filter(Task.id == task_id).first()
        if task:
            task.status = TaskStatus.FAILED
            task.finished_at = datetime.utcnow()
            task.progress_stage = "failed"
            task.progress_message = f"Task failed: {str(e)}"
            task_db.commit()

        logger.error(f"后台任务执行失败 - 任务ID: {task_id}, 错误: {str(e)}", exc_info=True)

    finally:
        task_db.close()

def execute_task_background(task_id: str, db: Session) -> None:
    """
    后台异步执行任务

    :param task_id: 任务ID
    :param db: 数据库会话
    """
    # 创建新的数据库会话（后台任务需要独立的session）
    from app.core.database import SessionLocal
    task_db = SessionLocal()

    try:
        # 获取任务
        task = task_db.query(Task).filter(Task.id == task_id).first()
        if not task:
            logger.error(f"后台任务执行失败 - 任务不存在: {task_id}")
            return

        logger.info(f"后台任务开始执行 - 任务ID: {task_id}, 文件类型: {task.file_type}")

        # 构建header_params
        header_params_dict = {
            'mawb_no': task.unique_code or '',
            'flight_no': task.flight_no or '',
            'arrival_date': task.declare_date or ''
        }

        # 获取任务文件存储目录
        task_dir = os.path.join(settings.TASKS_STORAGE_PATH, task_id)

        # 根据文件类型选择不同的处理器
        if task.file_type == FileType.CUSTOMS:
            from app.services.task_executor import TaskExecutor
            task_executor = TaskExecutor(
                db_session=task_db,
                task_id=task_id,
                file_type='CUSTOMS',
                header_params=header_params_dict
            )
            logger.info(f"使用TaskExecutor处理清关文件 - 任务ID: {task_id}")
        elif task.file_type == FileType.DELIVERY:
            from app.services.task_executor import TaskExecutor
            task_executor = TaskExecutor(
                db_session=task_db,
                task_id=task_id,
                file_type='DELIVERY',
                header_params=header_params_dict
            )
            logger.info(f"使用TaskExecutor处理派送文件 - 任务ID: {task_id}")
        else:
            from app.services.file_processor import FileProcessor
            processor = FileProcessor(task_dir, task.file_type.value, task_db)
            logger.info(f"使用通用文件处理器 - 任务ID: {task_id}")

        # 执行文件处理
        if task.file_type == FileType.CUSTOMS:
            original_file_path = os.path.join(task_dir, "original.xlsx")
            result_file_path = os.path.join(task_dir, f"result_{task.id}.xlsx")
            stats = task_executor.execute(original_file_path, result_file_path)
        elif task.file_type == FileType.DELIVERY:
            original_file_path = os.path.join(task_dir, "original.xlsx")
            result_file_path = os.path.join(task_dir, f"result_{task.id}.xlsx")
            stats = task_executor.execute(original_file_path, result_file_path)
        else:
            stats = processor.process()

        logger.info(f"文件处理完成 - 任务ID: {task_id}, 统计信息: {stats}")

        # 更新任务的文件信息
        task.files.update({
            "result": {
                "file_name": f"result_{task.id}.xlsx",
                "download_url": f"{settings.API_V1_STR}/tasks/{task_id}/files/result"
            },
            "diff": {
                "file_name": f"diff_{task.id}.xlsx",
                "download_url": f"{settings.API_V1_STR}/tasks/{task_id}/files/diff"
            }
        })

        # 更新任务统计信息
        task.stats = stats

        # 更新任务状态为成功
        task.status = TaskStatus.SUCCESS
        task.finished_at = datetime.utcnow()
        task.progress_stage = "done"
        task.progress_message = "Task completed successfully"

        task_db.commit()

        logger.info(f"后台任务执行成功 - 任务ID: {task_id}, 最终状态: {task.status}")

    except Exception as e:
        # 获取任务并更新状态为失败
        task = task_db.query(Task).filter(Task.id == task_id).first()
        if task:
            task.status = TaskStatus.FAILED
            task.finished_at = datetime.utcnow()
            task.progress_stage = "failed"
            task.progress_message = f"Task failed: {str(e)}"
            task_db.commit()

        logger.error(f"后台任务执行失败 - 任务ID: {task_id}, 错误: {str(e)}", exc_info=True)

    finally:
        task_db.close()


def validate_uploaded_file(file_path: str, file_type: str, db: Session) -> None:
    """
    验证上传的文件是否符合file_definitions的配置
    
    输入:
        - file_path: 文件路径
        - file_type: 文件类型（customs/delivery）
        - db: 数据库会话
    
    输出:
        - None
    
    抛出:
        - HTTPException: 当文件不符合配置时抛出400错误
    """
    logger.info(f"开始验证文件 - 文件路径: {file_path}, 文件类型: {file_type}")
    logger.debug(f"验证文件输入 - 文件路径: {file_path}, 文件类型: {file_type}")
    
    # 获取文件定义配置
    file_definition = db.query(FileDefinition).filter(
        FileDefinition.file_type == file_type.upper(),
        FileDefinition.file_role == 'source',
        FileDefinition.enabled == True
    ).first()
    
    if not file_definition:
        logger.error(f"文件定义不存在 - 文件类型: {file_type}, 角色: source")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File definition not found for file type: {file_type}"
        )
    
    logger.debug(f"获取文件定义成功 - 文件类型: {file_type}, 工作表名称: {file_definition.sheet_name}")
    
    # 加载Excel文件
    workbook = load_workbook(file_path)
    
    # 验证工作表名称
    if file_definition.sheet_name not in workbook.sheetnames:
        logger.error(f"工作表名称不匹配 - 期望: {file_definition.sheet_name}, 实际: {workbook.sheetnames}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Sheet name mismatch. Expected: {file_definition.sheet_name}, Actual: {', '.join(workbook.sheetnames)}"
        )
    
    logger.debug(f"工作表名称验证通过 - 工作表名称: {file_definition.sheet_name}")
    
    # 获取工作表
    worksheet = workbook[file_definition.sheet_name]
    
    # 验证列名
    columns_json = file_definition.columns_json or []
    expected_columns = {col.get('col'): col.get('header') for col in columns_json}
    
    # 获取表头行
    header_row = file_definition.header_row or 1
    actual_headers = {}
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True), start=header_row):
        for col_idx, cell_value in enumerate(row, start=1):
            if cell_value:
                col_letter = get_column_letter(col_idx)
                actual_headers[col_letter] = str(cell_value)
    
    logger.debug(f"期望列名: {expected_columns}")
    logger.debug(f"实际列名: {actual_headers}")
    
    # 比较列名：只检查文件中存在的列是否匹配配置
    mismatched_columns = []
    for col_letter, expected_header in expected_columns.items():
        if col_letter in actual_headers:
            if actual_headers[col_letter] != expected_header:
                mismatched_columns.append(f"Column {col_letter} header mismatch. Expected: {expected_header}, Actual: {actual_headers[col_letter]}")
        else:
            mismatched_columns.append(f"Column {col_letter} not found in file")
    
    if mismatched_columns:
        logger.error(f"列名不匹配 - 错误: {mismatched_columns}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Column validation failed: {'; '.join(mismatched_columns)}"
        )
    
    logger.info(f"文件验证通过 - 文件路径: {file_path}, 文件类型: {file_type}")
    logger.debug(f"文件验证完成 - 工作表名称: {file_definition.sheet_name}, 列名匹配: {expected_columns}")

# 创建任务（上传文件）
@router.post("", response_model=GeneralResponse[TaskCreateResponse], tags=["tasks"])
async def create_task(
        background_tasks: BackgroundTasks,
        file_type: str = Form(...),
        unique_code: str = Form(...),
        file: UploadFile = File(...),
        flight_no: Optional[str] = Form(None),
        declare_date: Optional[str] = Form(None),
        header_params: Optional[str] = Form(None),
        current_user: User = Depends(get_current_active_user),
        db: Session = Depends(get_db)
    ):
    """
    创建新任务并上传文件，任务会在后台异步执行

    :param background_tasks: 后台任务
    :param file_type: 文件类型（customs/delivery）
    :param unique_code: 唯一编码
    :param file: 上传的Excel文件
    :param flight_no: 航班号（customs类型必填）
    :param declare_date: 申报日期（customs类型必填）
    :param header_params: 表头参数JSON字符串
    :param current_user: 当前登录用户
    :param db: 数据库会话
    :return: 包含任务ID和状态的响应
    :raises HTTPException: 当文件格式错误或必填字段缺失时抛出400错误
    """
    logger.info(f"创建任务请求 - 用户ID: {current_user.id}, 文件名: {file.filename}, 文件类型: {file_type}, 唯一编码: {unique_code}")
    
    # 验证文件类型
    if not file.filename.endswith(".xlsx"):
        logger.warning(f"创建任务失败 - 文件格式错误: {file.filename}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx files are allowed"
        )
    
    # 验证必填字段
    file_type_enum = FileType(file_type)
    if file_type_enum == FileType.CUSTOMS:
        if not flight_no:
            logger.warning(f"创建任务失败 - 缺少航班号: {unique_code}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="flight_no is required for customs file type"
            )
        if not declare_date:
            logger.warning(f"创建任务失败 - 缺少申报日期: {unique_code}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="declare_date is required for customs file type"
            )
    
    # 生成任务ID
    task_id = f"t_{uuid4().hex[:8]}"
    logger.info(f"生成任务ID: {task_id}")
    
    # 创建任务记录
    task = Task(
        id=task_id,
        created_by_user_id=current_user.id,
        file_type=file_type_enum,
        unique_code=unique_code,
        flight_no=flight_no,
        declare_date=declare_date,
        header_params=header_params,
        status=TaskStatus.QUEUED,
        files={}
    )
    db.add(task)
    db.commit()
    db.refresh(task)
    
    # 创建任务文件存储目录
    task_dir = os.path.join(settings.TASKS_STORAGE_PATH, task_id)
    os.makedirs(task_dir, exist_ok=True)
    
    # 保存上传的文件
    file_path = os.path.join(task_dir, "original.xlsx")
    with open(file_path, "wb") as buffer:
        buffer.write(await file.read())
    
    logger.info(f"文件保存成功 - 路径: {file_path}, 大小: {os.path.getsize(file_path)} bytes")
    
    # 更新任务文件信息
    task.files = {
        "original": {
            "file_name": file.filename,
            "download_url": f"{settings.API_V1_STR}/tasks/{task_id}/files/original"
        }
    }
    db.commit()
    
    logger.info(f"任务创建成功 - 任务ID: {task_id}, 状态: {task.status}")

    # 添加后台任务（文件验证和处理都在后台执行）
    background_tasks.add_task(execute_task_background_with_validation, task_id, file_type, db)

    return GeneralResponse(
        data=TaskCreateResponse(
            task_id=task.id,
            status=task.status,
            message="Task has been created and will be processed in the background"
        )
    )

# 运行任务
@router.post("/{task_id}/run", response_model=GeneralResponse[TaskCreateResponse], tags=["tasks"])
async def run_task(
    task_id: str,
    background_tasks: BackgroundTasks,
    header_params: Optional[str] = Form(None),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    运行任务，执行文件处理（异步执行）
    
    :param task_id: 任务ID
    :param background_tasks: 后台任务
    :param header_params: 表头参数JSON字符串（可选）
    :param current_user: 当前登录用户
    :param db: 数据库会话
    :return: 包含任务ID和状态的响应
    :raises HTTPException: 当任务不存在、状态无效时抛出错误
    """
    logger.info(f"运行任务请求 - 任务ID: {task_id}, 用户ID: {current_user.id}")
    logger.debug(f"运行任务输入 - 任务ID: {task_id}, header_params: {header_params}")
    
    # 获取任务，检查用户权限
    task = db.query(Task).filter(
        Task.id == task_id,
        Task.created_by_user_id == current_user.id
    ).first()
    if not task:
        logger.warning(f"运行任务失败 - 任务不存在: {task_id}")
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Task not found"
        )
    
    logger.debug(f"获取任务成功 - 任务ID: {task_id}, 状态: {task.status}, 文件类型: {task.file_type}")
    
    # 验证任务状态：允许成功或失败的任务重新运行
    if task.status in [TaskStatus.PROCESSING]:
        logger.warning(f"运行任务失败 - 任务正在处理中: {task_id}, 当前状态: {task.status}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Task is already processing"
        )
    
    # 解析header_params参数
    import json

    # 从tasks表中查询header_params
    header_params_dict = {
        'mawb_no': task.unique_code or '',
        'flight_no': task.flight_no or '',
        'arrival_date': task.declare_date or ''
    }
    logger.info(f"从tasks表中查询header_params: {header_params_dict}")
    
    # 如果前端传递了header_params，则合并
    if header_params:
        try:
            user_header_params = json.loads(header_params)
            logger.info(f"解析header_params成功: {user_header_params}")
            logger.debug(f"解析header_params输入 - 原始字符串: {header_params}, 解析结果: {user_header_params}")
            # 合并用户提供的header_params
            header_params_dict.update(user_header_params)
        except json.JSONDecodeError:
            logger.error(f"解析header_params失败 - 无效的JSON格式: {header_params}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid header_params JSON format"
            )
    
    # 更新任务状态为处理中
    task.status = TaskStatus.PROCESSING
    task.started_at = datetime.utcnow()
    task.progress_stage = "processing"
    task.progress_message = "Task is running, please wait approximately 3 minutes"
    db.commit()
    
    logger.info(f"任务开始处理 - 任务ID: {task_id}, 文件类型: {task.file_type}")
    logger.debug(f"任务状态更新 - 任务ID: {task_id}, 状态: {task.status}, 开始时间: {task.started_at}")
    
    # 添加后台任务（文件验证和处理都在后台执行）
    background_tasks.add_task(execute_task_background_with_validation, task_id, task.file_type.value, db)

    return GeneralResponse(
        data=TaskCreateResponse(
            task_id=task.id,
            status=task.status,
            message="Task has been started and will be processed in the background"
        )
    )

# 获取任务列表
@router.get("", response_model=GeneralResponse[PaginatedTaskList], tags=["tasks"])
async def get_tasks(
    file_type: Optional[str] = None,
    status: Optional[str] = None,
    unique_code: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    获取任务列表（支持分页和筛选）
    
    :param file_type: 文件类型筛选（可选）
    :param status: 任务状态筛选（可选）
    :param unique_code: 唯一编码模糊查询（可选）
    :param page: 页码（默认1）
    :param page_size: 每页数量（默认20）
    :param current_user: 当前登录用户
    :param db: 数据库会话
    :return: 包含任务列表和分页信息的响应
    """
    logger.info(f"获取任务列表请求 - 用户ID: {current_user.id}, 文件类型: {file_type}, 状态: {status}, 唯一编码: {unique_code}, 页码: {page}, 每页数量: {page_size}")
    
    # 构建查询
    query = db.query(Task).filter(Task.created_by_user_id == current_user.id)
    
    # 应用过滤条件
    if file_type:
        query = query.filter(Task.file_type == FileType(file_type))
    if status:
        query = query.filter(Task.status == TaskStatus(status))
    if unique_code:
        query = query.filter(Task.unique_code.contains(unique_code))
    
    # 计算总数
    total = query.count()
    
    # 应用分页
    skip = (page - 1) * page_size
    tasks = query.order_by(Task.created_at.desc()).offset(skip).limit(page_size).all()
    
    # 转换为响应模型
    task_list = [TaskListResponse.from_orm(task) for task in tasks]
    
    logger.info(f"获取任务列表成功 - 用户ID: {current_user.id}, 返回数量: {len(task_list)}, 总数: {total}")
    
    return GeneralResponse(
        data=PaginatedTaskList(
            items=task_list,
            page=page,
            page_size=page_size,
            total=total
        )
    )

# 获取任务详情
@router.get("/{task_id}", response_model=GeneralResponse[TaskDetailResponse], tags=["tasks"])
async def get_task_detail(
    task_id: str,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    获取任务详情
    
    :param task_id: 任务ID
    :param current_user: 当前登录用户
    :param db: 数据库会话
    :return: 包含任务详情的响应
    :raises HTTPException: 当任务不存在时抛出404错误
    """
    logger.info(f"获取任务详情请求 - 任务ID: {task_id}, 用户ID: {current_user.id}")
    
    # 获取任务
    task = db.query(Task).filter(
        Task.id == task_id,
        Task.created_by_user_id == current_user.id
    ).first()
    
    if not task:
        logger.warning(f"获取任务详情失败 - 任务不存在: {task_id}")
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Task not found"
        )
    
    logger.info(f"获取任务详情成功 - 任务ID: {task_id}, 状态: {task.status}")
    
    return GeneralResponse(data=TaskDetailResponse.from_orm(task))

# 下载任务文件
@router.get("/{task_id}/files/{file_kind}", tags=["tasks"])
async def download_task_file(
    task_id: str,
    file_kind: str,  # original, result, diff
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    下载任务文件
    
    :param task_id: 任务ID
    :param file_kind: 文件类型（original/result/diff）
    :param current_user: 当前登录用户
    :param db: 数据库会话
    :return: Excel文件响应
    :raises HTTPException: 当任务不存在、文件类型无效或文件不存在时抛出错误
    """
    logger.info(f"下载任务文件请求 - 任务ID: {task_id}, 文件类型: {file_kind}, 用户ID: {current_user.id}")
    
    # 获取任务
    task = db.query(Task).filter(
        Task.id == task_id,
        Task.created_by_user_id == current_user.id
    ).first()
    
    if not task:
        logger.warning(f"下载任务文件失败 - 任务不存在: {task_id}")
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Task not found"
        )
    
    # 验证文件类型
    if file_kind not in ["original", "result", "diff"]:
        logger.warning(f"下载任务文件失败 - 无效的文件类型: {file_kind}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file kind"
        )
    
    # 验证文件是否存在
    task_dir = os.path.join(settings.TASKS_STORAGE_PATH, task_id)
    file_path = os.path.join(task_dir, f"{file_kind}.xlsx")
    
    if not os.path.exists(file_path):
        logger.warning(f"下载任务文件失败 - 文件不存在: {file_path}")
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="File not found"
        )
    
    logger.info(f"下载任务文件成功 - 任务ID: {task_id}, 文件类型: {file_kind}, 文件路径: {file_path}")
    
    # 返回文件
    from fastapi.responses import FileResponse
    return FileResponse(
        path=file_path,
        filename=task.files.get(file_kind, {}).get("file_name", f"{file_kind}.xlsx"),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
