from typing import Dict, Any, List, Optional
from openpyxl import load_workbook
import logging
import pymysql

logger = logging.getLogger(__name__)

def read_excel_file(
    file_path: str, 
    sheet_name: str = None,
    file_type: str = None,
    file_role: str = None
) -> Dict[str, Any]:
    """
    读取Excel文件
    
    输入：
        - file_path: Excel文件路径
        - sheet_name: 工作表名称（可选，如果为None则读取第一个工作表）
        - file_type: 文件类型（用于查询file_definitions配置）
        - file_role: 文件角色（用于查询file_definitions配置）
    
    输出：
        - {
            "worksheet": 工作表对象,
            "first_row": 第一行数据（供后续特别处理）,
            "column_data": [
                {"head": "会员编号", "data": ["DIDA", "DIDA"], "len": 126},
                ...
            ]
        }
    """
    try:
        logger.info(f"读取Excel文件：{file_path}")
        logger.debug(f"读取Excel文件输入 - 文件路径: {file_path}, 工作表名称: {sheet_name}, 文件类型: {file_type}, 文件角色: {file_role}")
        
        # 获取file_definitions配置（在加载工作表之前）
        columns_config = []
        header_row = 1
        data_start_row = 2
        if file_type and file_role:
            logger.info(f"开始查询file_definitions - file_type: {file_type}, file_role: {file_role}")
            connection = pymysql.connect(
                host='172.18.207.224',
                port=3306,
                user='app',
                password='app123456',
                database='demo',
                charset='utf8mb4'
            )
            try:
                with connection.cursor() as cursor:
                    sql = """
                    SELECT columns_json, header_row, data_start_row, sheet_name
                    FROM file_definitions
                    WHERE file_type = %s AND file_role = %s
                    """
                    cursor.execute(sql, (file_type, file_role))
                    result = cursor.fetchone()
                    
                    logger.info(f"SQL查询结果 - result: {result}")
                    
                    if result:
                        import json
                        columns_json_str = result[0]
                        logger.info(f"columns_json原始值（类型: {type(columns_json_str)}）: {columns_json_str[:200]}")  # 只打印前200个字符
                        
                        columns_json = json.loads(columns_json_str) if isinstance(columns_json_str, str) else columns_json_str
                        columns_config = columns_json
                        logger.info(f"columns_json解析后（类型: {type(columns_json)}，数量: {len(columns_json) if isinstance(columns_json, list) else 'N/A'}")
                        
                        header_row = result[1]
                        data_start_row = result[2]
                        sheet_name_from_db = result[3]
                        
                        if sheet_name_from_db and not sheet_name:
                            sheet_name = sheet_name_from_db
                            logger.info(f"从数据库获取sheet_name: {sheet_name}")
                        logger.info(f"从file_definitions获取配置 - 列数: {len(columns_json) if isinstance(columns_json, list) else 'N/A'}, header_row: {header_row}, data_start_row: {data_start_row}")
                        logger.debug(f"file_definitions配置 - columns_config: {columns_json}")
                    else:
                        logger.warning(f"file_definitions配置不存在 - file_type: {file_type}, file_role: {file_role}")
            finally:
                connection.close()
        
        workbook = load_workbook(file_path, data_only=True)
        
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                logger.error(f"工作表不存在 - 期望: {sheet_name}, 实际: {workbook.sheetnames}")
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active
        
        logger.debug(f"选择工作表 - 工作表名称: {worksheet.title}")
        
        first_row = []
        header_row_data = []
        column_data = {}
        data_row_count = 0
        
        logger.debug(f"初始化变量 - columns_config: {len(columns_config)}, column_data: {column_data}")
        
        # 读取第一行（特殊行）
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=False), start=1):
            # 读取表头行（如果header_row == 1，则第一行就是表头行）
            if row_idx == header_row:
                header_row_data = [cell.value if cell is not None else "" for cell in row]
                logger.info(f"读取表头行（第{header_row}行） - 数据: {header_row_data}")
                logger.debug(f"表头行数据 - 列数: {len(header_row_data)}, 内容: {header_row_data}")
                
                # 构建列索引映射
                column_map = {}
                logger.debug(f"开始构建列映射 - columns_config数量: {len(columns_config)}, header_row_data数量: {len(header_row_data)}")
                
                # 对表头行数据进行规范化处理（去除空格、统一编码）
                normalized_headers = []
                logger.debug(f"开始规范化表头 - header_row_data数量: {len(header_row_data)}")
                
                for idx, header in enumerate(header_row_data):
                    logger.debug(f"  处理表头 - idx: {idx}, header: {repr(header)}, 类型: {type(header)}")
                    
                    if header:
                        # 去除首尾空格
                        normalized = header.strip()
                        # 尝试规范化日文字符
                        try:
                            # 如果是字符串，尝试编码转换
                            if isinstance(header, str):
                                normalized = normalized
                        except Exception as e:
                            logger.warning(f"表头规范化失败: {header}, 错误: {e}")
                            normalized = header.strip()
                    else:
                        normalized = ""
                    
                    normalized_headers.append(normalized)
                    logger.debug(f"  规范化后 - idx: {idx}, normalized: {repr(normalized)}")
                
                logger.debug(f"规范化后的表头 - 数量: {len(normalized_headers)}, 前10个: {normalized_headers[:10]}")  # 打印前10个
                
                for col_def in columns_config:
                    col_letter = col_def.get('col')
                    expected_header = col_def.get('header')
                    logger.debug(f"尝试匹配列 - col_letter: {col_letter}, expected_header: {repr(expected_header)}")
                    
                    # 对期望的表头也进行规范化
                    normalized_expected = expected_header.strip() if expected_header else ""
                    
                    for idx, header in enumerate(normalized_headers):
                        logger.debug(f"  检查表头 - idx: {idx}, header: {repr(header)}, 是否匹配: {header == normalized_expected}")
                        if header == normalized_expected:
                            column_map[col_letter] = idx
                            logger.debug(f"列映射成功 - {col_letter}: {expected_header} -> 索引 {idx}")
                            break
                    else:
                        logger.warning(f"列映射失败 - {col_letter}: {expected_header} 未在表头行中找到")
                        logger.debug(f"  期望表头: {repr(expected_header)}, 规范化后: {repr(normalized_expected)}")
                        logger.debug(f"  表头行数据: {normalized_headers[:10]}")  # 只打印前10个，避免日志过长
                
                logger.debug(f"列映射完成 - {column_map}")
                
                # 如果header_row不是第一行，也需要保存为first_row
                if header_row != 1:
                    first_row = header_row_data[:]
                    logger.debug(f"表头行（非第一行）也保存为特殊第一行")
                continue
            
            # 读取数据行（从data_start_row开始）
            if row_idx >= data_start_row:
                # 如果没有列配置，跳过数据处理
                if not columns_config:
                    logger.debug(f"没有列配置，跳过数据处理 - 行号: {row_idx}")
                    continue
                
                # 检查第一列A是否有值（适用于所有文件类型：DELIVERY和CUSTOMS）
                if row and len(row) > 0:
                    first_cell_value = row[0].value if row[0] is not None else None
                    if first_cell_value is not None and str(first_cell_value).strip() != "":
                        # 第一列有值，继续处理数据
                        data_row_count += 1
                        logger.debug(f"数据行 {data_row_count} - 行号: {row_idx}, 第一列值: {first_cell_value}")
                    else:
                        # 第一列为空，停止读取
                        logger.info(f"第一列为空，停止读取 - 行号: {row_idx}, 已读取数据行数: {data_row_count}")
                        break
                else:
                    # 行为空，停止读取
                    logger.info(f"行为空，停止读取 - 行号: {row_idx}, 已读取数据行数: {data_row_count}")
                    break
                
                # 按列组织数据
                for col_def in columns_config:
                    col_letter = col_def.get('col')
                    col_idx = column_map.get(col_letter)
                    
                    if col_idx is not None and col_idx < len(row):
                        cell_value = row[col_idx].value if row[col_idx] is not None else None
                        if col_letter not in column_data:
                            column_data[col_letter] = []
                        column_data[col_letter].append(cell_value)
                    #    logger.debug(f"添加单元格数据 - 列: {col_letter}, 行号: {row_idx}, 值: {cell_value}")
                        
        
        # 计算每列的数据长度，用于验证数据一致性
        column_data_list = []
        for col_def in columns_config:
            col_letter = col_def.get('col')
            col_header = col_def.get('header')
            col_values = column_data.get(col_letter, [])
            column_data_list.append({
                "source_cols": col_letter,
                "head": col_header,
                "data": col_values,
                "len": len(col_values)
            })
        
        logger.info(f"读取完成 - 第一行: {len(first_row)}, 表头行: {len(header_row_data)}, 列数: {len(column_data_list)}, 数据行数: {data_row_count}")
        logger.debug(f"读取结果 - 第一行数据: {first_row}, 表头行数据: {header_row_data}")
        logger.info(f"读取结果 - 列数据格式: {column_data_list}")
        
        return {
            "worksheet": worksheet,
            "first_row": first_row,
            "column_data": column_data_list,
            "data_row_count": data_row_count
        }
        
    except Exception as e:
        logger.error(f"读取Excel文件失败：{str(e)}", exc_info=True)
        raise
