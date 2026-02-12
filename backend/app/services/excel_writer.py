from typing import List, Any, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime
import logging
import json

logger = logging.getLogger(__name__)

def load_excel_config(file_type: str, connection) -> Dict[str, Any]:
    """
    从数据库加载Excel配置
    
    输入：
        - file_type: 文件类型（如 'CUSTOMS'）
        - connection: 数据库连接对象
    
    输出：
        - 配置字典，包含 default_font, merge_ranges, style_rules
    """
    try:
        cursor = connection.cursor()
        sql = "SELECT default_font, merge_ranges, style_rules FROM excel_configs WHERE file_type = %s AND is_active = 1"
        cursor.execute(sql, (file_type,))
        result = cursor.fetchone()
        
        if result:
            config = {
                'default_font': result[0],
                'merge_ranges': json.loads(result[1]) if result[1] else [],
                'style_rules': json.loads(result[2]) if result[2] else []
            }
            logger.info(f"加载Excel配置成功 - 文件类型: {file_type}, 默认字体: {config['default_font']}")
            return config
        else:
            logger.warning(f"未找到Excel配置 - 文件类型: {file_type}")
            return {}
    except Exception as e:
        logger.error(f"加载Excel配置失败：{str(e)}", exc_info=True)
        return {}

def apply_styles_to_worksheet(worksheet, config: Dict[str, Any], num_rows: int, num_cols: int):
    """
    应用样式到工作表
    
    输入：
        - worksheet: 工作表对象
        - config: 配置字典
        - num_rows: 数据行数
        - num_cols: 数据列数
    """
    try:
        # 1. 应用默认字体
        if config.get('default_font'):
            font_name = config['default_font']
            logger.info(f"应用默认字体：{font_name}")
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.font:
                        cell.font.name = font_name
                    else:
                        cell.font = Font(name=font_name)
        
        # 2. 应用合并单元格
        for merge_rule in config.get('merge_ranges', []):
            range_str = merge_rule.get('range', '')
            if range_str:
                try:
                    worksheet.merge_cells(range_str)
                    logger.info(f"合并单元格：{range_str} - {merge_rule.get('description', '')}")
                except Exception as e:
                    logger.warning(f"合并单元格失败：{range_str} - {str(e)}")
        
        # 3. 应用样式规则
        for style_rule in config.get('style_rules', []):
            columns = style_rule.get('columns', [])
            fill_config = style_rule.get('fill', {})
            
            # 解析填充样式
            if fill_config:
                pattern_type = fill_config.get('patternType', 'solid')
                fg_color = fill_config.get('fgColor', 'FFFFFF')
                fill = PatternFill(patternType=pattern_type, fgColor=fg_color)
                
                # 应用到指定列的特定单元格（如B1、E1、H1）
                for col_ref in columns:
                    try:
                        # 解析列引用（如 'B1' -> 列2, 行1）
                        col_letter = col_ref[0]  # 取第一个字符作为列字母
                        col_idx = ord(col_letter.upper()) - ord('A') + 1
                        row_idx = int(col_ref[1:])  # 取剩余部分作为行号
                        
                        # 应用到指定单元格
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = fill
                        
                        logger.info(f"应用填充样式 - 单元格: {col_ref}, 颜色: {fg_color}")
                    except Exception as e:
                        logger.warning(f"应用填充样式失败：{col_ref} - {str(e)}")
        
        # 4. 应用表格边框
        logger.info("应用表格边框")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 应用到数据区域（从第2行开始，第1行是特殊第一行）
        for row_idx in range(2, num_rows + 3):  # +3 包含表头行和数据行
            for col_idx in range(1, num_cols + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
        
        logger.info("样式应用完成")
        
    except Exception as e:
        logger.error(f"应用样式失败：{str(e)}", exc_info=True)

def format_as_string(value: Any) -> str:
    """
    将值格式化为字符串
    
    输入：
        - value: 任意值
    
    输出：
        - 字符串值
    """
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return str(value)

def format_date_as_yyyy_mm_dd(value: Any) -> str:
    """
    将日期值格式化为YYYY-MM-DD格式的字符串
    
    输入：
        - value: 日期值（datetime对象、字符串等）
    
    输出：
        - YYYY-MM-DD格式的字符串
    """
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, str):
        # 如果已经是字符串，尝试解析后重新格式化
        try:
            # 尝试解析常见日期格式
            for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y年%m月%d日', '%d/%m/%Y', '%m/%d/%Y']:
                try:
                    parsed_date = datetime.strptime(value, fmt)
                    return parsed_date.strftime('%Y-%m-%d')
                except ValueError:
                    continue
        except Exception:
            pass
    # 如果无法解析，直接返回字符串
    return str(value)

def write_excel_file(
    file_path: str, 
    headers: List[str], 
    data: List[List[Any]],
    write_by_column: bool = False,
    special_first_row: List[str] = None
) -> None:
    """
    写入Excel文件
    
    输入：
        - file_path: Excel文件路径
        - headers: 表头列表
        - data: 数据列表（二维数组）
        - write_by_column: 是否按列写入（默认False）
        - special_first_row: 特殊第一行数据（可选，默认None，可以是稀疏列表）
    
    输出：
        - 无返回值，直接写入文件
    """
    try:
        logger.info(f"写入Excel文件：{file_path}")
        logger.debug(f"写入Excel文件输入 - 文件路径: {file_path}, 表头: {headers}, 数据行数: {len(data)}, 按列写入: {write_by_column}, 特殊第一行: {special_first_row}")
        
        workbook = Workbook()
        worksheet = workbook.active
        
        current_row = 1
        
        # 写入特殊第一行（如果提供）
        if special_first_row:
            # 计算最大列数
            max_cols = len(headers) if headers else 0
            if special_first_row:
                max_cols = max(max_cols, len(special_first_row))
            if data and len(data) > 0:
                max_cols = max(max_cols, max(len(row) for row in data))
            
            logger.info(f"写入特殊第一行（第{current_row}行） - 长度: {len(special_first_row)}, 最大列数: {max_cols}")
            
            for col_idx, value in enumerate(special_first_row, start=1):
                if value:  # 只写入非空值
                    worksheet.cell(row=current_row, column=col_idx, value=value)
                    logger.debug(f"写入特殊第一行单元格 - 行号: {current_row}, 列号: {col_idx}, 值: {value}")
            
            logger.debug(f"特殊第一行写入完成 - 行号: {current_row}, 列数: {len(special_first_row)}")
            current_row += 1
        
        # 写入表头
        if headers:
            for col_idx, header in enumerate(headers, start=1):
                worksheet.cell(row=current_row, column=col_idx, value=header)
            logger.info(f"写入表头（第{current_row}行）：{headers}")
            logger.debug(f"表头写入完成 - 行号: {current_row}, 列数: {len(headers)}")
            current_row += 1
        
        if write_by_column:
            logger.info("按列写入数据")
            logger.debug(f"按列写入 - 数据行数: {len(data)}, 数据列数: {len(data[0]) if data else 0}")
            
            # 按列写入，从current_row行开始
            if data and len(data) > 0:
                num_rows = len(data)
                num_cols = len(data[0]) if data[0] else 0
                
                for col_idx in range(num_cols):
                    for row_idx in range(num_rows):
                        cell_value = data[row_idx][col_idx] if col_idx < len(data[row_idx]) else None
                        if cell_value is not None:
                            worksheet.cell(row=row_idx + current_row, column=col_idx + 1, value=cell_value)
                            logger.debug(f"写入单元格 - 行号: {row_idx + current_row}, 列号: {col_idx + 1}, 值: {cell_value}")
                
                logger.info(f"按列写入完成 - 共 {num_rows} 行, {num_cols} 列")
        else:
            logger.info("按行写入数据")
            for row_idx, row in enumerate(data, start=current_row):
                worksheet.append(row)
                logger.debug(f"写入数据行 - 行号: {row_idx}, 数据: {row}")
        
        workbook.save(file_path)
        logger.info(f"写入完成，共 {len(data)} 行数据")
        logger.debug(f"文件保存成功 - 文件路径: {file_path}")
        
    except Exception as e:
        logger.error(f"写入Excel文件失败：{str(e)}", exc_info=True)
        raise

def write_excel_file_by_columns(
    file_path: str, 
    headers: List[str], 
    column_data: Dict[str, List[Any]],
    special_first_row: List[str] = None,
    file_type: str = None,
    connection = None,
    sheet_name: str = None
) -> None:
    """
    按列写入Excel文件（支持特殊第一行、表头行和数据行）
    
    输入：
        - file_path: Excel文件路径
        - headers: 表头列表
        - column_data: 按列组织的数据字典 {列名: [数据列表]}
        - special_first_row: 特殊第一行数据（可选，默认None，可以是稀疏列表）
        - file_type: 文件类型（可选，用于加载配置）
        - connection: 数据库连接对象（可选，用于加载配置）
        - sheet_name: 工作表名称（可选，默认None）
    
    输出：
        - 无返回值，直接写入文件
    """
    try:
        logger.info(f"按列写入Excel文件：{file_path}")
        logger.debug(f"按列写入Excel文件输入 - 文件路径: {file_path}, 表头: {headers}, 列数: {len(column_data)}, 特殊第一行: {special_first_row}, 文件类型: {file_type}, 工作表名称: {sheet_name}")
        
        workbook = Workbook()
        worksheet = workbook.active
        
        # 设置工作表名称
        if sheet_name:
            worksheet.title = sheet_name
            logger.info(f"设置工作表名称：{sheet_name}")
        
        current_row = 1
        num_rows = 0
        num_cols = len(headers) if headers else 0
        
        # 第1行：特殊第一行（如果提供）
        if special_first_row:
            # 计算最大列数（特殊第一行、表头和数据中的最大值）
            max_cols = len(headers) if headers else 0
            if special_first_row:
                max_cols = max(max_cols, len(special_first_row))
            
            logger.info(f"写入特殊第一行（第{current_row}行） - 长度: {len(special_first_row)}, 最大列数: {max_cols}")
            
            for col_idx, value in enumerate(special_first_row, start=1):
                if value:  # 只写入非空值
                    worksheet.cell(row=current_row, column=col_idx, value=value)
                    logger.debug(f"写入特殊第一行单元格 - 行号: {current_row}, 列号: {col_idx}, 值: {value}")
            
            logger.debug(f"特殊第一行写入完成 - 行号: {current_row}, 列数: {len(special_first_row)}")
            current_row += 1
        
        # 表头行
        if headers:
            for col_idx, header in enumerate(headers, start=1):
                worksheet.cell(row=current_row, column=col_idx, value=header)
            logger.info(f"写入表头（第{current_row}行）：{headers}")
            logger.debug(f"表头写入完成 - 行号: {current_row}, 列数: {len(headers)}")
            current_row += 1
        
        # 从current_row行开始，按列写入数据
        if column_data and len(column_data) > 0:
            num_cols = len(headers)
            num_rows = max(len(column_data.get(col, [])) for col in headers) if headers else 0
            
            logger.info(f"开始按列写入数据 - 共 {num_rows} 行, {num_cols} 列")

            for col_idx, col_name in enumerate(headers, start=1):
                if col_name in column_data:
                    for row_idx, value in enumerate(column_data[col_name], start=current_row):
                        # 根据列名进行特殊处理
                        if col_idx == 1 and col_name == 'お客様管理番号':
                            # A列：お客様管理番号，用字符串方式写入
                            formatted_value = format_as_string(value)
                        elif col_idx == 2 and col_name == '佐川問合せ番号HAWB':
                            # B列：佐川問合せ番号HAWB，用字符串方式写入
                            formatted_value = format_as_string(value)
                        elif col_idx == 3 and col_name == '配達指定日':
                            # C列：配達指定日，格式化为YYYY-MM-DD字符串
                            formatted_value = format_date_as_yyyy_mm_dd(value)
                        elif col_idx == 4 and col_name == '時間帯指定':
                            # D列：時間帯指定，特殊处理
                            # 对于D列，空值直接写入空字符串，不要转换为None
                            if value is None or (isinstance(value, str) and value.strip() == ''):
                                formatted_value = ''
                            else:
                                formatted_value = value
                        else:
                            # 其他列，统一使用format_as_string处理（将None转换为空字符串）
                            formatted_value = format_as_string(value)

                        # 写入单元格
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.value = formatted_value

                        logger.debug(f"写入单元格 - 行号: {row_idx}, 列号: {col_idx}, 列名: {col_name}, 值: {value}, 格式化值: {formatted_value}")

            logger.info(f"按列写入完成 - 共 {num_rows} 行, {num_cols} 列")
        
        # 应用Excel配置（如果提供了file_type和connection）
        if file_type and connection:
            logger.info(f"应用Excel配置 - 文件类型: {file_type}")
            config = load_excel_config(file_type, connection)
            if config:
                apply_styles_to_worksheet(worksheet, config, num_rows, num_cols)
        
        workbook.save(file_path)
        logger.info(f"文件保存成功 - 共 {num_rows} 行数据")
        logger.debug(f"文件保存成功 - 文件路径: {file_path}")
        
    except Exception as e:
        logger.error(f"按列写入Excel文件失败：{str(e)}", exc_info=True)
        raise
