from typing import Dict, List, Any, Optional
from datetime import datetime
import os
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
import logging

logger = logging.getLogger(__name__)

class BaseProcessor:
    """基础文件处理器，提供通用的Excel处理功能"""
    
    def __init__(self, task_dir: str, db_session=None, file_type=None):
        """
        初始化基础处理器
        
        Args:
            task_dir: 任务文件存储目录
            db_session: 数据库会话对象
            file_type: 文件类型
        """
        self.task_dir = task_dir
        self.original_file_path = os.path.join(task_dir, "original.xlsx")
        self.result_file_path = os.path.join(task_dir, "result.xlsx")
        self.diff_file_path = os.path.join(task_dir, "diff.xlsx")
        self.db_session = db_session
        self.file_type = file_type
    
    def process(self) -> Dict[str, Any]:
        """
        执行文件处理流程
        
        Returns:
            处理结果统计信息
        """
        try:
            logger.info(f"开始处理文件，目录：{self.task_dir}")
            
            # 1. 解析原始文件
            workbook, sheet = self._parse_original_file()
            
            # 2. 构建列索引
            column_index = self._build_column_index(sheet)
            
            # 3. MAP阶段：字段映射
            mapped_data = self._map_fields(sheet, column_index)
            
            # 4. PROCESS阶段：字段处理
            processed_data = self._process_fields(mapped_data)
            
            # 5. 生成结果文件
            self._generate_result_file(processed_data)
            
            # 6. 生成差异文件
            self._generate_diff_file(processed_data, mapped_data)
            
            # 7. 计算统计信息
            stats = self._calculate_stats(processed_data, mapped_data)
            
            logger.info(f"文件处理完成，结果：{stats}")
            return stats
            
        except Exception as e:
            logger.error(f"文件处理失败：{str(e)}", exc_info=True)
            raise
    
    def _parse_original_file(self) -> tuple[Workbook, Worksheet]:
        """
        解析原始Excel文件
        
        Returns:
            (workbook, sheet): 工作簿和工作表对象
        """
        logger.info(f"解析原始文件：{self.original_file_path}")
        workbook = load_workbook(self.original_file_path)
        sheet = workbook.active
        return workbook, sheet
    
    def _build_column_index(self, sheet: Worksheet) -> Dict[str, int]:
        """
        构建列索引，将列名映射到列索引
        
        Args:
            sheet: 工作表对象
        
        Returns:
            列索引字典 {列名: 列索引}
        """
        logger.info("构建列索引")
        column_index = {}
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        
        for idx, cell_value in enumerate(header_row):
            if cell_value:
                column_index[str(cell_value)] = idx
        
        logger.info(f"列索引：{column_index}")
        return column_index
    
    def _map_fields(self, sheet: Worksheet, column_index: Dict[str, int]) -> List[Dict[str, Any]]:
        """
        MAP阶段：字段映射，将原始字段映射到目标字段
        
        Args:
            sheet: 工作表对象
            column_index: 列索引字典
        
        Returns:
            映射后的数据列表
        """
        logger.info("开始MAP阶段：字段映射")
        mapped_data = []
        
        # 获取模板映射（由子类实现）
        template_mapping = self._get_template_mapping()
        
        # 遍历数据行（跳过表头）
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            mapped_row = {}
            
            # 遍历模板映射
            for target_field, mapping_info in template_mapping.items():
                source_field = mapping_info.get("source_field")
                default_value = mapping_info.get("default_value")
                
                # 复制原始值到目标字段
                if source_field in column_index:
                    col_idx = column_index[source_field]
                    mapped_row[target_field] = row[col_idx] if col_idx < len(row) else None
                else:
                    mapped_row[target_field] = None
                
                # 处理默认值
                if mapped_row[target_field] is None and default_value is not None:
                    mapped_row[target_field] = default_value
            
            mapped_data.append(mapped_row)
        
        logger.info(f"MAP阶段完成，处理了 {len(mapped_data)} 行数据")
        return mapped_data
    
    def _process_fields(self, mapped_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        PROCESS阶段：字段处理，应用规则进行验证、转换和计算
        
        Args:
            mapped_data: 映射后的数据列表
        
        Returns:
            处理后的数据列表
        """
        logger.info("开始PROCESS阶段：字段处理")
        processed_data = []
        
        # 获取处理规则（由子类实现）
        rules = self._get_processing_rules()
        
        # 按执行顺序排序规则
        sorted_rules = sorted(rules, key=lambda x: x.get("order_no", 0))
        
        # 遍历数据行
        for row in mapped_data:
            processed_row = row.copy()
            
            # 遍历规则
            for rule in sorted_rules:
                field_name = rule.get("field_name")
                transformation_type = rule.get("transformation_type")
                params = rule.get("params", {})
                
                # 确保字段存在
                if field_name not in processed_row:
                    processed_row[field_name] = None
                
                # 执行转换
                if transformation_type == "default_value":
                    if processed_row[field_name] is None:
                        processed_row[field_name] = params.get("value")
                elif transformation_type == "format_date":
                    if processed_row[field_name]:
                        processed_row[field_name] = self._format_date(processed_row[field_name], params.get("format"))
                elif transformation_type == "uppercase":
                    if processed_row[field_name]:
                        processed_row[field_name] = str(processed_row[field_name]).upper()
                elif transformation_type == "lowercase":
                    if processed_row[field_name]:
                        processed_row[field_name] = str(processed_row[field_name]).lower()
                elif transformation_type == "trim":
                    if processed_row[field_name]:
                        processed_row[field_name] = str(processed_row[field_name]).strip()
                elif transformation_type == "calculate":
                    processed_row[field_name] = self._calculate_value(processed_row, params)
                elif transformation_type == "validate_required":
                    if processed_row[field_name] is None or processed_row[field_name] == "":
                        processed_row[f"{field_name}_error"] = params.get("message", f"{field_name} 是必填字段")
                elif transformation_type == "ai_enhance":
                    # AI增强处理（示例实现）
                    if processed_row[field_name]:
                        processed_row[field_name] = self._ai_enhance(processed_row[field_name], params)
                elif transformation_type == "conditional_expr":
                    # 条件表达式转换
                    processed_row[field_name] = self._process_conditional_expr(processed_row, params)
            
            processed_data.append(processed_row)
        
        logger.info(f"PROCESS阶段完成，处理了 {len(processed_data)} 行数据")
        return processed_data
    
    def _generate_result_file(self, processed_data: List[Dict[str, Any]]):
        """
        生成结果文件
        
        Args:
            processed_data: 处理后的数据列表
        """
        logger.info(f"生成结果文件：{self.result_file_path}")
        
        # 创建新的工作簿和工作表
        workbook = Workbook()
        sheet = workbook.active
        
        # 写入标题行
        if processed_data:
            headers = list(processed_data[0].keys())
            sheet.append(headers)
            
            # 写入数据行
            for row in processed_data:
                data_row = [row.get(header) for header in headers]
                sheet.append(data_row)
        
        # 保存文件
        workbook.save(self.result_file_path)
        logger.info("结果文件生成完成")
    
    def _generate_diff_file(self, processed_data: List[Dict[str, Any]], mapped_data: List[Dict[str, Any]]):
        """
        生成差异文件，展示处理前后的变化
        
        Args:
            processed_data: 处理后的数据列表
            mapped_data: 映射后的数据列表
        """
        logger.info(f"生成差异文件：{self.diff_file_path}")
        
        # 创建新的工作簿和工作表
        workbook = Workbook()
        sheet = workbook.active
        
        # 写入标题行
        if processed_data and mapped_data:
            # 获取所有字段
            all_fields = list(set(list(processed_data[0].keys()) + list(mapped_data[0].keys())))
            
            # 构建差异表头
            diff_headers = []
            for field in all_fields:
                diff_headers.append(f"{field}_原始")
                diff_headers.append(f"{field}_处理后")
                diff_headers.append(f"{field}_状态")
            
            sheet.append(diff_headers)
            
            # 写入差异数据
            for mapped_row, processed_row in zip(mapped_data, processed_data):
                diff_row = []
                for field in all_fields:
                    original_value = mapped_row.get(field)
                    processed_value = processed_row.get(field)
                    
                    # 确定状态
                    if original_value == processed_value:
                        status = "未变化"
                    else:
                        status = "已更新"
                    
                    diff_row.extend([original_value, processed_value, status])
                
                sheet.append(diff_row)
        
        # 保存文件
        workbook.save(self.diff_file_path)
        logger.info("差异文件生成完成")
    
    def _calculate_stats(self, processed_data: List[Dict[str, Any]], mapped_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        计算处理结果统计信息
        
        Args:
            processed_data: 处理后的数据列表
            mapped_data: 映射后的数据列表
        
        Returns:
            统计信息字典
        """
        logger.info("计算处理结果统计信息")
        
        stats = {
            "total_rows": len(processed_data),
            "fixed_count": 0,
            "filled_count": 0,
            "fx_changed_rows": 0,
            "llm_filled_count": 0
        }
        
        # 计算变化行数
        for mapped_row, processed_row in zip(mapped_data, processed_data):
            row_changed = False
            for field in mapped_row:
                original_value = mapped_row[field]
                processed_value = processed_row[field]
                
                if original_value != processed_value:
                    row_changed = True
                    stats["fixed_count"] += 1
                    
                    # 统计AI填充数量（示例逻辑）
                    if "ai_" in field:
                        stats["llm_filled_count"] += 1
            
            if row_changed:
                stats["fx_changed_rows"] += 1
        
        # 统计填充数量
        for row in processed_data:
            for field, value in row.items():
                if value is not None and value != "":
                    stats["filled_count"] += 1
        
        return stats
    
    def _format_date(self, date_value: Any, format_str: str) -> str:
        """
        格式化日期
        
        Args:
            date_value: 原始日期值
            format_str: 目标格式字符串
        
        Returns:
            格式化后的日期字符串
        """
        if isinstance(date_value, datetime):
            return date_value.strftime(format_str)
        elif isinstance(date_value, str):
            # 尝试解析字符串日期
            try:
                date_obj = datetime.strptime(date_value, "%Y-%m-%d")
                return date_obj.strftime(format_str)
            except ValueError:
                return date_value
        return str(date_value)
    
    def _calculate_value(self, row: Dict[str, Any], params: Dict[str, Any]) -> Any:
        """
        执行计算逻辑
        
        Args:
            row: 当前行数据
            params: 计算参数
        
        Returns:
            计算结果
        """
        formula = params.get("formula")
        if not formula:
            return None
        
        # 简单的计算实现（示例）
        try:
            # 替换字段名为实际值
            for field, value in row.items():
                formula = formula.replace(f"{{{field}}}", str(value or 0))
            
            # 执行计算
            return eval(formula)
        except Exception as e:
            logger.error(f"计算失败，公式：{formula}，错误：{str(e)}")
            return None
    
    def _ai_enhance(self, value: Any, params: Dict[str, Any]) -> Any:
        """
        AI增强处理（示例实现）
        
        Args:
            value: 原始值
            params: AI增强参数
        
        Returns:
            AI增强后的值
        """
        # 这里可以调用实际的AI服务
        # 示例实现：简单返回原始值
        logger.info(f"AI增强处理：{value}")
        return value
    
    def _process_conditional_expr(self, row: Dict[str, Any], params: Dict[str, Any]) -> Any:
        """
        处理条件表达式转换
        
        Args:
            row: 当前行数据
            params: 条件表达式参数
        
        Returns:
            转换后的值
        """
        logger.info(f"处理条件表达式：{params}")
        logger.info(f"当前行数据：{row}")
        
        # 获取当前字段值
        target_col = params.get("target_col")
        logger.info(f"目标列：{target_col}")
        
        # 列名映射：将字母列名映射到实际的中文列名
        column_mapping = {
            "C": "配達指定日",
            "D": "時間帯指定"
        }
        
        # 获取实际的字段名
        actual_field = column_mapping.get(target_col, target_col)
        logger.info(f"实际字段名：{actual_field}")
        
        current_value = row.get(actual_field)
        logger.info(f"当前字段值：{current_value}")
        
        # 遍历规则列表
        for rule in params.get("rules", []):
            when_expr = rule.get("when")
            set_value = rule.get("set")
            logger.info(f"处理规则：when={when_expr}, set={set_value}")
            
            # 评估条件表达式
            if self._evaluate_condition(when_expr, row):
                logger.info(f"条件表达式匹配：{when_expr}，设置值：{set_value}")
                
                # 处理设置值
                if set_value == "KEEP":
                    return current_value
                return set_value
            else:
                logger.info(f"条件表达式不匹配：{when_expr}")
        
        # 所有规则都不匹配，处理else分支
        else_value = params.get("else", "KEEP")
        logger.info(f"所有规则都不匹配，处理else分支：{else_value}")
        if else_value == "KEEP":
            return current_value
        return else_value
    
    def _evaluate_condition(self, condition: str, row: Dict[str, Any]) -> bool:
        """
        评估条件表达式
        
        Args:
            condition: 条件表达式字符串
            row: 当前行数据
        
        Returns:
            条件是否成立
        """
        logger.info(f"评估条件：{condition}")
        logger.info(f"当前行数据：{row}")
        
        # 移除前后空格
        condition = condition.strip()
        
        # 列名映射：将字母列名映射到实际的中文列名
        column_mapping = {
            "C": "配達指定日",
            "D": "時間帯指定"
        }
        
        # 处理特定格式的条件表达式
        # 格式1: C != '' && (D == 0 || D == '0')
        # 格式2: C == '' && (D == 0 || D == '0')
        
        # 直接处理已知的条件格式
        if condition == "C != '' && (D == 0 || D == '0')":
            # 检查C列（配達指定日）是否不为空，且D列（時間帯指定）为0或'0'
            c_value = row.get(column_mapping["C"], "")
            d_value = row.get(column_mapping["D"])
            logger.info(f"条件1：C='{c_value}', D={d_value}")
            return c_value != "" and (d_value == 0 or d_value == "0")
        elif condition == "C == '' && (D == 0 || D == '0')":
            # 检查C列（配達指定日）是否为空，且D列（時間帯指定）为0或'0'
            c_value = row.get(column_mapping["C"], "")
            d_value = row.get(column_mapping["D"])
            logger.info(f"条件2：C='{c_value}', D={d_value}")
            return c_value == "" and (d_value == 0 or d_value == "0")
        
        # 处理其他条件格式（预留）
        logger.warning(f"未处理的条件格式：{condition}")
        return False
    
    def _compare_operands(self, left: str, right: str, row: Dict[str, Any]) -> bool:
        """
        比较两个操作数
        
        Args:
            left: 左操作数
            right: 右操作数
            row: 当前行数据
        
        Returns:
            比较结果
        """
        # 获取左操作数的值
        left_value = self._get_operand_value(left, row)
        # 获取右操作数的值
        right_value = self._get_operand_value(right, row)
        
        logger.info(f"比较操作数：{left} = {left_value}, {right} = {right_value}")
        return left_value == right_value
    
    def _get_operand_value(self, operand: str, row: Dict[str, Any]) -> Any:
        """
        获取操作数的值
        
        Args:
            operand: 操作数字符串
            row: 当前行数据
        
        Returns:
            操作数的值
        """
        # 去除引号
        if (operand.startswith("'") and operand.endswith("'") or 
            operand.startswith('"') and operand.endswith('"')):
            return operand[1:-1]
        
        # 检查是否为数字
        if operand.isdigit():
            return int(operand)
        
        # 检查是否为空字符串
        if operand == "''":
            return ""
        
        # 否则视为列名
        value = row.get(operand)
        
        # 处理None值
        if value is None:
            return ""
        
        return value
    
    def _get_template_mapping(self) -> Dict[str, Dict[str, Any]]:
        """
        获取模板映射（由子类实现）
        
        Returns:
            模板映射字典
        """
        raise NotImplementedError("子类必须实现_get_template_mapping方法")
    
    def _get_processing_rules(self) -> List[Dict[str, Any]]:
        """
        获取处理规则（由子类实现）
        
        Returns:
            处理规则列表
        """
        raise NotImplementedError("子类必须实现_get_processing_rules方法")
