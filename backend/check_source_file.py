"""
检查源文件结构
"""
import openpyxl
import os

source_file = os.path.join(
    os.path.dirname(__file__),
    "test_results",
    "delivery_original.xlsx"
)

print("=" * 80)
print("检查源文件结构")
print("=" * 80)

wb = openpyxl.load_workbook(source_file)
sheet = wb.active

print(f"\n工作表名称: {sheet.title}")
print(f"总行数: {sheet.max_row}")
print(f"总列数: {sheet.max_column}")

print("\n【前5行数据】")
for row_idx in range(1, min(6, sheet.max_row + 1)):
    print(f"\n第{row_idx}行:")
    for col in range(1, min(18, sheet.max_column + 1)):
        cell_value = sheet.cell(row=row_idx, column=col).value
        col_letter = openpyxl.utils.get_column_letter(col)
        if cell_value is not None:
            print(f"  {col_letter}: {cell_value}")

print("\n" + "=" * 80)
