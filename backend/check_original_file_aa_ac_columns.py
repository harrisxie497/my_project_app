"""
检查原始文件中AA列和AC列的数据
"""

import openpyxl
from openpyxl import load_workbook

def check_original_file_aa_ac_columns():
    """检查原始文件中AA列和AC列的数据"""
    print("=" * 100)
    print("检查原始文件中AA列和AC列的数据")
    print("=" * 100)
    
    # 原始文件路径
    original_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\original.xlsx'
    
    # 加载原始文件
    print(f"\n加载原始文件: {original_file_path}")
    workbook = load_workbook(original_file_path, data_only=True)
    
    # 获取工作表
    worksheet = workbook.active
    print(f"工作表名称: {worksheet.title}")
    print(f"总行数: {worksheet.max_row}")
    print(f"总列数: {worksheet.max_column}")
    
    # 检查AA列和AC列的数据
    # AA列是第27列，AC列是第29列
    aa_col_index = 27
    ac_col_index = 29
    
    print("\nAA列（第27列）的数据（前5行）:")
    for row_idx in range(1, min(6, worksheet.max_row + 1)):
        cell_value = worksheet.cell(row=row_idx, column=aa_col_index).value
        print(f"  行{row_idx}: {cell_value}")
    
    print("\nAC列（第29列）的数据（前5行）:")
    for row_idx in range(1, min(6, worksheet.max_row + 1)):
        cell_value = worksheet.cell(row=row_idx, column=ac_col_index).value
        print(f"  行{row_idx}: {cell_value}")
    
    workbook.close()
    
    print("\n" + "=" * 100)
    print("检查完成！")
    print("=" * 100)

if __name__ == "__main__":
    check_original_file_aa_ac_columns()
