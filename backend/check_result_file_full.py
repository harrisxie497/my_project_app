import openpyxl
import os

def check_result_file_full():
    """完整检查任务t_aa9d170a的结果文件"""
    print("=" * 100)
    print("完整检查任务t_aa9d170a的结果文件")
    print("=" * 100)
    
    try:
        # 结果文件路径
        result_file_path = r'C:\Users\harris.xie\Documents\trae_projects\japan\backend\storage\tasks\t_aa9d170a\result.xlsx'
        
        if not os.path.exists(result_file_path):
            print(f"\n❌ 结果文件不存在: {result_file_path}")
            return
        
        print(f"\n✅ 结果文件存在: {result_file_path}")
        
        # 读取结果文件
        wb = openpyxl.load_workbook(result_file_path)
        ws = wb.active
        
        print(f"\n工作表名称: {ws.title}")
        print(f"数据行数: {ws.max_row}")
        print(f"数据列数: {ws.max_column}")
        
        # 显示表头（第一行）
        print(f"\n表头（第一行）:")
        header_row = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            header_row.append(str(cell_value) if cell_value else '')
            print(f"  列{col_idx}: {cell_value}")
        
        # 显示前10行数据
        print(f"\n前10行数据:")
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(str(cell_value) if cell_value else '')
            print(f"  行{row_idx}: {', '.join(row_data)}")
        
        wb.close()
        
        print("\n" + "=" * 100)
        print("检查完成！")
        print("=" * 100)
        
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_result_file_full()
