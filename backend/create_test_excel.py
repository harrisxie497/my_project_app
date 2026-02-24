from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def create_test_excel():
    """创建测试Excel文件"""
    wb = Workbook()
    ws = wb.active
    
    # 修改工作表名称为"Customs"
    ws.title = "Customs"
    
    # 设置表头
    headers = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH"]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 添加测试数据（4行）
    test_data = [
        ["测试数据1", "数据1", "值1", "值2", "数据2", "重量1", "数据3", "商品1", "材料1", "日语1", "日语2", "格式1", "格式2", "数据4", "数据5", "数据6", "数据7", "数据8", "数据9", "USD", "100", "数据10", "数据11", "数据12", "数据13", "数据14", "数据15", "数据16", "数据17", "数据18", "数据19", "数据20", "平台代码", "平台名称"],
        ["测试数据2", "数据2", "值3", "值4", "数据3", "重量2", "数据4", "商品2", "材料2", "日语3", "日语4", "格式3", "格式4", "数据21", "数据22", "数据23", "数据24", "数据25", "数据26", "JPY", "200", "数据27", "数据28", "数据29", "数据30", "数据31", "数据32", "数据33", "数据34", "数据35", "平台代码2", "平台名称2"],
        ["测试数据3", "数据3", "值5", "值6", "数据4", "重量3", "数据5", "商品3", "材料3", "日语5", "日语6", "格式5", "格式6", "数据36", "数据37", "数据38", "数据39", "数据40", "数据41", "数据42", "EUR", "300", "数据43", "数据44", "数据45", "数据46", "数据47", "数据48", "数据49", "数据50", "平台代码3", "平台名称3"],
        ["测试数据4", "数据4", "值7", "值8", "数据5", "重量4", "数据6", "商品4", "材料4", "日语7", "日语8", "格式7", "格式8", "数据51", "数据52", "数据53", "数据54", "数据55", "数据56", "数据57", "GBP", "400", "数据58", "数据59", "数据60", "数据61", "数据62", "数据63", "数据64", "数据65", "平台代码4", "平台名称4"]
    ]
    
    # 添加最大行数标记
    test_data.append(["4"])
    
    # 写入数据
    for row_idx, row_data in enumerate(test_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 保存文件
    wb.save("test_data.xlsx")
    print("测试Excel文件创建成功: test_data.xlsx")

if __name__ == "__main__":
    create_test_excel()
