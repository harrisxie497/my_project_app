"""
完整的DELIVERY任务测试和验证流程 - 验证修改后的配置
"""
import sys
import os
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

from app.core.database import SessionLocal
from app.services.delivery_processor import DeliveryProcessor
from app.models.field_pipeline import FieldPipeline
import openpyxl
from datetime import datetime
import json

print("=" * 80)
print("DELIVERY任务测试 - 验证佐川顧客コード和記事欄2修改")
print("=" * 80)

# 测试任务目录
task_dir = "storage/tasks/test_delivery_jkm_001"
original_file = os.path.join(task_dir, "original.xlsx")
result_file = os.path.join(task_dir, "result.xlsx")

# 检查原始文件
if not os.path.exists(original_file):
    print(f"\n❌ 原始文件不存在: {original_file}")
    sys.exit(1)

print(f"\n✅ 测试目录: {task_dir}")
print(f"✅ 原始文件: {original_file}")

# 查询修改后的配置
print("\n" + "=" * 80)
print("步骤1: 验证配置修改")
print("=" * 80)

db = SessionLocal()

# 查询佐川顧客コード（固定）列
pipeline_customer = db.query(FieldPipeline).filter(
    FieldPipeline.target_header == '佐川顧客コード（固定）',
    FieldPipeline.file_type == 'DELIVERY'
).first()

if pipeline_customer:
    rule_params = pipeline_customer.rule_params_json
    if isinstance(rule_params, dict) and 'policy_const' in rule_params:
        customer_code = rule_params['policy_const'].get('value', '')
        print(f"\n佐川顧客コード（固定）列配置:")
        print(f"  固定值: {customer_code}")
        if customer_code == '148202040055':
            print(f"  ✅ 配置正确！")
        else:
            print(f"  ❌ 配置错误，应该是 148202040055")

# 查询記事欄2列
pipeline_kiji = db.query(FieldPipeline).filter(
    FieldPipeline.target_header == '記事欄2',
    FieldPipeline.file_type == 'DELIVERY'
).first()

if pipeline_kiji:
    print(f"\n記事欄2列配置:")
    print(f"  map_op: {pipeline_kiji.map_op}")
    print(f"  field_type: {pipeline_kiji.field_type}")
    print(f"  rule_params_json: {pipeline_kiji.rule_params_json}")
    if pipeline_kiji.map_op == 'DEFAULT' and pipeline_kiji.field_type == 'HEADER' and pipeline_kiji.rule_params_json == 'task_id':
        print(f"  ✅ 配置正确！")
    else:
        print(f"  ❌ 配置错误")

# 执行DELIVERY处理
print("\n" + "=" * 80)
print("步骤2: 执行DELIVERY处理")
print("=" * 80)

if os.path.exists(result_file):
    backup_file = result_file.replace('.xlsx', '_backup2.xlsx')
    os.rename(result_file, backup_file)
    print(f"已备份旧结果文件: {backup_file}")

try:
    # 获取task_id
    from app.models.task import Task
    task = db.query(Task).filter(Task.file_type == 'DELIVERY').first()
    task_id = task.id if task else 'test_delivery_001'

    header_params = {
        'mawb_no': 'TEST-001',
        'flight_no': 'CA123',
        'arrival_date': '2026-02-10',
        'task_id': task_id  # 添加task_id
    }

    print(f"\nHeader params: {header_params}")

    processor = DeliveryProcessor(
        task_dir=task_dir,
        db_session=db,
        file_type='DELIVERY',
        header_params=header_params
    )

    result = processor.process()

    print(f"\n✅ 处理完成")
    print(f"结果文件: {processor.result_file_path}")

    if not os.path.exists(result_file):
        print(f"❌ 结果文件不存在")
        sys.exit(1)

    file_size = os.path.getsize(result_file)
    print(f"结果文件大小: {file_size} 字节")

except Exception as e:
    print(f"\n❌ 处理失败: {str(e)}")
    import traceback
    traceback.print_exc()
    db.close()
    sys.exit(1)

# 验证结果
print("\n" + "=" * 80)
print("步骤3: 验证结果文件")
print("=" * 80)

wb_result = openpyxl.load_workbook(result_file)
ws_result = wb_result.active

# 获取表头
headers = [cell.value for cell in ws_result[1]]

# 找到佐川顧客コード（固定）列的索引
customer_col_idx = None
for i, header in enumerate(headers):
    if header == '佐川顧客コード（固定）':
        customer_col_idx = i + 1
        break

if customer_col_idx:
    # 检查第一行的值
    customer_value = ws_result.cell(row=2, column=customer_col_idx).value
    print(f"\n佐川顧客コード（固定）列（第{customer_col_idx}列）:")
    print(f"  第2行值: {customer_value}")
    if customer_value == '148202040055':
        print(f"  ✅ 固定值正确！")
    else:
        print(f"  ❌ 固定值错误，应该是 148202040055")

# 找到記事欄2列的索引
kiji_col_idx = None
for i, header in enumerate(headers):
    if header == '記事欄2':
        kiji_col_idx = i + 1
        break

if kiji_col_idx:
    # 检查第一行的值
    kiji_value = ws_result.cell(row=2, column=kiji_col_idx).value
    print(f"\n記事欄2列（第{kiji_col_idx}列）:")
    print(f"  第2行值: {kiji_value}")
    print(f"  task_id: {task_id}")
    if kiji_value == task_id:
        print(f"  ✅ task_id正确！")
    else:
        print(f"  ❌ task_id错误，应该是 {task_id}")

wb_result.close()
db.close()

print("\n" + "=" * 80)
print("测试完成")
print("=" * 80)
print(f"\n结果文件: {result_file}")
print(f"请打开Excel文件手动验证！")
