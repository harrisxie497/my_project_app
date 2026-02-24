"""
检查结果文件中的None值分布
"""
import openpyxl

result_file = "storage/tasks/test_delivery_jkm_001/result.xlsx"

print(f"检查文件: {result_file}\n")

wb = openpyxl.load_workbook(result_file)
ws = wb.active

# 获取表头
headers = [cell.value for cell in ws[1]]
print(f"表头: {headers}\n")

# 统计每一列的None值数量
none_counts = {header: 0 for header in headers}

for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    for col_idx, cell in enumerate(row):
        if cell.value is None:
            header = headers[col_idx]
            none_counts[header] += 1

print("各列None值数量:")
for header, count in none_counts.items():
    if count > 0:
        print(f"  {header}: {count} 个None值")

total_none = sum(none_counts.values())
print(f"\n总计: {total_none} 个None值")

# 显示一些None值的单元格位置
print("\n前20个None值的单元格位置:")
count = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    for col_idx, cell in enumerate(row):
        if cell.value is None and count < 20:
            print(f"  行{row_idx}, 列{col_idx+1}({headers[col_idx]}): None")
            count += 1

wb.close()
